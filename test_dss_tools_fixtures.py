"""Shared workbook / workspace helpers for DSS Tools tests."""

from __future__ import annotations

import uuid
from contextlib import contextmanager
from pathlib import Path
import unittest

from datetime import datetime

from openpyxl import Workbook


class DssToolsFixtures(unittest.TestCase):
    """Base with temp paths and minimal DSS workbooks (no test methods)."""

    @contextmanager
    def workspace_files(self, stem: str):
        source = Path.cwd() / f"{stem}_{uuid.uuid4().hex}.xlsx"
        try:
            yield source
        finally:
            if source.exists():
                source.unlink()

    @contextmanager
    def workspace_json(self, stem: str):
        path = Path.cwd() / f"{stem}_{uuid.uuid4().hex}.json"
        try:
            yield path
        finally:
            if path.exists():
                path.unlink()

    @contextmanager
    def workspace_dir(self, stem: str):
        path = Path.cwd() / f"{stem}_{uuid.uuid4().hex}"
        path.mkdir(parents=True, exist_ok=True)
        try:
            yield path
        finally:
            if path.exists():
                for child in sorted(path.rglob("*"), reverse=True):
                    if child.is_file():
                        child.unlink()
                    elif child.is_dir():
                        child.rmdir()
                path.rmdir()

    def build_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-07 R1"

        ws["AA25"] = 12345
        ws["T25"] = "Alice Smith"
        ws["AC25"] = 8
        ws["AD25"] = 1
        ws["AE25"] = 1
        ws["AC26"] = 2
        ws["AC27"] = 1

        ws["AT28"] = "Bob Jones"
        ws["AX28"] = 4
        ws["AY28"] = 4
        ws["AX29"] = 2
        ws["AZ30"] = 1

        ws2 = wb.create_sheet("2026-04-08 ")
        ws2["T31"] = "Alice Smith"
        ws2["AC31"] = 10
        ws2["AC32"] = 3

        ws3 = wb.create_sheet("Notes")
        ws3["A1"] = "ignore me"

        wb.save(path)

    def build_second_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-09 Phase 2"

        ws["T25"] = "Alice Smith"
        ws["AC25"] = 5
        ws["AC26"] = 1

        ws["AT28"] = "Charlie West"
        ws["AX28"] = 7

        wb.save(path)

    def build_revision_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-07"
        ws["T25"] = "Alice Smith"
        ws["AC25"] = 8

        rev_ws = wb.create_sheet("2026-04-07 R1")
        rev_ws["T25"] = "Alice Smith"
        rev_ws["AC25"] = 10

        base2_ws = wb.create_sheet("2026-04-08")
        base2_ws["T25"] = "Bob Jones"
        base2_ws["AC25"] = 6

        rev2_ws = wb.create_sheet("2026-04-08 rev 2")
        rev2_ws["T25"] = "Bob Jones"
        rev2_ws["AC25"] = 12

        wb.save(path)

    def build_multiweek_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        sheets = [
            ("2026-04-08", "Alice Smith", 8),
            ("2026-04-15", "Alice Smith", 9),
            ("2026-04-20", "Alice Smith", 10),
            ("2026-04-22", "Alice Smith", 11),
        ]
        for index, (sheet_name, employee, st_hours) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = sheet_name
            ws["T25"] = employee
            ws["AC25"] = st_hours

        wb.save(path)

    def build_signin_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["Q1"] = "SHIFT: Day R.00"
        ws["R1"] = "DATE:"
        ws["S1"] = datetime(2026, 6, 8)
        ws["A3"] = "NAME"
        ws["C3"] = "TIME   IN"
        ws["D3"] = "TIME OUT"
        ws["E3"] = "REG HOURS"
        ws["F3"] = "Hours OT"
        ws["J3"] = "JA TECH JOB #"
        ws["L3"] = "WORK  ORDER #"
        ws["M3"] = "OPERATION #"
        ws["N3"] = "JA TECH VEHICLE UNIT #"
        ws["O3"] = "DESCRIPTION OF WORK"

        ws["A5"] = "Lochlan Ross"
        ws["C5"] = "07:00"
        ws["D5"] = "09:00"
        ws["E5"] = 2
        ws["J5"] = "PF26005-3"
        ws["O5"] = "Transformer Testing Prep"

        ws["C6"] = "09:00"
        ws["D6"] = "17:30"
        ws["E6"] = 8
        ws["J6"] = "PF26044-4"
        ws["O6"] = "Online Battery Testing"

        ws["A7"] = "Hayden Roddis"
        ws["C7"] = "07:00"
        ws["D7"] = "17:30"
        ws["J7"] = "PF26005-3"
        ws["O7"] = "Transformer Testing Prep"

        ws["A35"] = "FOR OFFICE USE:"
        wb.save(path)

    def build_signin_weekly_workbook(self, path: Path) -> None:
        wb = Workbook()
        for index, (sheet_name, day_value) in enumerate((("2026-06-08", datetime(2026, 6, 8)), ("2026-06-09", datetime(2026, 6, 9)))):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = sheet_name
            ws["Q1"] = "SHIFT: Day R.00"
            ws["R1"] = "DATE:"
            ws["S1"] = day_value
            ws["A3"] = "NAME"
            ws["C3"] = "TIME   IN"
            ws["D3"] = "TIME OUT"
            ws["E3"] = "REG HOURS"
            ws["F3"] = "Hours OT"
            ws["J3"] = "JA TECH JOB #"
            ws["L3"] = "WORK  ORDER #"
            ws["M3"] = "OPERATION #"
            ws["N3"] = "JA TECH VEHICLE UNIT #"
            ws["O3"] = "DESCRIPTION OF WORK"
            ws["A5"] = "Lochlan Ross"
            ws["C5"] = "07:00"
            ws["D5"] = "17:30" if index == 1 else "09:00"
            ws["E5"] = 10 if index == 1 else 2
            ws["J5"] = "" if index == 1 else "PF26005-3"
            ws["A35"] = "FOR OFFICE USE:"
        wb.save(path)

    def build_signin_mismatch_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["Q1"] = "SHIFT: Day R.00"
        ws["R1"] = "DATE:"
        ws["S1"] = datetime(2026, 6, 10)
        ws["A3"] = "NAME"
        ws["C3"] = "TIME   IN"
        ws["D3"] = "TIME OUT"
        ws["E3"] = "REG HOURS"
        ws["F3"] = "Hours OT"
        ws["J3"] = "JA TECH JOB #"
        ws["L3"] = "WORK  ORDER #"
        ws["M3"] = "OPERATION #"
        ws["N3"] = "JA TECH VEHICLE UNIT #"
        ws["O3"] = "DESCRIPTION OF WORK"

        ws["A5"] = "Lochlan Ross"
        ws["C5"] = "07:00"
        ws["D5"] = "17:30"
        ws["E5"] = 10.5
        ws["J5"] = "PF26005-3"
        ws["O5"] = "Mismatch Example"

        ws["A7"] = "Hayden Roddis"
        ws["C7"] = "11:00"
        ws["D7"] = "13:00"
        ws["J7"] = "PF26005-3"
        ws["O7"] = "Fallback Lunch Example"

        ws["A35"] = "FOR OFFICE USE:"
        wb.save(path)
