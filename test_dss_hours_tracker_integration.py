"""
Integration tests: workbook IO, parse_daily_records, load_tracker_data, cache.

These are skipped unless RUN_SLOW_TESTS is set (1/true/yes/all) so default
`python -m unittest test_dss_hours_tracker` stays fast during iteration.

Full suite: RUN_SLOW_TESTS=1 python -m unittest discover -s . -p "test_dss*.py"
"""

from __future__ import annotations

import os
import unittest
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from dss_hours_tracker import (
    aggregate_daily,
    aggregate_weekly,
    build_daily_rollup,
    build_email_draft_requests,
    build_email_html,
    build_error_findings,
    build_week_totals,
    build_weekly_rollup,
    cache_file_path,
    clear_cache_files,
    compute_bytes_hash,
    compute_dss_semantic_hash,
    compute_workbook_content_hash,
    deserialize_daily_record,
    find_potential_name_typos,
    format_email_subject,
    FormattingProfile,
    load_cached_daily_records,
    load_cached_source_analysis,
    load_tracker_data,
    monday_week_start,
    OperationCancelled,
    parse_daily_records,
    purge_stale_cache,
    save_cached_daily_records,
    serialize_daily_record,
)
from test_dss_hours_tracker_fixtures import DssHoursTrackerFixtures

_SLOW = os.environ.get("RUN_SLOW_TESTS", "").strip().lower() in {"1", "true", "yes", "all"}


@unittest.skipUnless(
    _SLOW,
    "Slow integration tests skipped. Set RUN_SLOW_TESTS=1 to run (see README).",
)
class DssHoursTrackerIntegrationTests(DssHoursTrackerFixtures):
    def test_parse_daily_records_ignores_numeric_name_cells(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            records = parse_daily_records(source)
            employees = {record.employee for record in records}

            self.assertEqual(employees, {"Alice Smith", "Bob Jones"})
            self.assertNotIn("12345", employees)

    def test_aggregate_weekly_groups_monday_through_sunday(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            weekly = aggregate_weekly(parse_daily_records(source))
            alice = next(record for record in weekly if record.employee == "Alice Smith")

            self.assertEqual(monday_week_start(date(2026, 4, 7)), date(2026, 4, 6))
            self.assertEqual(alice.week_start, date(2026, 4, 6))
            self.assertEqual(alice.st, 20.0)
            self.assertEqual(alice.ot, 5.0)
            self.assertEqual(alice.dt, 1.0)

    def test_rollup_and_week_totals_keep_crew_rows_separate(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            weekly = aggregate_weekly(parse_daily_records(source))
            rollup = build_weekly_rollup(weekly)
            totals = build_week_totals(weekly)

            crew_row = next(row for row in rollup if row.row_type == "Crew Total")
            self.assertEqual(crew_row.employee, "Whole Crew")
            self.assertEqual(crew_row.st, 28.0)
            self.assertEqual(crew_row.ot, 7.0)
            self.assertEqual(crew_row.dt, 2.0)

            self.assertEqual(len(totals), 1)
            self.assertEqual(totals[0].total, 37.0)

    def test_daily_rollup_adds_whole_crew_per_day(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            daily = aggregate_daily(parse_daily_records(source))
            rollup = build_daily_rollup(daily)
            crew_rows = [row for row in rollup if row.row_type == "Crew Total"]
            self.assertEqual(len(crew_rows), 2)
            self.assertTrue(all(row.employee == "Whole Crew" for row in crew_rows))
            self.assertEqual(sum(row.total for row in crew_rows), sum(row.total for row in daily))

    def test_load_tracker_data_builds_gui_view_model(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)

            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            self.assertEqual(len(tracker_data.daily_records), 3)
            self.assertEqual(tracker_data.employee_names, ["Alice Smith", "Bob Jones"])
            self.assertEqual(len(tracker_data.weekly_summary), 2)
            self.assertEqual(len(tracker_data.weekly_rollup), 3)
            self.assertEqual(len(tracker_data.daily_summary), 3)
            self.assertEqual(len(tracker_data.daily_rollup), 5)
            self.assertEqual(len(tracker_data.week_totals), 1)
            self.assertEqual(len(tracker_data.combined_weekly_summary), 2)
            self.assertEqual(len(tracker_data.combined_daily_summary), 3)

    def test_load_tracker_data_combines_multiple_dsss(self) -> None:
        with self.workspace_files("source1") as source1, self.workspace_files("source2") as source2, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source1)
            self.build_second_source_workbook(source2)

            tracker_data = load_tracker_data([source1, source2], cache_dir=cache_dir)

            self.assertEqual(len(tracker_data.source_paths), 2)
            self.assertEqual(len(tracker_data.daily_records), 5)
            self.assertEqual(len(tracker_data.weekly_summary), 4)
            self.assertEqual(len(tracker_data.combined_weekly_summary), 3)
            self.assertEqual(len(tracker_data.daily_summary), 5)
            self.assertEqual(len(tracker_data.daily_rollup), 8)
            self.assertEqual(len(tracker_data.combined_daily_summary), 5)

            alice = next(record for record in tracker_data.combined_weekly_summary if record.employee == "Alice Smith")
            self.assertEqual(alice.st, 25.0)
            self.assertEqual(alice.ot, 6.0)
            self.assertEqual(alice.dt, 1.0)

    def test_disk_cache_load_requires_workbook_content_hash(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            content_hash = compute_workbook_content_hash(source.read_bytes())
            wrong_key = "not_the_bytes_hash"
            save_cached_daily_records(cache_dir, source, wrong_key, records)
            self.assertIsNone(load_cached_daily_records(cache_dir, source, content_hash))
            save_cached_daily_records(cache_dir, source, content_hash, records)
            loaded = load_cached_daily_records(cache_dir, source, content_hash)
            self.assertIsNotNone(loaded)
            self.assertEqual(len(loaded), len(records))

    def test_parse_daily_records_uses_revised_sheet_when_present(self) -> None:
        with self.workspace_files("revisions") as source:
            self.build_revision_source_workbook(source)

            records = parse_daily_records(source)
            by_day = {(record.work_date, record.employee): record for record in records}

            self.assertEqual(by_day[(date(2026, 4, 7), "Alice Smith")].st, 10.0)
            self.assertEqual(by_day[(date(2026, 4, 8), "Bob Jones")].st, 12.0)

    def test_parse_daily_records_starts_with_most_recent_sheet(self) -> None:
        with self.workspace_files("multiweek") as source:
            self.build_multiweek_source_workbook(source)

            records = parse_daily_records(source)

            self.assertTrue(records)
            self.assertEqual(records[0].work_date, date(2026, 4, 22))

    def test_find_potential_name_typos_reports_locations(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            typo_records = [
                record
                for record in records
                if record.employee == "Alice Smith"
            ]
            typo_records = [
                type(record)(
                    source_path=record.source_path,
                    source_file=record.source_file,
                    work_date=record.work_date,
                    source_sheet=record.source_sheet,
                    employee="Alic Smith",
                    st=record.st,
                    ot=record.ot,
                    dt=record.dt,
                    source_ranges=record.source_ranges,
                )
                for record in typo_records[:1]
            ]
            warnings = find_potential_name_typos(
                ["Alic Smith"],
                ["Alice Smith", "Bob Jones", "Alic Smith"],
                records + typo_records,
            )

            self.assertEqual(len(warnings), 1)
            warning = warnings[0]
            self.assertEqual(warning.employee, "Alic Smith")
            self.assertEqual(warning.similar_employee, "Alice Smith")
            self.assertTrue(warning.locations)
            self.assertTrue(any(typo_records[0].work_date.isoformat() in location for location in warning.locations))

    def test_compute_workbook_content_hash_ignores_metadata_only_changes(self) -> None:
        with self.workspace_files("hash_meta_1") as source1, self.workspace_files("hash_meta_2") as source2:
            wb = Workbook()
            ws = wb.active
            ws.title = "2026-04-07"
            ws["T25"] = "Alice Smith"
            ws["AC25"] = 8
            wb.properties.creator = "User One"
            wb.save(source1)

            wb.properties.creator = "User Two"
            wb.save(source2)

            self.assertEqual(
                compute_workbook_content_hash(source1.read_bytes()),
                compute_workbook_content_hash(source2.read_bytes()),
            )

    def test_compute_dss_semantic_hash_ignores_irrelevant_cells(self) -> None:
        with self.workspace_files("hash_irrelevant_1") as source1, self.workspace_files("hash_irrelevant_2") as source2:
            wb = Workbook()
            ws = wb.active
            ws.title = "2026-04-07"
            ws["T25"] = "Alice Smith"
            ws["AC25"] = 8
            ws["B2"] = "outside range"
            wb.save(source1)

            ws["B2"] = "changed outside range"
            wb.save(source2)

            self.assertEqual(
                compute_dss_semantic_hash(source1.read_bytes()),
                compute_dss_semantic_hash(source2.read_bytes()),
            )

    def test_compute_dss_semantic_hash_ignores_non_dated_sheets(self) -> None:
        with self.workspace_files("hash_notes_1") as source1, self.workspace_files("hash_notes_2") as source2:
            wb = Workbook()
            ws = wb.active
            ws.title = "2026-04-07"
            ws["T25"] = "Alice Smith"
            ws["AC25"] = 8
            notes = wb.create_sheet("Notes")
            notes["A1"] = "alpha"
            wb.save(source1)

            notes["A1"] = "beta"
            wb.save(source2)

            self.assertEqual(
                compute_dss_semantic_hash(source1.read_bytes()),
                compute_dss_semantic_hash(source2.read_bytes()),
            )

    def test_compute_dss_semantic_hash_changes_when_relevant_cells_change(self) -> None:
        with self.workspace_files("hash_relevant_1") as source1, self.workspace_files("hash_relevant_2") as source2:
            wb = Workbook()
            ws = wb.active
            ws.title = "2026-04-07"
            ws["T25"] = "Alice Smith"
            ws["AC25"] = 8
            wb.save(source1)

            ws["AC25"] = 9
            wb.save(source2)

            self.assertNotEqual(
                compute_dss_semantic_hash(source1.read_bytes()),
                compute_dss_semantic_hash(source2.read_bytes()),
            )

    def test_build_error_findings_reports_threshold_crossing_day(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            profile = FormattingProfile(name="Job A", st_threshold=15.0, ot_threshold=None, daily_st_threshold=None, max_hours_per_day=None)

            findings = build_error_findings(records, profile)

            self.assertEqual(len(findings), 1)
            finding = findings[0]
            self.assertEqual(finding.employee, "Alice Smith")
            self.assertEqual(finding.hour_type, "ST")
            self.assertEqual(finding.actual_total, 20.0)
            self.assertEqual(finding.delta, 5.0)
            self.assertEqual(finding.trigger_date, date(2026, 4, 8))

    def test_build_error_findings_reports_daily_rules(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            profile = FormattingProfile(name="Job A", st_threshold=None, ot_threshold=None, daily_st_threshold=9.0, max_hours_per_day=10.0)

            findings = build_error_findings(records, profile)

            alice_findings = [finding for finding in findings if finding.employee == "Alice Smith"]
            rule_names = {finding.hour_type for finding in alice_findings}
            self.assertIn("Daily ST", rule_names)
            self.assertIn("Daily Total", rule_names)

    def test_build_email_draft_requests_groups_by_employee_for_week(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            requests = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com"},
                date(2026, 4, 6),
            )

            self.assertEqual(len(requests), 2)
            alice = next(request for request in requests if request.employee == "Alice Smith")
            self.assertEqual(alice.email, "alice@example.com")
            self.assertEqual(len(alice.records), 2)

    def test_email_draft_requests_can_be_filtered_to_selected_employees(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            requests = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com", "Bob Jones": "bob@example.com"},
                date(2026, 4, 6),
            )
            selected = {"Bob Jones"}
            filtered = [request for request in requests if request.employee in selected]

            self.assertEqual(len(filtered), 1)
            self.assertEqual(filtered[0].employee, "Bob Jones")

    def test_email_subject_and_html_include_weekly_details(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)
            request = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com"},
                date(2026, 4, 6),
            )[0]

            subject = format_email_subject(
                "Hours for {first_name} - {week_start} to {week_end}",
                request.employee,
                request.week_start,
                request.week_end,
                request.records,
            )
            html = build_email_html(
                request.employee,
                request.week_start,
                request.week_end,
                request.records,
                "<p>Hi {first_name},</p>{hours_table}",
            )

            self.assertIn("Alice", subject)
            self.assertIn("2026-04-06", subject)
            self.assertIn("Hi Alice", html)
            self.assertIn("Source File", html)
            self.assertIn("Week Total", html)

    def test_load_tracker_data_reuses_unchanged_file_records(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            initial = load_tracker_data(source, cache_dir=cache_dir)

            with mock.patch("dss_hours_tracker.parse_daily_records_from_bytes", side_effect=AssertionError("should not reparse")):
                reloaded = load_tracker_data(source, previous_data=initial, cache_dir=cache_dir)

            self.assertEqual(len(reloaded.daily_records), len(initial.daily_records))
            self.assertEqual(reloaded.file_hashes[source], initial.file_hashes[source])

    def test_load_tracker_data_reports_overall_progress(self) -> None:
        with self.workspace_files("source1") as source1, self.workspace_files("source2") as source2, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source1)
            self.build_second_source_workbook(source2)

            progress_updates: list[float] = []
            load_tracker_data(
                [source1, source2],
                cache_dir=cache_dir,
                progress_callback=lambda fraction, _message: progress_updates.append(fraction),
            )

            self.assertTrue(progress_updates)
            self.assertGreaterEqual(progress_updates[0], 0.0)
            self.assertEqual(progress_updates[-1], 1.0)
            self.assertTrue(all(left <= right for left, right in zip(progress_updates, progress_updates[1:])))

    def test_load_tracker_data_emits_partial_recent_weeks_preview(self) -> None:
        with self.workspace_files("multiweek") as source, self.workspace_dir("cache") as cache_dir:
            self.build_multiweek_source_workbook(source)

            partial_snapshots = []
            load_tracker_data(
                source,
                cache_dir=cache_dir,
                partial_callback=lambda tracker_data, message: partial_snapshots.append((tracker_data, message)),
            )

            self.assertTrue(partial_snapshots)
            tracker_data, message = partial_snapshots[0]
            self.assertIn("recent", message.lower())
            self.assertEqual(
                [record.work_date for record in tracker_data.daily_records],
                [date(2026, 4, 15), date(2026, 4, 20), date(2026, 4, 22)],
            )

    def test_load_tracker_data_can_be_cancelled(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)

            with self.assertRaises(OperationCancelled):
                load_tracker_data(
                    source,
                    cache_dir=cache_dir,
                    should_cancel=lambda: True,
                )

    def test_cache_round_trip_for_daily_records(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())

            save_cached_daily_records(cache_dir, source, workbook_hash, records)
            loaded = load_cached_daily_records(cache_dir, source, workbook_hash)

            self.assertIsNotNone(loaded)
            self.assertEqual(len(loaded), len(records))
            self.assertEqual(serialize_daily_record(loaded[0]), serialize_daily_record(records[0]))
            self.assertEqual(deserialize_daily_record(serialize_daily_record(records[0])), records[0])

    def test_cache_round_trip_for_source_analysis(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            warnings = []
            health = []

            save_cached_daily_records(
                cache_dir,
                source,
                workbook_hash,
                records,
                parse_warnings=warnings,
                workbook_health=health,
            )
            loaded = load_cached_source_analysis(cache_dir, source, workbook_hash)

            self.assertIsNotNone(loaded)
            loaded_warnings, loaded_health = loaded
            self.assertEqual(loaded_warnings, warnings)
            self.assertEqual(loaded_health, health)

    def test_purge_stale_cache_removes_old_files(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            save_cached_daily_records(cache_dir, source, workbook_hash, records)

            cache_path = cache_file_path(cache_dir, source)
            old_timestamp = (datetime.now() - timedelta(days=8)).timestamp()
            os.utime(cache_path, (old_timestamp, old_timestamp))

            purge_stale_cache(cache_dir)
            self.assertFalse(cache_path.exists())

    def test_clear_cache_files_removes_cache_jsons(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            save_cached_daily_records(cache_dir, source, workbook_hash, records)

            deleted = clear_cache_files(cache_dir)

            self.assertEqual(deleted, 1)
            self.assertEqual(list(cache_dir.glob("*.json")), [])

    def test_load_tracker_data_uses_disk_cache_when_available(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            first = load_tracker_data(source, cache_dir=cache_dir)

            with mock.patch("dss_hours_tracker.parse_daily_records_from_bytes", side_effect=AssertionError("should use cache")):
                second = load_tracker_data(source, previous_data=None, cache_dir=cache_dir)

            self.assertEqual(len(second.daily_records), len(first.daily_records))


if __name__ == "__main__":
    unittest.main()
