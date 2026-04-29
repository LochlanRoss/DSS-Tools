from __future__ import annotations

import os
import unittest
import uuid
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from dss_hours_tracker import (
    aggregate_weekly,
    AppSettings,
    build_email_draft_requests,
    build_email_html,
    build_bug_report_html,
    build_error_findings,
    build_permission_denied_message,
    build_week_totals,
    build_weekly_rollup,
    cache_file_path,
    clear_cache_files,
    compute_bytes_hash,
    compute_workbook_content_hash,
    format_email_subject,
    default_formatting_profiles,
    deserialize_daily_record,
    extract_pf_identifier,
    FilterSelection,
    find_potential_name_typos,
    find_open_excel_workbook,
    filter_employee_names,
    FormattingProfile,
    get_app_root,
    is_alert_triggered,
    load_cached_daily_records,
    load_cached_source_analysis,
    load_tracker_data,
    load_email_templates,
    load_formatting_profiles,
    load_app_settings,
    load_employee_emails,
    load_employee_groups,
    load_table_layouts,
    monday_week_start,
    normalize_person_name,
    normalize_windows_path,
    OperationCancelled,
    parse_sheet_revision,
    parse_daily_records,
    parse_threshold_value,
    purge_stale_cache,
    remove_config_keys,
    save_employee_emails,
    save_employee_groups,
    save_cached_daily_records,
    save_formatting_profiles,
    save_email_templates,
    save_app_settings,
    serialize_daily_record,
    save_table_layout,
    select_preferred_dated_sheets,
)


class DssHoursTrackerTests(unittest.TestCase):
    @contextmanager
    def workspace_files(self, stem: str):
        source = Path.cwd() / f"{stem}_{uuid.uuid4().hex}.xlsx"
        try:
            yield source
        finally:
            if source.exists():
                source.unlink()

    @contextmanager
    def workspace_json(self, stem: str):
        path = Path.cwd() / f"{stem}_{uuid.uuid4().hex}.json"
        try:
            yield path
        finally:
            if path.exists():
                path.unlink()

    @contextmanager
    def workspace_dir(self, stem: str):
        path = Path.cwd() / f"{stem}_{uuid.uuid4().hex}"
        path.mkdir(parents=True, exist_ok=True)
        try:
            yield path
        finally:
            if path.exists():
                for child in sorted(path.rglob("*"), reverse=True):
                    if child.is_file():
                        child.unlink()
                    elif child.is_dir():
                        child.rmdir()
                path.rmdir()

    def build_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-07 R1"

        ws["AA25"] = 12345
        ws["T25"] = "Alice Smith"
        ws["AC25"] = 8
        ws["AD25"] = 1
        ws["AE25"] = 1
        ws["AC26"] = 2
        ws["AC27"] = 1

        ws["AT28"] = "Bob Jones"
        ws["AX28"] = 4
        ws["AY28"] = 4
        ws["AX29"] = 2
        ws["AZ30"] = 1

        ws2 = wb.create_sheet("2026-04-08 ")
        ws2["T31"] = "Alice Smith"
        ws2["AC31"] = 10
        ws2["AC32"] = 3

        ws3 = wb.create_sheet("Notes")
        ws3["A1"] = "ignore me"

        wb.save(path)

    def build_second_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-09 Phase 2"

        ws["T25"] = "Alice Smith"
        ws["AC25"] = 5
        ws["AC26"] = 1

        ws["AT28"] = "Charlie West"
        ws["AX28"] = 7

        wb.save(path)

    def build_revision_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026-04-07"
        ws["T25"] = "Alice Smith"
        ws["AC25"] = 8

        rev_ws = wb.create_sheet("2026-04-07 R1")
        rev_ws["T25"] = "Alice Smith"
        rev_ws["AC25"] = 10

        base2_ws = wb.create_sheet("2026-04-08")
        base2_ws["T25"] = "Bob Jones"
        base2_ws["AC25"] = 6

        rev2_ws = wb.create_sheet("2026-04-08 rev 2")
        rev2_ws["T25"] = "Bob Jones"
        rev2_ws["AC25"] = 12

        wb.save(path)

    def build_multiweek_source_workbook(self, path: Path) -> None:
        wb = Workbook()
        sheets = [
            ("2026-04-08", "Alice Smith", 8),
            ("2026-04-15", "Alice Smith", 9),
            ("2026-04-20", "Alice Smith", 10),
            ("2026-04-22", "Alice Smith", 11),
        ]
        for index, (sheet_name, employee, st_hours) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = sheet_name
            ws["T25"] = employee
            ws["AC25"] = st_hours

        wb.save(path)

    def test_parse_daily_records_ignores_numeric_name_cells(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            records = parse_daily_records(source)
            employees = {record.employee for record in records}

            self.assertEqual(employees, {"Alice Smith", "Bob Jones"})
            self.assertNotIn("12345", employees)

    def test_aggregate_weekly_groups_monday_through_sunday(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            weekly = aggregate_weekly(parse_daily_records(source))
            alice = next(record for record in weekly if record.employee == "Alice Smith")

            self.assertEqual(monday_week_start(date(2026, 4, 7)), date(2026, 4, 6))
            self.assertEqual(alice.week_start, date(2026, 4, 6))
            self.assertEqual(alice.st, 20.0)
            self.assertEqual(alice.ot, 5.0)
            self.assertEqual(alice.dt, 1.0)

    def test_rollup_and_week_totals_keep_crew_rows_separate(self) -> None:
        with self.workspace_files("sample") as source:
            self.build_source_workbook(source)

            weekly = aggregate_weekly(parse_daily_records(source))
            rollup = build_weekly_rollup(weekly)
            totals = build_week_totals(weekly)

            crew_row = next(row for row in rollup if row.row_type == "Crew Total")
            self.assertEqual(crew_row.employee, "Whole Crew")
            self.assertEqual(crew_row.st, 28.0)
            self.assertEqual(crew_row.ot, 7.0)
            self.assertEqual(crew_row.dt, 2.0)

            self.assertEqual(len(totals), 1)
            self.assertEqual(totals[0].total, 37.0)

    def test_load_tracker_data_builds_gui_view_model(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)

            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            self.assertEqual(len(tracker_data.daily_records), 3)
            self.assertEqual(tracker_data.employee_names, ["Alice Smith", "Bob Jones"])
            self.assertEqual(len(tracker_data.weekly_summary), 2)
            self.assertEqual(len(tracker_data.weekly_rollup), 3)
            self.assertEqual(len(tracker_data.week_totals), 1)
            self.assertEqual(len(tracker_data.combined_weekly_summary), 2)

    def test_load_tracker_data_combines_multiple_dsss(self) -> None:
        with self.workspace_files("source1") as source1, self.workspace_files("source2") as source2, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source1)
            self.build_second_source_workbook(source2)

            tracker_data = load_tracker_data([source1, source2], cache_dir=cache_dir)

            self.assertEqual(len(tracker_data.source_paths), 2)
            self.assertEqual(len(tracker_data.daily_records), 5)
            self.assertEqual(len(tracker_data.weekly_summary), 4)
            self.assertEqual(len(tracker_data.combined_weekly_summary), 3)

            alice = next(record for record in tracker_data.combined_weekly_summary if record.employee == "Alice Smith")
            self.assertEqual(alice.st, 25.0)
            self.assertEqual(alice.ot, 6.0)
            self.assertEqual(alice.dt, 1.0)

    def test_parse_sheet_revision_handles_common_patterns(self) -> None:
        self.assertEqual(parse_sheet_revision("2026-04-07"), 0)
        self.assertEqual(parse_sheet_revision("2026-04-07 R1"), 1)
        self.assertEqual(parse_sheet_revision("2026-04-07 r 2"), 2)
        self.assertEqual(parse_sheet_revision("2026-04-07 rev 3"), 3)
        self.assertEqual(parse_sheet_revision("2026-04-07 Revision_4"), 4)

    def test_extract_pf_identifier_prefers_pf_token(self) -> None:
        self.assertEqual(
            extract_pf_identifier("PF26024-2 Electrical Tech Maintenance Ongoing DSS.xlsx"),
            "PF26024-2",
        )
        self.assertEqual(
            extract_pf_identifier("ja tech pf26024-7 phase dss.xlsx"),
            "PF26024-7",
        )

    def test_extract_pf_identifier_falls_back_to_stem(self) -> None:
        self.assertEqual(extract_pf_identifier("Electrical Tech Maintenance.xlsx"), "Electrical Tech Maintenance")

    def test_select_preferred_dated_sheets_prefers_highest_revision(self) -> None:
        selected = select_preferred_dated_sheets(
            ["2026-04-07", "2026-04-07 R1", "2026-04-08", "2026-04-08 rev 2", "notes"]
        )
        self.assertEqual(
            selected,
            [
                (date(2026, 4, 7), "2026-04-07 R1"),
                (date(2026, 4, 8), "2026-04-08 rev 2"),
            ],
        )

    def test_parse_daily_records_uses_revised_sheet_when_present(self) -> None:
        with self.workspace_files("revisions") as source:
            self.build_revision_source_workbook(source)

            records = parse_daily_records(source)
            by_day = {(record.work_date, record.employee): record for record in records}

            self.assertEqual(by_day[(date(2026, 4, 7), "Alice Smith")].st, 10.0)
            self.assertEqual(by_day[(date(2026, 4, 8), "Bob Jones")].st, 12.0)

    def test_parse_daily_records_starts_with_most_recent_sheet(self) -> None:
        with self.workspace_files("multiweek") as source:
            self.build_multiweek_source_workbook(source)

            records = parse_daily_records(source)

            self.assertTrue(records)
            self.assertEqual(records[0].work_date, date(2026, 4, 22))

    def test_permission_message_mentions_onedrive_guidance(self) -> None:
        message = build_permission_denied_message(
            Path(r"C:\Users\Test\OneDrive - JA Tech\SharePoint - PF26024\Outage DSS.xlsx")
        )
        self.assertIn("OneDrive / SharePoint", message)
        self.assertIn("Always keep on this device", message)
        self.assertIn("Update View", message)

    def test_normalize_windows_path(self) -> None:
        self.assertEqual(
            normalize_windows_path(r"C:/Users/Test/OneDrive/File.xlsx\\"),
            r"c:\users\test\onedrive\file.xlsx",
        )

    def test_normalize_person_name(self) -> None:
        self.assertEqual(normalize_person_name("  Alice   Smith "), "alice smith")

    def test_find_potential_name_typos_reports_locations(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            typo_records = [
                record
                for record in records
                if record.employee == "Alice Smith"
            ]
            typo_records = [
                type(record)(
                    source_path=record.source_path,
                    source_file=record.source_file,
                    work_date=record.work_date,
                    source_sheet=record.source_sheet,
                    employee="Alic Smith",
                    st=record.st,
                    ot=record.ot,
                    dt=record.dt,
                    source_ranges=record.source_ranges,
                )
                for record in typo_records[:1]
            ]
            warnings = find_potential_name_typos(
                ["Alic Smith"],
                ["Alice Smith", "Bob Jones", "Alic Smith"],
                records + typo_records,
            )

            self.assertEqual(len(warnings), 1)
            warning = warnings[0]
            self.assertEqual(warning.employee, "Alic Smith")
            self.assertEqual(warning.similar_employee, "Alice Smith")
            self.assertTrue(warning.locations)
            self.assertTrue(any(typo_records[0].work_date.isoformat() in location for location in warning.locations))

    def test_find_open_excel_workbook_matches_by_full_path_or_unique_name(self) -> None:
        class FakeWorkbook:
            def __init__(self, full_name: str, name: str):
                self.FullName = full_name
                self.Name = name

        class FakeExcel:
            def __init__(self, workbooks):
                self.Workbooks = workbooks

        exact = FakeWorkbook(
            r"C:\Users\Test\OneDrive - JA Tech\SharePoint - PF26024\Outage DSS.xlsx",
            "Outage DSS.xlsx",
        )
        result = find_open_excel_workbook(FakeExcel([exact]), Path(exact.FullName))
        self.assertIs(result, exact)

        by_name = FakeWorkbook(r"C:\Other\Folder\Outage DSS.xlsx", "Outage DSS.xlsx")
        result = find_open_excel_workbook(FakeExcel([by_name]), Path(r"C:\Missing\Outage DSS.xlsx"))
        self.assertIs(result, by_name)

    def test_compute_bytes_hash_is_stable(self) -> None:
        payload = b"example workbook bytes"
        self.assertEqual(compute_bytes_hash(payload), compute_bytes_hash(payload))
        self.assertNotEqual(compute_bytes_hash(payload), compute_bytes_hash(payload + b"!"))

    def test_compute_workbook_content_hash_ignores_metadata_only_changes(self) -> None:
        with self.workspace_files("hash_meta_1") as source1, self.workspace_files("hash_meta_2") as source2:
            wb = Workbook()
            ws = wb.active
            ws.title = "2026-04-07"
            ws["T25"] = "Alice Smith"
            ws["AC25"] = 8
            wb.properties.creator = "User One"
            wb.save(source1)

            wb.properties.creator = "User Two"
            wb.save(source2)

            self.assertEqual(
                compute_workbook_content_hash(source1.read_bytes()),
                compute_workbook_content_hash(source2.read_bytes()),
            )

    def test_parse_threshold_value_allows_blank(self) -> None:
        self.assertIsNone(parse_threshold_value(""))
        self.assertEqual(parse_threshold_value("40"), 40.0)

    def test_is_alert_triggered_uses_configured_thresholds(self) -> None:
        profile = FormattingProfile(name="Job A", st_threshold=40.0, ot_threshold=10.0, daily_st_threshold=8.0, max_hours_per_day=12.0)
        self.assertFalse(is_alert_triggered(40.0, 10.0, 2.0, profile))
        self.assertTrue(is_alert_triggered(41.0, 0.0, 0.0, profile))
        self.assertTrue(is_alert_triggered(0.0, 11.0, 0.0, profile))

    def test_build_error_findings_reports_threshold_crossing_day(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            profile = FormattingProfile(name="Job A", st_threshold=15.0, ot_threshold=None, daily_st_threshold=None, max_hours_per_day=None)

            findings = build_error_findings(records, profile)

            self.assertEqual(len(findings), 1)
            finding = findings[0]
            self.assertEqual(finding.employee, "Alice Smith")
            self.assertEqual(finding.hour_type, "ST")
            self.assertEqual(finding.actual_total, 20.0)
            self.assertEqual(finding.delta, 5.0)
            self.assertEqual(finding.trigger_date, date(2026, 4, 8))

    def test_build_error_findings_reports_daily_rules(self) -> None:
        with self.workspace_files("source") as source:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            profile = FormattingProfile(name="Job A", st_threshold=None, ot_threshold=None, daily_st_threshold=9.0, max_hours_per_day=10.0)

            findings = build_error_findings(records, profile)

            alice_findings = [finding for finding in findings if finding.employee == "Alice Smith"]
            rule_names = {finding.hour_type for finding in alice_findings}
            self.assertIn("Daily ST", rule_names)
            self.assertIn("Daily Total", rule_names)

    def test_formatting_profiles_round_trip(self) -> None:
        with self.workspace_json("formatting") as path:
            profiles = default_formatting_profiles()
            profiles["Job B"] = FormattingProfile(
                name="Job B",
                st_threshold=36.0,
                ot_threshold=8.0,
                daily_st_threshold=9.0,
                max_hours_per_day=12.0,
            )
            save_formatting_profiles(path, profiles, "Job B")

            loaded_profiles, current_profile = load_formatting_profiles(path)

            self.assertEqual(current_profile, "Job B")
            self.assertEqual(loaded_profiles["Job B"].st_threshold, 36.0)
            self.assertEqual(loaded_profiles["Job B"].ot_threshold, 8.0)
            self.assertEqual(loaded_profiles["Job B"].daily_st_threshold, 9.0)
            self.assertEqual(loaded_profiles["Job B"].max_hours_per_day, 12.0)

    def test_employee_emails_round_trip(self) -> None:
        with self.workspace_json("emails") as path:
            save_employee_emails(path, {"Alice Smith": "alice@example.com", "Bob Jones": "bob@example.com"})
            loaded = load_employee_emails(path)
            self.assertEqual(loaded["Alice Smith"], "alice@example.com")
            self.assertEqual(loaded["Bob Jones"], "bob@example.com")

    def test_employee_groups_round_trip(self) -> None:
        with self.workspace_json("groups") as path:
            save_employee_groups(path, {"Crew A": ["Alice Smith", "Bob Jones"], "Crew B": ["Charlie West"]})
            loaded = load_employee_groups(path)
            self.assertEqual(loaded["Crew A"], ["Alice Smith", "Bob Jones"])
            self.assertEqual(loaded["Crew B"], ["Charlie West"])

    def test_table_layout_round_trip(self) -> None:
        with self.workspace_json("table_layout") as path:
            save_table_layout(
                path,
                "weekly_summary",
                ["week_start", "employee", "st"],
                {"week_start": 120, "employee": 240, "st": 90},
                sort_column="week_start",
                sort_descending=True,
            )

            layouts = load_table_layouts(path)

            self.assertEqual(layouts["weekly_summary"]["visible_columns"], ["week_start", "employee", "st"])
            self.assertEqual(layouts["weekly_summary"]["column_widths"]["employee"], 240)
            self.assertEqual(layouts["weekly_summary"]["sort_column"], "week_start")
            self.assertTrue(layouts["weekly_summary"]["sort_descending"])

    def test_app_settings_round_trip(self) -> None:
        with self.workspace_json("app_settings") as path:
            save_app_settings(
                path,
                AppSettings(
                    disable_name_typo_notifications=True,
                    hash_poll_minutes=12,
                    show_daily_raw_tab=False,
                ),
            )

            loaded = load_app_settings(path)

            self.assertTrue(loaded.disable_name_typo_notifications)
            self.assertEqual(loaded.hash_poll_minutes, 12)
            self.assertFalse(loaded.show_daily_raw_tab)

    def test_filter_employee_names_supports_all_employee_and_group_modes(self) -> None:
        employees = ["Alice Smith", "Bob Jones", "Charlie West"]
        groups = {"Crew A": ["Alice Smith", "Charlie West"]}
        self.assertEqual(
            filter_employee_names(employees, FilterSelection(mode="all", value="All Employees"), groups),
            {"Alice Smith", "Bob Jones", "Charlie West"},
        )
        self.assertEqual(
            filter_employee_names(employees, FilterSelection(mode="employee", value="Bob Jones"), groups),
            {"Bob Jones"},
        )
        self.assertEqual(
            filter_employee_names(employees, FilterSelection(mode="group", value="Crew A"), groups),
            {"Alice Smith", "Charlie West"},
        )
        self.assertEqual(
            filter_employee_names(
                employees,
                FilterSelection(mode="employee_multi", value="Alice Smith, Bob Jones", values=("Alice Smith", "Bob Jones")),
                groups,
            ),
            {"Alice Smith", "Bob Jones"},
        )

    def test_build_email_draft_requests_groups_by_employee_for_week(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            requests = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com"},
                date(2026, 4, 6),
            )

            self.assertEqual(len(requests), 2)
            alice = next(request for request in requests if request.employee == "Alice Smith")
            self.assertEqual(alice.email, "alice@example.com")
            self.assertEqual(len(alice.records), 2)

    def test_email_draft_requests_can_be_filtered_to_selected_employees(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)

            requests = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com", "Bob Jones": "bob@example.com"},
                date(2026, 4, 6),
            )
            selected = {"Bob Jones"}
            filtered = [request for request in requests if request.employee in selected]

            self.assertEqual(len(filtered), 1)
            self.assertEqual(filtered[0].employee, "Bob Jones")

    def test_email_subject_and_html_include_weekly_details(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            tracker_data = load_tracker_data(source, cache_dir=cache_dir)
            request = build_email_draft_requests(
                tracker_data.daily_records,
                {"Alice Smith": "alice@example.com"},
                date(2026, 4, 6),
            )[0]

            subject = format_email_subject(
                "Hours for {first_name} - {week_start} to {week_end}",
                request.employee,
                request.week_start,
                request.week_end,
            )
            html = build_email_html(
                request.employee,
                request.week_start,
                request.week_end,
                request.records,
                "<p>Hi {first_name},</p>{hours_table}",
            )

            self.assertIn("Alice", subject)
            self.assertIn("2026-04-06", subject)
            self.assertIn("Hi Alice", html)
            self.assertIn("Source File", html)
            self.assertIn("Week Total", html)

    def test_email_templates_round_trip(self) -> None:
        with self.workspace_json("email_templates") as path:
            save_email_templates(path, "Subject {first_name}", "<p>Hello {first_name}</p>{hours_table}")
            subject_template, body_template = load_email_templates(path)

            self.assertEqual(subject_template, "Subject {first_name}")
            self.assertIn("{hours_table}", body_template)

    def test_build_bug_report_html_includes_key_fields(self) -> None:
        html = build_bug_report_html(
            current_profile_name="Default",
            app_root=Path(r"C:\Temp\DSSHoursTracker"),
            snapshot_path=Path(r"C:\Temp\DSSHoursTracker\diagnostic_snapshot.json"),
            loaded_sources=[Path(r"C:\Work\Alpha.xlsx")],
            cache_status_by_path={Path(r"C:\Work\Alpha.xlsx"): "Disk Hit"},
        )

        self.assertIn("Summary", html)
        self.assertIn("Steps to reproduce", html)
        self.assertIn("diagnostic_snapshot.json", html)
        self.assertIn("Alpha.xlsx", html)
        self.assertIn("Disk Hit", html)

    def test_load_tracker_data_reuses_unchanged_file_records(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            initial = load_tracker_data(source, cache_dir=cache_dir)

            with mock.patch("dss_hours_tracker.parse_daily_records_from_bytes", side_effect=AssertionError("should not reparse")):
                reloaded = load_tracker_data(source, previous_data=initial, cache_dir=cache_dir)

            self.assertEqual(len(reloaded.daily_records), len(initial.daily_records))
            self.assertEqual(reloaded.file_hashes[source], initial.file_hashes[source])

    def test_load_tracker_data_reports_overall_progress(self) -> None:
        with self.workspace_files("source1") as source1, self.workspace_files("source2") as source2, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source1)
            self.build_second_source_workbook(source2)

            progress_updates: list[float] = []
            load_tracker_data(
                [source1, source2],
                cache_dir=cache_dir,
                progress_callback=lambda fraction, _message: progress_updates.append(fraction),
            )

            self.assertTrue(progress_updates)
            self.assertGreaterEqual(progress_updates[0], 0.0)
            self.assertEqual(progress_updates[-1], 1.0)
            self.assertTrue(all(left <= right for left, right in zip(progress_updates, progress_updates[1:])))

    def test_load_tracker_data_emits_partial_recent_weeks_preview(self) -> None:
        with self.workspace_files("multiweek") as source, self.workspace_dir("cache") as cache_dir:
            self.build_multiweek_source_workbook(source)

            partial_snapshots = []
            load_tracker_data(
                source,
                cache_dir=cache_dir,
                partial_callback=lambda tracker_data, message: partial_snapshots.append((tracker_data, message)),
            )

            self.assertTrue(partial_snapshots)
            tracker_data, message = partial_snapshots[0]
            self.assertIn("recent", message.lower())
            self.assertEqual(
                [record.work_date for record in tracker_data.daily_records],
                [date(2026, 4, 15), date(2026, 4, 20), date(2026, 4, 22)],
            )

    def test_load_tracker_data_can_be_cancelled(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)

            with self.assertRaises(OperationCancelled):
                load_tracker_data(
                    source,
                    cache_dir=cache_dir,
                    should_cancel=lambda: True,
                )

    def test_cache_round_trip_for_daily_records(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())

            save_cached_daily_records(cache_dir, source, workbook_hash, records)
            loaded = load_cached_daily_records(cache_dir, source, workbook_hash)

            self.assertIsNotNone(loaded)
            self.assertEqual(len(loaded), len(records))
            self.assertEqual(serialize_daily_record(loaded[0]), serialize_daily_record(records[0]))
            self.assertEqual(deserialize_daily_record(serialize_daily_record(records[0])), records[0])

    def test_cache_round_trip_for_source_analysis(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            warnings = []
            health = []

            save_cached_daily_records(
                cache_dir,
                source,
                workbook_hash,
                records,
                parse_warnings=warnings,
                workbook_health=health,
            )
            loaded = load_cached_source_analysis(cache_dir, source, workbook_hash)

            self.assertIsNotNone(loaded)
            loaded_warnings, loaded_health = loaded
            self.assertEqual(loaded_warnings, warnings)
            self.assertEqual(loaded_health, health)

    def test_purge_stale_cache_removes_old_files(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            save_cached_daily_records(cache_dir, source, workbook_hash, records)

            cache_path = cache_file_path(cache_dir, source)
            old_timestamp = (datetime.now() - timedelta(days=8)).timestamp()
            os.utime(cache_path, (old_timestamp, old_timestamp))

            purge_stale_cache(cache_dir)
            self.assertFalse(cache_path.exists())

    def test_clear_cache_files_removes_cache_jsons(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            records = parse_daily_records(source)
            workbook_hash = compute_bytes_hash(source.read_bytes())
            save_cached_daily_records(cache_dir, source, workbook_hash, records)

            deleted = clear_cache_files(cache_dir)

            self.assertEqual(deleted, 1)
            self.assertEqual(list(cache_dir.glob("*.json")), [])

    def test_load_tracker_data_uses_disk_cache_when_available(self) -> None:
        with self.workspace_files("source") as source, self.workspace_dir("cache") as cache_dir:
            self.build_source_workbook(source)
            first = load_tracker_data(source, cache_dir=cache_dir)

            with mock.patch("dss_hours_tracker.parse_daily_records_from_bytes", side_effect=AssertionError("should use cache")):
                second = load_tracker_data(source, previous_data=None, cache_dir=cache_dir)

            self.assertEqual(len(second.daily_records), len(first.daily_records))

    def test_get_app_root_uses_localappdata_when_available(self) -> None:
        with mock.patch.dict(os.environ, {"LOCALAPPDATA": r"C:\Temp\AppData"}, clear=False):
            app_root = get_app_root()
            self.assertTrue(str(app_root).endswith("DSSHoursTracker"))

    def test_remove_config_keys_keeps_remaining_payload(self) -> None:
        with self.workspace_json("config_keys") as path:
            path.write_text(
                '{"employee_emails": {"Alice Smith": "alice@example.com"}, "table_layouts": {"daily_raw": {}}}',
                encoding="utf-8",
            )

            remove_config_keys(path, ["table_layouts"])

            payload = path.read_text(encoding="utf-8")
            self.assertIn("employee_emails", payload)
            self.assertNotIn("table_layouts", payload)


if __name__ == "__main__":
    unittest.main()
