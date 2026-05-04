from __future__ import annotations

import argparse
import concurrent.futures
import difflib
import html
import hashlib
import io
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import tomllib
import urllib.error
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from importlib import metadata as importlib_metadata
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable

import tkinter as tk
from tkinter import colorchooser, filedialog, font as tkfont, messagebox, simpledialog, ttk

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import pythoncom
    import win32com.client
except ImportError:  # pragma: no cover - optional Windows integration
    pythoncom = None
    win32com = None


DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})")
# Revision suffix after the date token: rev 1, rv1, r1, r 1, r-1, r.1, Revision 1, etc.
REVISION_PATTERN = re.compile(
    r"(?:^|[\s._-])(?:rev(?:ision)?[\s._-]*|rv[\s._-]*|r[\s._-]*)(\d+)(?=$|[\s._-])",
    re.IGNORECASE,
)
DSS_HASH_AZ2_COL = 52  # AZ
DSS_HASH_AZ2_ROW = 2
PF_PATTERN = re.compile(r"\b(PF\d+(?:-\d+)?)\b", re.IGNORECASE)
WORKBOOK_CALC_ID_PATTERN = re.compile(br'\s(?:calcId|fullCalcOnLoad|forceFullCalc|calcCompleted)="[^"]*"')
EXCEL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
EXCEL_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
EXCEL_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DSS_HASH_MIN_COL = 11  # K
DSS_HASH_MAX_COL = 52  # AZ
DSS_HASH_MIN_ROW = 25
DSS_HASH_MAX_ROW = 36
BLOCK_START_ROWS = (25, 28, 31, 34)

LEFT_NAME_COLS = tuple(range(20, 28))   # T:AA
LEFT_HOUR_COLS = tuple(range(29, 32))   # AC:AE
RIGHT_NAME_COLS = tuple(range(29, 48))  # AC:AV
RIGHT_HOUR_COLS = tuple(range(50, 53))  # AX:AZ
CONFIG_FILENAME = "dss_hours_tracker_config.json"
DEFAULT_PROFILE_NAME = "Default"
DISPLAY_APP_NAME = "DSS Tools"
APP_DIRNAME = "DSSTools"
LEGACY_APP_DIRNAME = "DSSHoursTracker"
DISTRIBUTION_PACKAGE_NAMES = ("dss-tools", "dss-hours-tracker")
CACHE_DIRNAME = "cache"
CACHE_RETENTION_DAYS = 7
HASH_CHECK_INTERVAL_MS = 300000
AUTO_OUTLOOK_SYNC_DELAY_MS = 60000
AUTO_UPDATE_CHECK_DELAY_MS = 15000
DEFAULT_HASH_POLL_MINUTES = 5
BUG_REPORT_EMAIL = "lross@jatechpowersystems.com"
MAX_PARALLEL_PARSE_WORKERS = 2
UPDATE_DIRNAME = "updates"
UPDATER_EXE_NAME = "DSSToolsUpdater.exe"
INSTALLER_EXTENSIONS = (".exe", ".msi", ".msix", ".msixbundle")
CHECKSUM_ASSET_NAMES = ("checksums.txt", "sha256sums.txt", "sha256sums", "sha256sum.txt")
GITHUB_REPO_SLUG = "LochlanRoss/DSS-Tools"
GITHUB_LATEST_RELEASE_URL = f"https://api.github.com/repos/{GITHUB_REPO_SLUG}/releases/latest"
APP_ICON_CANDIDATE_NAMES = (
    "dss_tools.ico",
    "DSSTools.ico",
    "app_icon.ico",
    "icon.ico",
    "app.ico",
)


class OperationCancelled(RuntimeError):
    pass


def discover_app_version() -> str:
    """Resolve display / compare version: env, frozen bundle file, installed package, then pyproject."""
    env_version = os.environ.get("DSS_APP_VERSION", "").strip()
    if env_version:
        return env_version
    if getattr(sys, "frozen", False):
        bundle_root = Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
        bundled = bundle_root / "dss_app_version.txt"
        try:
            text = bundled.read_text(encoding="utf-8").strip()
        except OSError:
            text = ""
        if text:
            return text
    for package_name in DISTRIBUTION_PACKAGE_NAMES:
        try:
            return importlib_metadata.version(package_name)
        except importlib_metadata.PackageNotFoundError:
            continue
    pyproject_path = Path(__file__).with_name("pyproject.toml")
    try:
        with pyproject_path.open("rb") as handle:
            payload = tomllib.load(handle)
    except (OSError, tomllib.TOMLDecodeError):
        return "0.0.0"
    project = payload.get("project", {})
    return str(project.get("version", "0.0.0")).strip() or "0.0.0"


def normalize_release_version(version_text: str) -> str:
    return version_text.strip().lstrip("vV")


def version_key(version_text: str) -> tuple[tuple[int, object], ...]:
    normalized = normalize_release_version(version_text)
    tokens = re.findall(r"\d+|[A-Za-z]+", normalized)
    if not tokens:
        return ((0, 0),)
    key: list[tuple[int, object]] = []
    for token in tokens:
        if token.isdigit():
            key.append((0, int(token)))
        else:
            key.append((1, token.casefold()))
    return tuple(key)


def is_newer_version(candidate_version: str, current_version: str) -> bool:
    return version_key(candidate_version) > version_key(current_version)


def parse_latest_release_payload(payload: dict) -> dict[str, object]:
    tag_name = str(payload.get("tag_name", "")).strip()
    version = normalize_release_version(tag_name)
    assets: list[dict[str, object]] = []
    for asset in payload.get("assets", []):
        if not isinstance(asset, dict):
            continue
        name = str(asset.get("name", "")).strip()
        download_url = str(asset.get("browser_download_url", "")).strip()
        if not name:
            continue
        assets.append(
            {
                "name": name,
                "download_url": download_url,
                "size": int(asset.get("size", 0) or 0),
                "content_type": str(asset.get("content_type", "")).strip(),
            }
        )
    return {
        "tag_name": tag_name,
        "version": version,
        "name": str(payload.get("name", "")).strip(),
        "html_url": str(payload.get("html_url", "")).strip(),
        "published_at": str(payload.get("published_at", "")).strip(),
        "body": str(payload.get("body", "")),
        "asset_names": [str(asset["name"]) for asset in assets],
        "assets": assets,
    }


def fetch_latest_release_info(url: str = GITHUB_LATEST_RELEASE_URL, timeout: int = 10) -> dict[str, object]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": f"dss-tools/{discover_app_version()}",
            "Accept": "application/vnd.github+json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"GitHub release check failed with HTTP {exc.code}.") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Could not reach GitHub to check for updates.") from exc
    except (OSError, json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise RuntimeError("Could not read the GitHub release response.") from exc
    if not isinstance(payload, dict):
        raise RuntimeError("Unexpected GitHub release response.")
    return parse_latest_release_payload(payload)


def choose_release_installer_asset(release_info: dict[str, object]) -> dict[str, object] | None:
    assets = release_info.get("assets", [])
    if not isinstance(assets, list):
        return None
    ranked_assets: list[tuple[int, dict[str, object]]] = []
    for asset in assets:
        if not isinstance(asset, dict):
            continue
        name = str(asset.get("name", "")).strip()
        suffix = Path(name).suffix.lower()
        if suffix not in INSTALLER_EXTENSIONS:
            continue
        ranked_assets.append((INSTALLER_EXTENSIONS.index(suffix), asset))
    if not ranked_assets:
        return None
    ranked_assets.sort(key=lambda item: (item[0], str(item[1].get("name", "")).casefold()))
    return ranked_assets[0][1]


def choose_release_checksum_asset(release_info: dict[str, object]) -> dict[str, object] | None:
    assets = release_info.get("assets", [])
    if not isinstance(assets, list):
        return None
    for preferred_name in CHECKSUM_ASSET_NAMES:
        for asset in assets:
            if isinstance(asset, dict) and str(asset.get("name", "")).strip().casefold() == preferred_name.casefold():
                return asset
    return None


def parse_checksum_manifest(manifest_text: str) -> dict[str, str]:
    checksums: dict[str, str] = {}
    for raw_line in manifest_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        match = re.match(r"^([0-9a-fA-F]{64})\s+[* ]?(.+)$", line)
        if not match:
            continue
        checksums[Path(match.group(2).strip()).name] = match.group(1).lower()
    return checksums


def checksum_for_asset_name(manifest_text: str, asset_name: str) -> str | None:
    return parse_checksum_manifest(manifest_text).get(Path(asset_name).name)


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def download_url_bytes(url: str, timeout: int = 60) -> bytes:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": f"dss-tools/{APP_VERSION}",
            "Accept": "application/octet-stream, text/plain, application/vnd.github+json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"Download failed with HTTP {exc.code}.") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError("Could not reach GitHub to download the update asset.") from exc
    except OSError as exc:
        raise RuntimeError("Could not download the update asset.") from exc


def download_release_asset(url: str, destination: Path, timeout: int = 300) -> Path:
    payload = download_url_bytes(url, timeout=timeout)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(payload)
    return destination


def get_windows_network_profile(timeout: int = 10) -> dict[str, object]:
    if os.name != "nt":
        return {"connected": False, "supported": False, "reason": "Windows-only check"}
    command = [
        "powershell",
        "-NoProfile",
        "-Command",
        "[Windows.Networking.Connectivity.NetworkInformation, Windows, ContentType = WindowsRuntime] | Out-Null; "
        "$profile = [Windows.Networking.Connectivity.NetworkInformation]::GetInternetConnectionProfile(); "
        "if ($null -eq $profile) { @{ connected = $false; supported = $true } | ConvertTo-Json -Compress; exit 0 }; "
        "$cost = $profile.GetConnectionCost(); $level = $profile.GetNetworkConnectivityLevel().ToString(); "
        "[ordered]@{ connected = ($level -ne 'None'); supported = $true; connectivity_level = $level; is_wlan = [bool]$profile.IsWlanConnectionProfile; network_cost_type = $cost.NetworkCostType.ToString(); roaming = [bool]$cost.Roaming; over_data_limit = [bool]$cost.OverDataLimit; approaching_data_limit = [bool]$cost.ApproachingDataLimit; background_restricted = [bool]$cost.BackgroundDataUsageRestricted } | ConvertTo-Json -Compress",
    ]
    try:
        completed = subprocess.run(command, capture_output=True, text=True, timeout=timeout, check=True)
    except (OSError, subprocess.SubprocessError) as exc:
        return {"connected": False, "supported": False, "reason": str(exc)}
    try:
        payload = json.loads(completed.stdout.strip() or "{}")
    except json.JSONDecodeError:
        return {"connected": False, "supported": False, "reason": "Invalid network profile response"}
    return payload if isinstance(payload, dict) else {"connected": False, "supported": False}


def is_unmetered_wifi_profile(profile: dict[str, object]) -> bool:
    if not isinstance(profile, dict):
        return False
    return (
        bool(profile.get("supported", False))
        and bool(profile.get("connected", False))
        and bool(profile.get("is_wlan", False))
        and str(profile.get("network_cost_type", "")).casefold() == "unrestricted"
        and not bool(profile.get("roaming", False))
        and not bool(profile.get("over_data_limit", False))
        and not bool(profile.get("approaching_data_limit", False))
        and not bool(profile.get("background_restricted", False))
    )


def describe_network_profile(profile: dict[str, object]) -> str:
    if not isinstance(profile, dict):
        return "network state unavailable"
    if not bool(profile.get("supported", False)):
        return str(profile.get("reason", "network state unavailable")).strip() or "network state unavailable"
    if not bool(profile.get("connected", False)):
        return "no internet connection"
    connection_type = "Wi-Fi" if bool(profile.get("is_wlan", False)) else "non-Wi-Fi network"
    cost = str(profile.get("network_cost_type", "Unknown")).strip() or "Unknown"
    return f"{connection_type}, {cost}"


APP_VERSION = discover_app_version()


def directional_sort_key(value: str, descending: bool = False) -> tuple[int, object]:
    text = str(value)
    try:
        number = float(text)
    except ValueError:
        folded = text.casefold()
        if descending:
            return (1, tuple(-ord(char) for char in folded))
        return (1, folded)
    return (0, -number if descending else number)


def weekly_rollup_sort_key(column: str, row: tuple[str, ...], descending: bool) -> tuple[object, ...]:
    columns = {
        "source_file": 0,
        "week_start": 1,
        "week_end": 2,
        "employee": 3,
        "st": 4,
        "ot": 5,
        "dt": 6,
        "total": 7,
        "expanded": 8,
        "row_type": 9,
    }
    primary_index = columns.get(column)
    if primary_index is None:
        return tuple()
    row_type_rank = 1 if row[9] == "Crew Total" else 0
    return (
        directional_sort_key(row[primary_index], descending),
        directional_sort_key(row[0], False),
        directional_sort_key(row[1], descending if column in {"week_start", "week_end"} else False),
        row_type_rank,
        directional_sort_key(row[3], False),
    )


def daily_rollup_sort_key(column: str, row: tuple[str, ...], descending: bool) -> tuple[object, ...]:
    columns = {
        "source_file": 0,
        "work_date": 1,
        "employee": 2,
        "st": 3,
        "ot": 4,
        "dt": 5,
        "total": 6,
        "expanded": 7,
        "row_type": 8,
    }
    primary_index = columns.get(column)
    if primary_index is None:
        return tuple()
    row_type_rank = 1 if row[8] == "Crew Total" else 0
    return (
        directional_sort_key(row[primary_index], descending),
        directional_sort_key(row[0], False),
        directional_sort_key(row[1], descending if column == "work_date" else False),
        row_type_rank,
        directional_sort_key(row[2], False),
    )


@dataclass(frozen=True)
class DailyRecord:
    source_path: Path
    source_file: str
    work_date: date
    source_sheet: str
    employee: str
    st: float
    ot: float
    dt: float
    source_ranges: str

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class WeeklyRecord:
    source_file: str
    week_start: date
    week_end: date
    employee: str
    st: float
    ot: float
    dt: float

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class WeeklyRollupRow:
    source_file: str
    week_start: date
    week_end: date
    employee: str
    st: float
    ot: float
    dt: float
    row_type: str

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class DailySummaryRecord:
    source_file: str
    work_date: date
    employee: str
    st: float
    ot: float
    dt: float

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class DailyRollupRow:
    source_file: str
    work_date: date
    employee: str
    st: float
    ot: float
    dt: float
    row_type: str

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class WeekTotalRow:
    week_start: date
    week_end: date
    st: float
    ot: float
    dt: float

    @property
    def total(self) -> float:
        return round(self.st + self.ot + self.dt, 2)


@dataclass(frozen=True)
class TrackerData:
    source_paths: list[Path]
    file_hashes: dict[Path, str]
    reused_paths: list[Path]
    reloaded_paths: list[Path]
    cache_status_by_path: dict[Path, str]
    daily_records: list[DailyRecord]
    employee_names: list[str]
    weekly_summary: list[WeeklyRecord]
    weekly_rollup: list[WeeklyRollupRow]
    daily_summary: list[DailySummaryRecord]
    daily_rollup: list[DailyRollupRow]
    week_totals: list[WeekTotalRow]
    combined_weekly_summary: list[WeeklyRecord]
    combined_daily_summary: list[DailySummaryRecord]
    parse_warnings: list["SheetParseWarning"]
    workbook_health: list["WorkbookHealthItem"]


@dataclass(frozen=True)
class FormattingProfile:
    name: str
    st_threshold: float | None
    ot_threshold: float | None
    daily_st_threshold: float | None
    max_hours_per_day: float | None


@dataclass(frozen=True)
class ErrorFinding:
    employee: str
    week_start: date
    week_end: date
    hour_type: str
    threshold: float
    actual_total: float
    delta: float
    trigger_date: date
    trigger_day_st: float
    trigger_day_ot: float
    trigger_day_dt: float
    source_files: str
    reason: str
    breakdown: str


@dataclass(frozen=True)
class NameTypoWarning:
    employee: str
    similar_employee: str
    similarity: float
    locations: list[str]


@dataclass(frozen=True)
class EmailDraftRequest:
    employee: str
    email: str
    week_start: date
    week_end: date
    records: list[DailyRecord]


@dataclass(frozen=True)
class FilterSelection:
    mode: str
    value: str
    values: tuple[str, ...] = ()


@dataclass(frozen=True)
class UiThemeColors:
    """Semantic UI colours (hex #RRGGBB) for tables, tooltips, and main chrome."""

    alert_row_background: str = "#fde2e7"
    alert_row_foreground: str = "#9f1239"
    crew_total_background: str = "#e0f2f1"
    crew_total_foreground: str = "#134e4a"
    tooltip_background: str = "#f1f5f9"
    tooltip_foreground: str = "#334155"
    reports_outline_background: str = "#fbcfe8"
    reports_outline_foreground: str = "#be123c"
    table_background: str = "#f4f4f5"
    content_chrome_background: str = "#ffffff"


DEFAULT_UI_THEME = UiThemeColors()

# (human label, UiThemeColors attribute name) for Configuration → Appearance.
UI_THEME_CONFIG_FIELDS: tuple[tuple[str, str], ...] = (
    ("Alert table row — background", "alert_row_background"),
    ("Alert table row — text", "alert_row_foreground"),
    ("Crew total row — background", "crew_total_background"),
    ("Crew total row — text", "crew_total_foreground"),
    ("Tooltip — background", "tooltip_background"),
    ("Tooltip — text", "tooltip_foreground"),
    ("Main content area — background", "content_chrome_background"),
    ("Table — cell background", "table_background"),
)

DSS_TABLE_TREEVIEW_STYLE = "DssTable.Treeview"
DSS_TAB_SWATCH_W = 6
DSS_TAB_SWATCH_H = 18


def normalize_ui_hex_color(value: str) -> str | None:
    """Return canonical #rrggbb or None if the string is not a valid 24-bit hex colour."""
    text = str(value).strip().lower()
    if not text.startswith("#"):
        return None
    body = text[1:]
    if len(body) == 3 and all(c in "0123456789abcdef" for c in body):
        body = "".join(c * 2 for c in body)
    if len(body) != 6 or any(c not in "0123456789abcdef" for c in body):
        return None
    return f"#{body}"


def _hex_to_rgb_triplet(hex_color: str) -> tuple[int, int, int]:
    h = normalize_ui_hex_color(hex_color) or "#000000"
    body = h[1:]
    return int(body[0:2], 16), int(body[2:4], 16), int(body[4:6], 16)


def lighten_hex_color(hex_color: str, delta: int) -> str:
    """Lighten (or darken with negative delta) a #rrggbb colour for derived UI tones."""
    r, g, b = _hex_to_rgb_triplet(hex_color)
    r = min(255, max(0, r + delta))
    g = min(255, max(0, g + delta))
    b = min(255, max(0, b + delta))
    return f"#{r:02x}{g:02x}{b:02x}"


def solid_tab_swatch_photo(master: tk.Misc, color: str, width: int = DSS_TAB_SWATCH_W, height: int = DSS_TAB_SWATCH_H) -> tk.PhotoImage:
    """Small vertical stripe for ttk.Notebook tabs (per-tab alert tint without restyling the whole tab bar)."""
    photo = tk.PhotoImage(master=master, width=width, height=height)
    row_pixels = "{" + " ".join([color] * width) + "}"
    data = " ".join([row_pixels] * height)
    photo.put(data)
    return photo


def configure_dss_table_treeview_style(master: tk.Misc, theme: UiThemeColors) -> None:
    """Shared Treeview style for all DataTable instances."""
    style = ttk.Style(master)
    bg = theme.table_background
    heading_bg = lighten_hex_color(bg, -18)
    heading_fg = "#18181b"
    cell_fg = "#18181b"
    try:
        style.configure(
            DSS_TABLE_TREEVIEW_STYLE,
            background=bg,
            fieldbackground=bg,
            foreground=cell_fg,
        )
        style.configure(f"{DSS_TABLE_TREEVIEW_STYLE}.Heading", background=heading_bg, foreground=heading_fg)
        style.map(
            DSS_TABLE_TREEVIEW_STYLE,
            background=[("selected", "#bfdbfe")],
            foreground=[("selected", "#1e3a8a")],
        )
    except tk.TclError:
        pass


def parse_ui_theme_payload(raw: object, defaults: UiThemeColors = DEFAULT_UI_THEME) -> UiThemeColors:
    if not isinstance(raw, dict):
        return defaults
    kwargs: dict[str, str] = {}
    for key in (
        "alert_row_background",
        "alert_row_foreground",
        "crew_total_background",
        "crew_total_foreground",
        "tooltip_background",
        "tooltip_foreground",
        "reports_outline_background",
        "reports_outline_foreground",
        "table_background",
        "content_chrome_background",
    ):
        fallback = getattr(defaults, key)
        raw_val = raw.get(key)
        if raw_val is None or (isinstance(raw_val, str) and not str(raw_val).strip()):
            kwargs[key] = fallback
            continue
        normalized = normalize_ui_hex_color(str(raw_val).strip())
        kwargs[key] = normalized if normalized is not None else fallback
    return UiThemeColors(**kwargs)


@dataclass(frozen=True)
class AppSettings:
    disable_name_typo_notifications: bool = False
    hash_poll_minutes: int = DEFAULT_HASH_POLL_MINUTES
    show_daily_raw_tab: bool = True
    quickload_last_sources_enabled: bool = True
    quickload_cancel_hotkey: str = "<Escape>"
    auto_update_check_enabled: bool = True
    auto_download_updates_on_unmetered_wifi: bool = True
    ui_theme: UiThemeColors = field(default_factory=lambda: DEFAULT_UI_THEME)


@dataclass(frozen=True)
class SheetParseWarning:
    source_file: str
    source_sheet: str
    work_date: str
    issue: str
    details: str


@dataclass(frozen=True)
class WorkbookHealthItem:
    source_file: str
    status: str
    details: str


def _app_root_primary_and_legacy(base_dir: Path) -> Path:
    primary = base_dir / APP_DIRNAME
    legacy = base_dir / LEGACY_APP_DIRNAME
    if primary.exists():
        return primary
    if legacy.exists():
        return legacy
    return primary


def get_app_root() -> Path:
    if os.name == "nt":
        base = os.environ.get("LOCALAPPDATA")
        if base:
            return _app_root_primary_and_legacy(Path(base))
    if getattr(sys, "frozen", False):
        return _app_root_primary_and_legacy(Path(sys.executable).resolve().parent)
    primary = Path.home() / f".{APP_DIRNAME}"
    legacy = Path.home() / f".{LEGACY_APP_DIRNAME}"
    if primary.exists():
        return primary
    if legacy.exists():
        return legacy
    return primary


def ensure_app_directories() -> tuple[Path, Path]:
    app_root = get_app_root()
    cache_dir = app_root / CACHE_DIRNAME
    app_root.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)
    return app_root, cache_dir


def resolve_app_icon_path() -> Path | None:
    """Bundled onefile: MEIPASS first; dev: next to this module. Optional single *.ico in repo root."""
    roots: list[Path] = []
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        roots.append(Path(sys._MEIPASS))
    source_root = Path(__file__).resolve().parent
    roots.append(source_root)
    for base in roots:
        for name in APP_ICON_CANDIDATE_NAMES:
            cand = base / name
            if cand.is_file():
                return cand
    if not getattr(sys, "frozen", False):
        try:
            loose = sorted(source_root.glob("*.ico"))
            if len(loose) == 1:
                return loose[0]
        except OSError:
            pass
    return None


def apply_tk_window_icon(root: tk.Misc) -> None:
    path = resolve_app_icon_path()
    if path is None:
        return
    resolved = path.resolve()
    try:
        if os.name == "nt" and resolved.suffix.lower() == ".ico":
            root.iconbitmap(default=str(resolved))
            return
    except tk.TclError:
        return
    if resolved.suffix.lower() in {".png", ".gif", ".ppm", ".pgm"}:
        try:
            photo = tk.PhotoImage(master=root, file=str(resolved))
            root.iconphoto(True, photo)
            setattr(root, "_dss_app_icon_photo", photo)
        except tk.TclError:
            pass


def parse_sheet_date(sheet_name: str) -> date | None:
    match = DATE_PATTERN.search(sheet_name)
    if not match:
        return None
    return datetime.strptime(match.group(1), "%Y-%m-%d").date()


def parse_sheet_revision(sheet_name: str) -> int:
    match = REVISION_PATTERN.search(sheet_name)
    if not match:
        return 0
    return int(match.group(1))


def _normalize_az2_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float) and math.isnan(value):
            return ""
        fv = float(value)
        if abs(fv - round(fv)) < 1e-9 and -1e9 < fv < 1e9:
            return str(int(round(fv)))
        return str(value).strip()
    return str(value).strip()


def revision_level_from_az2(value: object) -> int | None:
    """Return revision level 0,1,2,... from AZ2, or None if the cell does not encode a level (blank / unknown)."""
    text = _normalize_az2_cell_text(value)
    if not text:
        return None
    lowered = text.casefold()
    if lowered in {"n", "no", "false", "0", "non", "none", "base", "orig", "original"}:
        return 0
    if lowered in {"y", "yes", "true", "rev", "revision"}:
        return None
    match = REVISION_PATTERN.search(text)
    if match:
        return int(match.group(1))
    try:
        as_float = float(text.replace(",", "."))
    except ValueError:
        return None
    if math.isnan(as_float):
        return None
    if abs(as_float - round(as_float)) < 1e-9 and 0 <= round(as_float) <= 99:
        return int(round(as_float))
    return None


def revision_presence_from_az2(value: object) -> bool | None:
    """True = cell says this is a revision sheet; False = not a revision; None = unknown."""
    text = _normalize_az2_cell_text(value)
    if not text:
        return None
    lowered = text.casefold()
    if lowered in {"n", "no", "false", "0", "non", "none", "base", "orig", "original"}:
        return False
    if lowered in {"y", "yes", "true", "rev", "revision"}:
        return True
    if REVISION_PATTERN.search(text):
        return None
    try:
        as_float = float(text.replace(",", "."))
    except ValueError:
        return None
    if math.isnan(as_float):
        return None
    if abs(as_float - round(as_float)) < 1e-9 and 0 <= round(as_float) <= 99:
        return None
    return None


def az2_revision_matches_sheet_name(sheet_name: str, az2_value: object) -> tuple[bool, str | None]:
    """
    AZ2 should agree with the revision encoded in the sheet tab name.
    Returns (ok, warning_detail_or_none).
    """
    name_rev = parse_sheet_revision(sheet_name)
    text = _normalize_az2_cell_text(az2_value)
    if not text:
        if name_rev > 0:
            return False, "AZ2 is blank but the sheet name includes a revision suffix."
        return True, None
    level = revision_level_from_az2(az2_value)
    if level is not None:
        if level != name_rev:
            return (
                False,
                f"AZ2 indicates revision level {level} but the sheet name encodes revision {name_rev} "
                f"(parsed from '{sheet_name}').",
            )
        return True, None
    presence = revision_presence_from_az2(az2_value)
    if presence is None:
        return True, None
    name_has_rev = name_rev > 0
    if presence != name_has_rev:
        return (
            False,
            f"AZ2 indicates {'a' if presence else 'no'} revision sheet, but the sheet name "
            f"{'has' if name_has_rev else 'has no'} a revision suffix (parsed revision {name_rev}).",
        )
    return True, None


def extract_pf_identifier(source_name: str) -> str:
    match = PF_PATTERN.search(source_name)
    if match:
        return match.group(1).upper()
    return Path(source_name).stem.strip() or source_name.strip()


_QUICKLOAD_CANCEL_MODIFIER_KEYSYMS: frozenset[str] = frozenset(
    {
        "Shift_L",
        "Shift_R",
        "Control_L",
        "Control_R",
        "Alt_L",
        "Alt_R",
        "Meta_L",
        "Meta_R",
        "Super_L",
        "Super_R",
        "Caps_Lock",
        "Num_Lock",
        "ISO_Level3_Shift",
    }
)

QUICKLOAD_CANCEL_HOTKEY_PRESETS: tuple[str, ...] = (
    "<Escape>",
    "<F8>",
    "<F9>",
    "<F10>",
    "<Control-q>",
    "<Control-w>",
    "<Shift-Escape>",
)


def normalize_quickload_cancel_hotkey(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return "<Escape>"
    if not s.startswith("<"):
        s = f"<{s}"
    if not s.endswith(">"):
        s = f"{s}>"
    return s


def is_allowed_quickload_cancel_hotkey(sequence: str) -> bool:
    seq = normalize_quickload_cancel_hotkey(sequence)
    if seq in QUICKLOAD_CANCEL_HOTKEY_PRESETS:
        return True
    if re.fullmatch(r"<F([1-9]|1\d|2[0-4])>", seq):
        return True
    if seq in {"<Escape>", "<Pause>", "<Delete>", "<Insert>", "<Home>", "<End>", "<Next>", "<Prior>", "<Tab>"}:
        return True
    if re.fullmatch(r"<Shift-F([1-9]|1\d|2[0-4])>", seq):
        return True
    if re.fullmatch(r"<Control-Key-[a-z0-9,\.]>", seq, re.IGNORECASE):
        return True
    if re.fullmatch(r"<Control-[a-z]>", seq, re.IGNORECASE):
        return True
    return False


def binding_sequence_from_keypress_event(event: tk.Event) -> str | None:
    """Build a Tk virtual event string from a key press, or None if the key should be ignored."""
    keysym = event.keysym
    if keysym in _QUICKLOAD_CANCEL_MODIFIER_KEYSYMS or keysym in {"", "??"}:
        return None
    state = int(getattr(event, "state", 0) or 0)
    ctrl = (state & 0x4) != 0
    shift = (state & 0x1) != 0
    if keysym.startswith("F"):
        suffix = keysym[1:]
        if suffix.isdigit():
            if shift and not ctrl:
                return f"<Shift-{keysym}>"
            if ctrl:
                return f"<Control-{keysym}>"
            return f"<{keysym}>"
    if keysym == "Escape":
        if shift and not ctrl:
            return "<Shift-Escape>"
        if not shift:
            return "<Escape>"
    if ctrl:
        if len(keysym) == 1 and (keysym.isalpha() or keysym in "0123456789,."):
            return f"<Control-Key-{keysym.lower()}>"
        if keysym in ("BackSpace", "Return", "space", "Tab"):
            return f"<Control-Key-{keysym}>"
    return None


def default_formatting_profiles() -> dict[str, FormattingProfile]:
    return {
        DEFAULT_PROFILE_NAME: FormattingProfile(
            name=DEFAULT_PROFILE_NAME,
            st_threshold=40.0,
            ot_threshold=10.0,
            daily_st_threshold=None,
            max_hours_per_day=None,
        )
    }


def read_config_payload(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    try:
        payload = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return {}
    return payload if isinstance(payload, dict) else {}


def parse_threshold_value(raw_value: str) -> float | None:
    text = raw_value.strip()
    if not text:
        return None
    return float(text)


def load_formatting_profiles(config_path: Path) -> tuple[dict[str, FormattingProfile], str]:
    defaults = default_formatting_profiles()
    payload = read_config_payload(config_path)
    if not payload:
        return defaults, DEFAULT_PROFILE_NAME

    profiles: dict[str, FormattingProfile] = {}
    for item in payload.get("profiles", []):
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        profiles[name] = FormattingProfile(
            name=name,
            st_threshold=item.get("st_threshold"),
            ot_threshold=item.get("ot_threshold"),
            daily_st_threshold=item.get("daily_st_threshold"),
            max_hours_per_day=item.get("max_hours_per_day"),
        )

    if not profiles:
        profiles = defaults

    current_profile_name = str(payload.get("current_profile", DEFAULT_PROFILE_NAME))
    if current_profile_name not in profiles:
        current_profile_name = sorted(profiles)[0]
    return profiles, current_profile_name


def save_formatting_profiles(
    config_path: Path,
    profiles: dict[str, FormattingProfile],
    current_profile_name: str,
) -> None:
    payload = read_config_payload(config_path)
    payload["current_profile"] = current_profile_name
    payload["profiles"] = [
        {
                "name": profile.name,
                "st_threshold": profile.st_threshold,
                "ot_threshold": profile.ot_threshold,
                "daily_st_threshold": profile.daily_st_threshold,
                "max_hours_per_day": profile.max_hours_per_day,
            }
        for profile in sorted(profiles.values(), key=lambda item: item.name.lower())
    ]
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_employee_emails(config_path: Path) -> dict[str, str]:
    payload = read_config_payload(config_path)
    raw_map = payload.get("employee_emails", {})
    if not isinstance(raw_map, dict):
        return {}
    emails: dict[str, str] = {}
    for name, email_value in raw_map.items():
        employee = str(name).strip()
        email = str(email_value).strip()
        if employee:
            emails[employee] = email
    return emails


def save_employee_emails(config_path: Path, employee_emails: dict[str, str]) -> None:
    payload = read_config_payload(config_path)
    payload["employee_emails"] = {
        name: email.strip()
        for name, email in sorted(employee_emails.items(), key=lambda item: item[0].lower())
        if name.strip()
    }
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_ignored_name_typos(config_path: Path) -> set[str]:
    payload = read_config_payload(config_path)
    raw_values = payload.get("ignored_name_typos", [])
    if not isinstance(raw_values, list):
        return set()
    return {str(value).strip() for value in raw_values if str(value).strip()}


def save_ignored_name_typos(config_path: Path, ignored_name_typos: set[str]) -> None:
    payload = read_config_payload(config_path)
    payload["ignored_name_typos"] = sorted(value for value in ignored_name_typos if value.strip())
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_employee_groups(config_path: Path) -> dict[str, list[str]]:
    payload = read_config_payload(config_path)
    raw_groups = payload.get("employee_groups", {})
    if not isinstance(raw_groups, dict):
        return {}
    groups: dict[str, list[str]] = {}
    for group_name, employees in raw_groups.items():
        name = str(group_name).strip()
        if not name:
            continue
        if isinstance(employees, list):
            groups[name] = sorted({str(employee).strip() for employee in employees if str(employee).strip()})
    return groups


def save_employee_groups(config_path: Path, employee_groups: dict[str, list[str]]) -> None:
    payload = read_config_payload(config_path)
    payload["employee_groups"] = {
        name: sorted({employee.strip() for employee in employees if employee.strip()})
        for name, employees in sorted(employee_groups.items(), key=lambda item: item[0].lower())
        if name.strip()
    }
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_employee_notes(config_path: Path) -> dict[str, str]:
    payload = read_config_payload(config_path)
    raw_map = payload.get("employee_notes", {})
    if not isinstance(raw_map, dict):
        return {}
    notes: dict[str, str] = {}
    for name, note_value in raw_map.items():
        employee = str(name).strip()
        note = str(note_value).strip()
        if employee:
            notes[employee] = note
    return notes


def save_employee_notes(config_path: Path, employee_notes: dict[str, str]) -> None:
    payload = read_config_payload(config_path)
    payload["employee_notes"] = {
        name: note.strip()
        for name, note in sorted(employee_notes.items(), key=lambda item: item[0].lower())
        if name.strip()
    }
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_job_presets(config_path: Path) -> dict[str, str]:
    payload = read_config_payload(config_path)
    raw_map = payload.get("job_presets", {})
    if not isinstance(raw_map, dict):
        return {}
    presets: dict[str, str] = {}
    for job_name, profile_name in raw_map.items():
        job = str(job_name).strip()
        profile = str(profile_name).strip()
        if job and profile:
            presets[job] = profile
    return presets


def save_job_presets(config_path: Path, job_presets: dict[str, str]) -> None:
    payload = read_config_payload(config_path)
    payload["job_presets"] = {
        job: profile
        for job, profile in sorted(job_presets.items(), key=lambda item: item[0].lower())
        if job.strip() and profile.strip()
    }
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def default_email_body_template() -> str:
    return (
        "<p>Hi {first_name},</p>\n"
        "<p>Here are your DSS hours for the week of {week_start} to {week_end}. "
        "Please use these details to complete your timesheet.</p>\n"
        "{hours_table}\n"
        "<p>Thanks.</p>"
    )


def load_email_templates(config_path: Path) -> tuple[str, str]:
    payload = read_config_payload(config_path)
    subject_template = str(payload.get("email_subject_template", "Timesheet Reminder - Week of {week_start}")).strip()
    if not subject_template:
        subject_template = "Timesheet Reminder - Week of {week_start}"
    body_template = str(payload.get("email_body_template", default_email_body_template()))
    if not body_template.strip():
        body_template = default_email_body_template()
    return subject_template, body_template


def save_last_open_dss_paths(config_path: Path, paths: Iterable[Path]) -> None:
    payload = read_config_payload(config_path)
    payload["last_open_dss_paths"] = [str(Path(path).expanduser().resolve()) for path in paths]
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_last_open_dss_paths(config_path: Path) -> list[Path]:
    payload = read_config_payload(config_path)
    raw = payload.get("last_open_dss_paths", [])
    if not isinstance(raw, list):
        return []
    results: list[Path] = []
    for item in raw:
        try:
            candidate = Path(str(item)).expanduser()
        except (TypeError, ValueError):
            continue
        if str(candidate).strip():
            results.append(candidate)
    return results


def save_email_templates(config_path: Path, subject_template: str, body_template: str) -> None:
    payload = read_config_payload(config_path)
    payload["email_subject_template"] = subject_template
    payload["email_body_template"] = body_template
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_app_settings(config_path: Path) -> AppSettings:
    payload = read_config_payload(config_path)
    raw_settings = payload.get("app_settings", {})
    if not isinstance(raw_settings, dict):
        return AppSettings()

    raw_minutes = raw_settings.get("hash_poll_minutes", DEFAULT_HASH_POLL_MINUTES)
    try:
        hash_poll_minutes = max(1, int(raw_minutes))
    except (TypeError, ValueError):
        hash_poll_minutes = DEFAULT_HASH_POLL_MINUTES

    raw_hotkey = str(raw_settings.get("quickload_cancel_hotkey", "<Escape>")).strip()
    cancel_hotkey = normalize_quickload_cancel_hotkey(raw_hotkey)
    if not is_allowed_quickload_cancel_hotkey(cancel_hotkey):
        cancel_hotkey = "<Escape>"

    return AppSettings(
        disable_name_typo_notifications=bool(raw_settings.get("disable_name_typo_notifications", False)),
        hash_poll_minutes=hash_poll_minutes,
        show_daily_raw_tab=bool(raw_settings.get("show_daily_raw_tab", True)),
        quickload_last_sources_enabled=bool(raw_settings.get("quickload_last_sources_enabled", True)),
        quickload_cancel_hotkey=cancel_hotkey,
        auto_update_check_enabled=bool(raw_settings.get("auto_update_check_enabled", True)),
        auto_download_updates_on_unmetered_wifi=bool(raw_settings.get("auto_download_updates_on_unmetered_wifi", True)),
        ui_theme=parse_ui_theme_payload(raw_settings.get("ui_theme")),
    )


def save_app_settings(config_path: Path, settings: AppSettings) -> None:
    payload = read_config_payload(config_path)
    payload["app_settings"] = {
        "disable_name_typo_notifications": settings.disable_name_typo_notifications,
        "hash_poll_minutes": settings.hash_poll_minutes,
        "show_daily_raw_tab": settings.show_daily_raw_tab,
        "quickload_last_sources_enabled": settings.quickload_last_sources_enabled,
        "quickload_cancel_hotkey": settings.quickload_cancel_hotkey,
        "auto_update_check_enabled": settings.auto_update_check_enabled,
        "auto_download_updates_on_unmetered_wifi": settings.auto_download_updates_on_unmetered_wifi,
        "ui_theme": asdict(settings.ui_theme),
    }
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def remove_config_keys(config_path: Path, keys: Iterable[str]) -> None:
    payload = read_config_payload(config_path)
    changed = False
    for key in keys:
        if key in payload:
            del payload[key]
            changed = True
    if changed:
        if payload:
            config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        elif config_path.exists():
            config_path.unlink()


def load_table_layouts(config_path: Path) -> dict[str, dict]:
    payload = read_config_payload(config_path)
    raw_layouts = payload.get("table_layouts", {})
    if not isinstance(raw_layouts, dict):
        return {}

    layouts: dict[str, dict] = {}
    for table_id, layout in raw_layouts.items():
        if not isinstance(layout, dict):
            continue
        visible_columns = layout.get("visible_columns", [])
        column_widths = layout.get("column_widths", {})
        sort_column = str(layout.get("sort_column", "")).strip()
        sort_descending = bool(layout.get("sort_descending", False))
        column_filters: dict[str, set[str]] = {}
        raw_filters = layout.get("column_filters")
        if isinstance(raw_filters, dict):
            for column, values in raw_filters.items():
                col_key = str(column).strip()
                if not col_key or not isinstance(values, list):
                    continue
                allowed = {str(value) for value in values}
                if allowed:
                    column_filters[col_key] = allowed
        layouts[str(table_id)] = {
            "visible_columns": [str(column) for column in visible_columns if str(column).strip()],
            "column_widths": {
                str(column): int(width)
                for column, width in column_widths.items()
                if str(column).strip() and isinstance(width, (int, float)) and int(width) > 0
            },
            "sort_column": sort_column,
            "sort_descending": sort_descending,
            "column_filters": column_filters,
        }
    return layouts


def save_table_layout(
    config_path: Path,
    table_id: str,
    visible_columns: list[str],
    column_widths: dict[str, int],
    sort_column: str = "",
    sort_descending: bool = False,
    column_filters: dict[str, set[str]] | None = None,
) -> None:
    payload = read_config_payload(config_path)
    raw_layouts = payload.get("table_layouts", {})
    layouts = raw_layouts if isinstance(raw_layouts, dict) else {}
    entry: dict[str, object] = {
        "visible_columns": list(visible_columns),
        "column_widths": {
            column: int(width)
            for column, width in column_widths.items()
            if int(width) > 0
        },
        "sort_column": sort_column,
        "sort_descending": sort_descending,
    }
    if column_filters:
        entry["column_filters"] = {
            column: sorted(values, key=lambda item: item.casefold())
            for column, values in column_filters.items()
            if column.strip() and values
        }
    layouts[table_id] = entry
    payload["table_layouts"] = layouts
    config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def clear_cache_files(cache_dir: Path) -> int:
    deleted = 0
    if not cache_dir.exists():
        return deleted
    for cache_file in cache_dir.glob("*.json"):
        try:
            cache_file.unlink()
            deleted += 1
        except OSError:
            continue
    return deleted


def source_cache_key(source_path: Path) -> str:
    normalized = normalize_windows_path(str(source_path.resolve()))
    return hashlib.blake2b(normalized.encode("utf-8"), digest_size=16).hexdigest()


def cache_file_path(cache_dir: Path, source_path: Path) -> Path:
    return cache_dir / f"{source_cache_key(source_path)}.json"


def serialize_daily_record(record: DailyRecord) -> dict:
    return {
        "source_path": str(record.source_path),
        "source_file": record.source_file,
        "work_date": record.work_date.isoformat(),
        "source_sheet": record.source_sheet,
        "employee": record.employee,
        "st": record.st,
        "ot": record.ot,
        "dt": record.dt,
        "source_ranges": record.source_ranges,
    }


def deserialize_daily_record(payload: dict) -> DailyRecord:
    return DailyRecord(
        source_path=Path(payload["source_path"]),
        source_file=str(payload["source_file"]),
        work_date=datetime.strptime(str(payload["work_date"]), "%Y-%m-%d").date(),
        source_sheet=str(payload["source_sheet"]),
        employee=str(payload["employee"]),
        st=float(payload["st"]),
        ot=float(payload["ot"]),
        dt=float(payload["dt"]),
        source_ranges=str(payload["source_ranges"]),
    )


def serialize_sheet_parse_warning(warning: SheetParseWarning) -> dict:
    return {
        "source_file": warning.source_file,
        "source_sheet": warning.source_sheet,
        "work_date": warning.work_date,
        "issue": warning.issue,
        "details": warning.details,
    }


def deserialize_sheet_parse_warning(payload: dict) -> SheetParseWarning:
    return SheetParseWarning(
        source_file=str(payload["source_file"]),
        source_sheet=str(payload["source_sheet"]),
        work_date=str(payload["work_date"]),
        issue=str(payload["issue"]),
        details=str(payload["details"]),
    )


def serialize_workbook_health_item(item: WorkbookHealthItem) -> dict:
    return {
        "source_file": item.source_file,
        "status": item.status,
        "details": item.details,
    }


def deserialize_workbook_health_item(payload: dict) -> WorkbookHealthItem:
    return WorkbookHealthItem(
        source_file=str(payload["source_file"]),
        status=str(payload["status"]),
        details=str(payload["details"]),
    )


def purge_stale_cache(cache_dir: Path, now: datetime | None = None) -> None:
    current_time = now or datetime.now()
    cutoff = current_time - timedelta(days=CACHE_RETENTION_DAYS)
    for cache_file in cache_dir.glob("*.json"):
        try:
            modified = datetime.fromtimestamp(cache_file.stat().st_mtime)
        except OSError:
            continue
        if modified < cutoff:
            try:
                cache_file.unlink()
            except OSError:
                pass


def load_cached_daily_records(cache_dir: Path, source_path: Path, file_hash: str) -> list[DailyRecord] | None:
    cache_path = cache_file_path(cache_dir, source_path)
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if payload.get("file_hash") != file_hash:
        return None
    if normalize_windows_path(str(payload.get("source_path", ""))) != normalize_windows_path(str(source_path)):
        return None
    records_payload = payload.get("records", [])
    if not isinstance(records_payload, list):
        return None
    try:
        return [deserialize_daily_record(item) for item in records_payload]
    except (KeyError, TypeError, ValueError):
        return None


def load_cached_source_analysis(
    cache_dir: Path,
    source_path: Path,
    file_hash: str,
) -> tuple[list[SheetParseWarning], list[WorkbookHealthItem]] | None:
    cache_path = cache_file_path(cache_dir, source_path)
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if payload.get("file_hash") != file_hash:
        return None
    if normalize_windows_path(str(payload.get("source_path", ""))) != normalize_windows_path(str(source_path)):
        return None
    if "parse_warnings" not in payload or "workbook_health" not in payload:
        return None
    warnings_payload = payload.get("parse_warnings", [])
    health_payload = payload.get("workbook_health", [])
    if not isinstance(warnings_payload, list) or not isinstance(health_payload, list):
        return None
    try:
        return (
            [deserialize_sheet_parse_warning(item) for item in warnings_payload],
            [deserialize_workbook_health_item(item) for item in health_payload],
        )
    except (KeyError, TypeError, ValueError):
        return None


def read_workbook_cache_payload(cache_dir: Path, source_path: Path) -> dict | None:
    cache_path = cache_file_path(cache_dir, source_path)
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if normalize_windows_path(str(payload.get("source_path", ""))) != normalize_windows_path(str(source_path)):
        return None
    return payload if isinstance(payload, dict) else None


def merge_workbook_from_cache_by_sheet_hashes(
    *,
    source_path: Path,
    workbook_bytes: bytes,
    cache_dir: Path,
    should_cancel: Callable[[], bool],
    progress_callback: Callable[[float, str], None] | None,
) -> tuple[list[DailyRecord], list[SheetParseWarning], list[WorkbookHealthItem], dict[str, str], str] | None:
    """
    When the combined workbook fingerprint changed but preferred sheets per date are unchanged,
    re-parse only dated sheets whose per-sheet digest changed (or were removed), and merge with
    cached records and warnings for unchanged sheets.
    """
    if should_cancel():
        return None
    payload = read_workbook_cache_payload(cache_dir, source_path)
    if not payload:
        return None
    old_sh_raw = payload.get("sheet_hashes")
    if not isinstance(old_sh_raw, dict) or not old_sh_raw:
        return None
    old_sh = {str(name): str(value) for name, value in old_sh_raw.items() if str(name).strip() and str(value).strip()}
    if not old_sh:
        return None
    records_payload = payload.get("records", [])
    if not isinstance(records_payload, list):
        return None
    try:
        old_records = [deserialize_daily_record(item) for item in records_payload]
    except (KeyError, TypeError, ValueError):
        return None
    warnings_payload = payload.get("parse_warnings", [])
    health_payload = payload.get("workbook_health", [])
    if not isinstance(warnings_payload, list):
        warnings_payload = []
    if not isinstance(health_payload, list):
        health_payload = []
    try:
        old_warns = [deserialize_sheet_parse_warning(item) for item in warnings_payload if isinstance(item, dict)]
        old_health = [deserialize_workbook_health_item(item) for item in health_payload if isinstance(item, dict)]
    except (KeyError, TypeError, ValueError):
        return None

    new_sh = compute_all_dated_sheet_hashes(workbook_bytes)
    if not new_sh:
        return None
    new_fp = combine_sheet_hashes(new_sh)
    old_pref = dict(select_preferred_dated_sheets(old_sh.keys()))
    new_pref = dict(select_preferred_dated_sheets(new_sh.keys()))
    if old_pref != new_pref:
        return None

    stale_dates: set[date] = set()
    for sheet_name in set(old_sh) | set(new_sh):
        if old_sh.get(sheet_name) != new_sh.get(sheet_name):
            sheet_date = parse_sheet_date(sheet_name)
            if sheet_date is not None:
                stale_dates.add(sheet_date)
    if not stale_dates:
        return None

    to_repars = frozenset({new_pref[d] for d in stale_dates if d in new_pref})
    if not to_repars:
        return None

    removed_sheets = set(old_sh) - set(new_sh)
    drop_sheets = to_repars | removed_sheets
    kept_records = [r for r in old_records if r.source_sheet not in drop_sheets]
    kept_warns = [w for w in old_warns if w.source_sheet not in drop_sheets]

    if progress_callback:
        progress_callback(0.05, f"Re-parsing {len(to_repars)} changed sheet(s) in {source_path.name}")

    def reparsing_progress(fraction: float, message: str) -> None:
        if progress_callback:
            progress_callback(0.05 + 0.95 * fraction, message)

    new_records, new_warns, new_health = process_workbook_bytes(
        source_path,
        workbook_bytes,
        progress_callback=reparsing_progress if progress_callback else None,
        should_cancel=should_cancel,
        preview_callback=None,
        restrict_to_sheet_names=to_repars,
    )
    if should_cancel():
        return None

    merged_records = kept_records + new_records
    merged_warns = kept_warns + new_warns
    merged_health = new_health or old_health
    return merged_records, merged_warns, merged_health, new_sh, new_fp


def save_cached_daily_records(
    cache_dir: Path,
    source_path: Path,
    file_hash: str,
    records: list[DailyRecord],
    parse_warnings: list[SheetParseWarning] | None = None,
    workbook_health: list[WorkbookHealthItem] | None = None,
    sheet_hashes: dict[str, str] | None = None,
) -> None:
    cache_path = cache_file_path(cache_dir, source_path)
    payload: dict[str, object] = {
        "source_path": str(source_path),
        "file_hash": file_hash,
        "cached_at": datetime.now().isoformat(),
        "records": [serialize_daily_record(record) for record in records],
        "parse_warnings": [serialize_sheet_parse_warning(warning) for warning in (parse_warnings or [])],
        "workbook_health": [serialize_workbook_health_item(item) for item in (workbook_health or [])],
    }
    if sheet_hashes:
        payload["sheet_hashes"] = dict(sorted(sheet_hashes.items(), key=lambda item: item[0].casefold()))
    cache_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def is_alert_triggered(st: float, ot: float, dt: float, profile: FormattingProfile) -> bool:
    return (
        (profile.st_threshold is not None and st > profile.st_threshold)
        or (profile.ot_threshold is not None and ot > profile.ot_threshold)
    )


def build_error_findings(records: Iterable[DailyRecord], profile: FormattingProfile) -> list[ErrorFinding]:
    employee_week_records: dict[tuple[str, date], list[DailyRecord]] = {}
    for record in records:
        week_start = monday_week_start(record.work_date)
        employee_week_records.setdefault((record.employee, week_start), []).append(record)

    findings: list[ErrorFinding] = []
    for (employee, week_start), week_records in sorted(employee_week_records.items(), key=lambda item: (item[0][1], item[0][0])):
        week_end = week_start + timedelta(days=6)
        daily_totals: list[dict[str, object]] = []
        by_day: dict[date, dict[str, object]] = {}
        for record in sorted(week_records, key=lambda item: (item.work_date, item.source_file, item.source_sheet)):
            bucket = by_day.setdefault(
                record.work_date,
                {
                    "date": record.work_date,
                    "st": 0.0,
                    "ot": 0.0,
                    "dt": 0.0,
                    "sources": set(),
                },
            )
            bucket["st"] = float(bucket["st"]) + record.st
            bucket["ot"] = float(bucket["ot"]) + record.ot
            bucket["dt"] = float(bucket["dt"]) + record.dt
            cast_sources = bucket["sources"]
            if isinstance(cast_sources, set):
                cast_sources.add(record.source_file)
        for work_date in sorted(by_day):
            daily_totals.append(by_day[work_date])

        weekly_st = round(sum(float(day["st"]) for day in daily_totals), 2)
        weekly_ot = round(sum(float(day["ot"]) for day in daily_totals), 2)
        source_files = ", ".join(sorted({source for day in daily_totals for source in day["sources"]}))  # type: ignore[index]
        breakdown = " | ".join(
            (
                f"{day['date'].isoformat()} "
                f"ST {fmt_hours(float(day['st']))} "
                f"OT {fmt_hours(float(day['ot']))} "
                f"DT {fmt_hours(float(day['dt']))}"
            )
            for day in daily_totals
        )

        for hour_type, threshold, actual_total in (
            ("ST", profile.st_threshold, weekly_st),
            ("OT", profile.ot_threshold, weekly_ot),
        ):
            if threshold is None or actual_total <= threshold:
                continue
            cumulative = 0.0
            trigger_day = daily_totals[-1]
            for day in daily_totals:
                cumulative += float(day[hour_type.lower()])
                if cumulative > threshold:
                    trigger_day = day
                    break
            delta = round(actual_total - threshold, 2)
            reason = (
                f"{hour_type} weekly total {fmt_hours(actual_total)} exceeded the limit "
                f"{fmt_hours(threshold)} by {fmt_hours(delta)}."
            )
            findings.append(
                ErrorFinding(
                    employee=employee,
                    week_start=week_start,
                    week_end=week_end,
                    hour_type=hour_type,
                    threshold=threshold,
                    actual_total=actual_total,
                    delta=delta,
                    trigger_date=trigger_day["date"],  # type: ignore[arg-type]
                    trigger_day_st=round(float(trigger_day["st"]), 2),
                    trigger_day_ot=round(float(trigger_day["ot"]), 2),
                    trigger_day_dt=round(float(trigger_day["dt"]), 2),
                    source_files=source_files,
                    reason=reason,
                    breakdown=breakdown,
                )
            )

        for day in daily_totals:
            day_st = round(float(day["st"]), 2)
            day_ot = round(float(day["ot"]), 2)
            day_dt = round(float(day["dt"]), 2)
            day_total = round(day_st + day_ot + day_dt, 2)
            trigger_date = day["date"]  # type: ignore[assignment]
            day_sources = ", ".join(sorted(day["sources"]))  # type: ignore[arg-type]
            day_breakdown = (
                f"{trigger_date.isoformat()} "
                f"ST {fmt_hours(day_st)} OT {fmt_hours(day_ot)} DT {fmt_hours(day_dt)} Total {fmt_hours(day_total)}"
            )

            if profile.daily_st_threshold is not None and day_st > profile.daily_st_threshold:
                delta = round(day_st - profile.daily_st_threshold, 2)
                findings.append(
                    ErrorFinding(
                        employee=employee,
                        week_start=week_start,
                        week_end=week_end,
                        hour_type="Daily ST",
                        threshold=profile.daily_st_threshold,
                        actual_total=day_st,
                        delta=delta,
                        trigger_date=trigger_date,
                        trigger_day_st=day_st,
                        trigger_day_ot=day_ot,
                        trigger_day_dt=day_dt,
                        source_files=day_sources,
                        reason=(
                            f"Daily ST total {fmt_hours(day_st)} exceeded the limit "
                            f"{fmt_hours(profile.daily_st_threshold)} by {fmt_hours(delta)}."
                        ),
                        breakdown=day_breakdown,
                    )
                )

            if profile.max_hours_per_day is not None and day_total > profile.max_hours_per_day:
                delta = round(day_total - profile.max_hours_per_day, 2)
                findings.append(
                    ErrorFinding(
                        employee=employee,
                        week_start=week_start,
                        week_end=week_end,
                        hour_type="Daily Total",
                        threshold=profile.max_hours_per_day,
                        actual_total=day_total,
                        delta=delta,
                        trigger_date=trigger_date,
                        trigger_day_st=day_st,
                        trigger_day_ot=day_ot,
                        trigger_day_dt=day_dt,
                        source_files=day_sources,
                        reason=(
                            f"Daily total {fmt_hours(day_total)} exceeded the max-hours limit "
                            f"{fmt_hours(profile.max_hours_per_day)} by {fmt_hours(delta)}."
                        ),
                        breakdown=day_breakdown,
                    )
                )
    return findings


def filter_employee_names(
    employee_names: Iterable[str],
    filter_selection: FilterSelection,
    employee_groups: dict[str, list[str]],
) -> set[str]:
    mode = filter_selection.mode
    value = filter_selection.value.strip()
    names = {employee for employee in employee_names if employee.strip()}
    if filter_selection.values:
        return {employee for employee in filter_selection.values if employee in names}
    if mode == "employee" and value:
        return {value} if value in names else set()
    if mode == "group" and value:
        return {employee for employee in employee_groups.get(value, []) if employee in names}
    return names


def normalize_person_name(name: str) -> str:
    return " ".join(name.split()).casefold()


def typo_warning_key(employee: str, similar_employee: str) -> str:
    return f"{normalize_person_name(employee)}|{normalize_person_name(similar_employee)}"


def find_potential_name_typos(
    unresolved_names: Iterable[str],
    all_employee_names: Iterable[str],
    daily_records: Iterable[DailyRecord],
    similarity_threshold: float = 0.82,
) -> list[NameTypoWarning]:
    all_names = sorted({name for name in all_employee_names if name.strip()})
    normalized_names = {name: normalize_person_name(name) for name in all_names}
    warnings: list[NameTypoWarning] = []

    for unresolved in sorted({name for name in unresolved_names if name.strip()}):
        unresolved_normalized = normalize_person_name(unresolved)
        best_match = ""
        best_similarity = 0.0
        for candidate in all_names:
            if candidate == unresolved:
                continue
            similarity = difflib.SequenceMatcher(None, unresolved_normalized, normalized_names[candidate]).ratio()
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = candidate

        if not best_match or best_similarity < similarity_threshold:
            continue

        locations = []
        for record in daily_records:
            if record.employee != unresolved:
                continue
            locations.append(f"{record.work_date.isoformat()} | {record.source_sheet} | {record.source_file}")

        warnings.append(
            NameTypoWarning(
                employee=unresolved,
                similar_employee=best_match,
                similarity=best_similarity,
                locations=locations,
            )
        )
    return warnings


def find_similar_employee_name_pairs(
    all_employee_names: Iterable[str],
    daily_records: Iterable[DailyRecord],
    similarity_threshold: float = 0.82,
) -> list[NameTypoWarning]:
    all_names = sorted({name for name in all_employee_names if name.strip()})
    normalized_names = {name: normalize_person_name(name) for name in all_names}
    warnings: list[NameTypoWarning] = []
    seen_pairs: set[tuple[str, str]] = set()

    for index, left_name in enumerate(all_names):
        for right_name in all_names[index + 1:]:
            similarity = difflib.SequenceMatcher(None, normalized_names[left_name], normalized_names[right_name]).ratio()
            if similarity < similarity_threshold:
                continue
            pair_key = tuple(sorted((left_name, right_name), key=str.casefold))
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)
            locations = []
            for record in daily_records:
                if record.employee not in pair_key:
                    continue
                locations.append(f"{record.employee} | {record.work_date.isoformat()} | {record.source_sheet} | {record.source_file}")
            warnings.append(
                NameTypoWarning(
                    employee=pair_key[1],
                    similar_employee=pair_key[0],
                    similarity=similarity,
                    locations=locations,
                )
            )
    return warnings


def extract_smtp_address(recipient) -> str:
    try:
        address_entry = recipient.AddressEntry
    except Exception:
        address_entry = None

    if address_entry is not None:
        try:
            exchange_user = address_entry.GetExchangeUser()
            if exchange_user and getattr(exchange_user, "PrimarySmtpAddress", ""):
                return str(exchange_user.PrimarySmtpAddress).strip()
        except Exception:
            pass
        try:
            exchange_dist = address_entry.GetExchangeDistributionList()
            if exchange_dist and getattr(exchange_dist, "PrimarySmtpAddress", ""):
                return str(exchange_dist.PrimarySmtpAddress).strip()
        except Exception:
            pass

    address = getattr(recipient, "Address", "")
    return str(address).strip()


def lookup_outlook_emails(
    employee_names: Iterable[str],
    should_cancel: Callable[[], bool] | None = None,
) -> dict[str, str]:
    if pythoncom is None or win32com is None:
        raise RuntimeError("Outlook lookup requires pywin32 and desktop Outlook.")

    names = [name for name in employee_names if name.strip()]
    if not names:
        return {}
    check_cancel = should_cancel or (lambda: False)

    pythoncom.CoInitialize()
    try:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
        except Exception as exc:
            raise RuntimeError("Could not open Outlook to query email addresses.") from exc

        results: dict[str, str] = {}
        for employee in names:
            if check_cancel():
                raise OperationCancelled("Cancelled Outlook email sync.")
            try:
                recipient = namespace.CreateRecipient(employee)
                recipient.Resolve()
            except Exception:
                continue
            if not getattr(recipient, "Resolved", False):
                continue
            email = extract_smtp_address(recipient)
            if email:
                results[employee] = email
        return results
    finally:
        pythoncom.CoUninitialize()


def format_week_label(week_start: date, week_end: date) -> str:
    return f"{week_start.isoformat()} to {week_end.isoformat()}"


def collect_week_ranges(records: Iterable[DailyRecord]) -> list[tuple[date, date]]:
    week_starts = sorted({monday_week_start(record.work_date) for record in records})
    return [(week_start, week_start + timedelta(days=6)) for week_start in week_starts]


def records_for_week(records: Iterable[DailyRecord], week_start: date) -> list[DailyRecord]:
    week_end = week_start + timedelta(days=6)
    return [record for record in records if week_start <= record.work_date <= week_end]


def extract_pf_number(source_name: str) -> str | None:
    match = PF_PATTERN.search(source_name)
    if not match:
        return None
    return match.group(1).upper()


def pf_numbers_for_records(records: Iterable[DailyRecord]) -> str:
    pf_numbers = sorted({pf_number for record in records if (pf_number := extract_pf_number(record.source_file))})
    return ", ".join(pf_numbers)


def build_email_draft_requests(
    records: Iterable[DailyRecord],
    employee_emails: dict[str, str],
    week_start: date,
) -> list[EmailDraftRequest]:
    grouped: dict[str, list[DailyRecord]] = {}
    week_records = records_for_week(records, week_start)
    for record in week_records:
        grouped.setdefault(record.employee, []).append(record)

    requests: list[EmailDraftRequest] = []
    week_end = week_start + timedelta(days=6)
    for employee in sorted(grouped):
        requests.append(
            EmailDraftRequest(
                employee=employee,
                email=employee_emails.get(employee, "").strip(),
                week_start=week_start,
                week_end=week_end,
                records=sorted(grouped[employee], key=lambda item: (item.work_date, item.source_file, item.source_sheet)),
            )
        )
    return requests


def format_email_subject(
    template: str,
    employee: str,
    week_start: date,
    week_end: date,
    records: Iterable[DailyRecord] | None = None,
) -> str:
    subject_template = template.strip() or "Timesheet Reminder - Week of {week_start}"
    first_name = employee.strip().split()[0] if employee.strip() else employee
    pf_numbers = pf_numbers_for_records(records or [])
    if pf_numbers and "{pf_numbers}" not in subject_template:
        subject_template = f"{subject_template} - {pf_numbers}"
    return subject_template.format(
        employee=employee,
        first_name=first_name,
        week_start=week_start.isoformat(),
        week_end=week_end.isoformat(),
        pf_numbers=pf_numbers,
    )


def build_hours_table_html(records: list[DailyRecord]) -> str:
    total_st = round(sum(record.st for record in records), 2)
    total_ot = round(sum(record.ot for record in records), 2)
    total_dt = round(sum(record.dt for record in records), 2)
    total_hours = round(total_st + total_ot + total_dt, 2)

    rows = []
    for record in records:
        rows.append(
            "<tr>"
            f"<td>{html.escape(record.work_date.isoformat())}</td>"
            f"<td>{html.escape(record.work_date.strftime('%A'))}</td>"
            f"<td>{html.escape(record.source_file)}</td>"
            f"<td>{html.escape(record.source_sheet)}</td>"
            f"<td>{html.escape(fmt_hours(record.st))}</td>"
            f"<td>{html.escape(fmt_hours(record.ot))}</td>"
            f"<td>{html.escape(fmt_hours(record.dt))}</td>"
            f"<td>{html.escape(fmt_hours(record.total))}</td>"
            "</tr>"
        )
    rows_html = "".join(rows)

    return (
        "<table border='1' cellspacing='0' cellpadding='4' style='border-collapse:collapse;'>"
        "<thead><tr>"
        "<th>Date</th><th>Day</th><th>Source File</th><th>Source Sheet</th><th>ST</th><th>OT</th><th>DT</th><th>Total</th>"
        "</tr></thead>"
        f"<tbody>{rows_html}</tbody>"
        "<tfoot><tr>"
        "<th colspan='4'>Week Total</th>"
        f"<th>{html.escape(fmt_hours(total_st))}</th>"
        f"<th>{html.escape(fmt_hours(total_ot))}</th>"
        f"<th>{html.escape(fmt_hours(total_dt))}</th>"
        f"<th>{html.escape(fmt_hours(total_hours))}</th>"
        "</tr></tfoot>"
        "</table>"
    )


def build_email_html(
    employee: str,
    week_start: date,
    week_end: date,
    records: list[DailyRecord],
    body_template: str,
) -> str:
    first_name = employee.strip().split()[0] if employee.strip() else employee
    template = body_template.strip() or default_email_body_template()
    return template.format(
        employee=html.escape(employee),
        first_name=html.escape(first_name),
        week_start=html.escape(week_start.isoformat()),
        week_end=html.escape(week_end.isoformat()),
        hours_table=build_hours_table_html(records),
    )


def build_bug_report_html(
    current_profile_name: str,
    app_root: Path,
    snapshot_path: Path,
    loaded_sources: list[Path],
    cache_status_by_path: dict[Path, str],
) -> str:
    source_items = "".join(
        f"<li>{html.escape(str(path))} ({html.escape(cache_status_by_path.get(path, 'Unknown'))})</li>"
        for path in loaded_sources
    ) or "<li>None loaded</li>"
    return (
        "<p>Hi Lochlan,</p>"
        f"<p>Please find a bug report for {DISPLAY_APP_NAME} below.</p>"
        "<p><strong>Summary:</strong><br>[Describe the bug briefly]</p>"
        "<p><strong>What I was trying to do:</strong><br>[Describe the task]</p>"
        "<p><strong>What happened:</strong><br>[Describe the actual result or error]</p>"
        "<p><strong>What I expected to happen:</strong><br>[Describe the expected result]</p>"
        "<p><strong>Steps to reproduce:</strong><br>"
        "1. [Step one]<br>"
        "2. [Step two]<br>"
        "3. [Step three]</p>"
        "<p><strong>When it happened:</strong><br>[Date / time]</p>"
        "<p><strong>Relevant screenshots or notes:</strong><br>[Add anything helpful here]</p>"
        "<hr>"
        "<p><strong>Attached diagnostic snapshot:</strong><br>"
        f"{html.escape(str(snapshot_path))}</p>"
        "<p><strong>Current profile:</strong><br>"
        f"{html.escape(current_profile_name)}</p>"
        "<p><strong>App data folder:</strong><br>"
        f"{html.escape(str(app_root))}</p>"
        "<p><strong>Loaded DSS files:</strong></p>"
        f"<ul>{source_items}</ul>"
    )


# OlAttachmentType.olByValue — file is embedded; Outlook is picky about Source paths (Unicode, length).
_OUTLOOK_ATTACHMENT_BY_VALUE = 1


def _bug_report_attachment_strings_to_try(snapshot_path: Path) -> tuple[list[str], list[Path]]:
    """Return path strings to try with ``Attachments.Add``, plus temp copies to delete after ``Save()``."""
    cleanup: list[Path] = []
    resolved = snapshot_path.expanduser().resolve()
    if not resolved.is_file():
        raise OSError(f"Bug report snapshot is not readable: {resolved}")
    ordered: list[str] = []
    seen_norm: set[str] = set()

    def push(path_str: str) -> None:
        key = os.path.normcase(path_str)
        if key in seen_norm:
            return
        seen_norm.add(key)
        ordered.append(path_str)

    primary = os.path.normpath(str(resolved))
    if os.name == "nt":
        try:
            import win32api

            primary = win32api.GetShortPathName(primary)
        except Exception:
            pass
    push(primary)

    temp_path = Path(tempfile.gettempdir()) / f"dssbug_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.getpid()}.json"
    shutil.copy2(resolved, temp_path)
    cleanup.append(temp_path)
    alt = os.path.normpath(str(temp_path))
    if os.name == "nt":
        try:
            import win32api

            alt = win32api.GetShortPathName(alt)
        except Exception:
            pass
    push(alt)

    if os.name == "nt":
        for candidate in list(ordered):
            if "\\" in candidate:
                push(candidate.replace("\\", "/"))

    return ordered, cleanup


def create_bug_report_draft(
    recipient_email: str,
    subject: str,
    html_body: str,
    attachment_path: Path | None = None,
) -> str | None:
    """Create a saved Outlook draft. Returns a warning string if the snapshot could not be attached."""
    if pythoncom is None or win32com is None:
        raise RuntimeError("Bug report draft creation requires desktop Outlook with pywin32 available.")

    pythoncom.CoInitialize()
    cleanup_paths: list[Path] = []
    attachment_warning: str | None = None
    try:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
        except Exception as exc:
            raise RuntimeError("Could not open Outlook to create the bug report draft.") from exc

        mail_item = outlook.CreateItem(0)
        mail_item.To = recipient_email
        mail_item.Subject = subject

        if attachment_path is not None and Path(attachment_path).exists():
            path_strings: list[str] = []
            try:
                path_strings, cleanup_paths = _bug_report_attachment_strings_to_try(Path(attachment_path))
            except OSError as exc:
                attachment_warning = f"Could not prepare the diagnostic file for Outlook: {exc}"
            last_error: str | None = None
            attached = False
            for candidate in path_strings:
                try:
                    mail_item.Attachments.Add(candidate, _OUTLOOK_ATTACHMENT_BY_VALUE)
                    attached = True
                    break
                except Exception as exc:
                    last_error = str(exc)
            if path_strings and not attached:
                attachment_warning = (
                    "Outlook could not attach the diagnostic snapshot (path or permission issue). "
                    "The draft was saved without an attachment; use Export Diagnostic Snapshot and attach that file manually.\n\n"
                    f"Technical detail: {last_error or 'Unknown COM error'}"
                )

        mail_item.HTMLBody = html_body
        mail_item.Save()
    finally:
        for temp in cleanup_paths:
            try:
                temp.unlink(missing_ok=True)
            except OSError:
                pass
        pythoncom.CoUninitialize()

    return attachment_warning


def create_outlook_drafts(
    requests: list[EmailDraftRequest],
    subject_template: str,
    body_template: str,
) -> tuple[int, list[str]]:
    if pythoncom is None or win32com is None:
        raise RuntimeError("Outlook draft creation requires desktop Outlook with pywin32 available.")

    created = 0
    skipped: list[str] = []
    pythoncom.CoInitialize()
    try:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
        except Exception as exc:
            raise RuntimeError("Could not open Outlook to create draft emails.") from exc

        for request in requests:
            if not request.email:
                skipped.append(request.employee)
                continue
            mail_item = outlook.CreateItem(0)
            mail_item.To = request.email
            mail_item.Subject = format_email_subject(
                subject_template,
                request.employee,
                request.week_start,
                request.week_end,
                request.records,
            )
            mail_item.HTMLBody = build_email_html(
                request.employee,
                request.week_start,
                request.week_end,
                request.records,
                body_template,
            )
            mail_item.Save()
            created += 1
    finally:
        pythoncom.CoUninitialize()
    return created, skipped


def select_preferred_dated_sheets(sheet_names: Iterable[str]) -> list[tuple[date, str]]:
    best_by_date: dict[date, tuple[int, str]] = {}
    for sheet_name in sheet_names:
        sheet_date = parse_sheet_date(sheet_name)
        if not sheet_date:
            continue
        revision = parse_sheet_revision(sheet_name)
        current = best_by_date.get(sheet_date)
        if current is None or revision > current[0] or (revision == current[0] and sheet_name.strip() > current[1].strip()):
            best_by_date[sheet_date] = (revision, sheet_name)
    return [(sheet_date, best_by_date[sheet_date][1]) for sheet_date in sorted(best_by_date)]


def is_text_name(value: object) -> bool:
    return isinstance(value, str) and bool(value.strip())


def to_number(value: object) -> float:
    if value in (None, "") or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        return round(float(value), 2)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0.0
        try:
            return round(float(Decimal(text)), 2)
        except (InvalidOperation, ValueError):
            return 0.0
    return 0.0


def detect_employee_name(ws, start_row: int, column_indexes: Iterable[int]) -> str | None:
    for col_idx in reversed(tuple(column_indexes)):
        for row_idx in range(start_row, start_row + 3):
            value = ws.cell(row=row_idx, column=col_idx).value
            if is_text_name(value):
                return value.strip()
    return None


def sum_hour_row(ws, row_idx: int, hour_columns: Iterable[int]) -> float:
    return round(sum(to_number(ws.cell(row=row_idx, column=col_idx).value) for col_idx in hour_columns), 2)


def build_source_ranges(label: str, start_row: int, name_cols: tuple[int, ...], hour_cols: tuple[int, ...]) -> str:
    return (
        f"{label} name {get_column_letter(name_cols[0])}{start_row}:{get_column_letter(name_cols[-1])}{start_row + 2}; "
        f"{label} hours {get_column_letter(hour_cols[0])}{start_row}:{get_column_letter(hour_cols[-1])}{start_row + 2}"
    )


def find_name_cell(ws, start_row: int, column_indexes: Iterable[int]) -> tuple[str, str] | tuple[None, None]:
    for col_idx in reversed(tuple(column_indexes)):
        for row_idx in range(start_row, start_row + 3):
            value = ws.cell(row=row_idx, column=col_idx).value
            if is_text_name(value):
                return value.strip(), f"{get_column_letter(col_idx)}{row_idx}"
    return None, None


def process_workbook_bytes(
    source_path: Path,
    workbook_bytes: bytes,
    progress_callback: Callable[[float, str], None] | None = None,
    should_cancel: Callable[[], bool] | None = None,
    preview_callback: Callable[[list[DailyRecord], list[SheetParseWarning], list[WorkbookHealthItem], str], None] | None = None,
    restrict_to_sheet_names: frozenset[str] | None = None,
) -> tuple[list[DailyRecord], list[SheetParseWarning], list[WorkbookHealthItem]]:
    emit_progress = progress_callback or (lambda _fraction, _message: None)
    check_cancel = should_cancel or (lambda: False)

    def raise_if_cancelled() -> None:
        if check_cancel():
            raise OperationCancelled(f"Cancelled {source_path.name}")

    raise_if_cancelled()
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    records: list[DailyRecord] = []
    warnings: list[SheetParseWarning] = []
    health: list[WorkbookHealthItem] = []

    try:
        raise_if_cancelled()
        emit_progress(0.1, f"Opened {source_path.name}")
        selected_sheets = sorted(
            select_preferred_dated_sheets(workbook.sheetnames),
            key=lambda item: (item[0], item[1].strip().casefold()),
            reverse=True,
        )
        emit_progress(0.2, f"Selected sheets for {source_path.name}")
        dated_sheets = [sheet_name for sheet_name in workbook.sheetnames if parse_sheet_date(sheet_name) is not None]
        if not dated_sheets:
            health.append(
                WorkbookHealthItem(
                    source_file=source_path.name,
                    status="Warning",
                    details="No dated DSS sheets matching YYYY-MM-DD were found.",
                )
            )
        if "onedrive" in str(source_path).lower() or "sharepoint" in str(source_path).lower():
            health.append(
                WorkbookHealthItem(
                    source_file=source_path.name,
                    status="Info",
                    details="Workbook is in a OneDrive / SharePoint path and may be temporarily locked during sync or save.",
                )
            )

        by_date: dict[date, list[str]] = {}
        for sheet_name in dated_sheets:
            sheet_date = parse_sheet_date(sheet_name)
            if sheet_date is not None:
                by_date.setdefault(sheet_date, []).append(sheet_name)
        preferred_sheet_by_date = dict(selected_sheets)
        if restrict_to_sheet_names is None:
            for sheet_date, sheet_names in sorted(by_date.items()):
                if len(sheet_names) > 1:
                    health.append(
                        WorkbookHealthItem(
                            source_file=source_path.name,
                            status="Info",
                            details=(
                                f"{sheet_date.isoformat()} has {len(sheet_names)} revision candidate sheets. "
                                f"Using '{preferred_sheet_by_date.get(sheet_date, '')}'."
                            ),
                        )
                    )

        if restrict_to_sheet_names is None:
            az2_targets = dated_sheets
        else:
            az2_targets = [name for name in dated_sheets if name in restrict_to_sheet_names]
        for sheet_name in az2_targets:
            sheet_date_az = parse_sheet_date(sheet_name)
            if sheet_date_az is None:
                continue
            ws_az = workbook[sheet_name]
            az2_raw = ws_az.cell(row=DSS_HASH_AZ2_ROW, column=DSS_HASH_AZ2_COL).value
            ok_az, detail_az = az2_revision_matches_sheet_name(sheet_name, az2_raw)
            if not ok_az and detail_az:
                extra = ""
                if az2_raw is not None and str(az2_raw).strip() != "":
                    extra = f" (AZ2 raw value: {az2_raw!r})"
                else:
                    extra = " (AZ2 is blank or non-numeric)"
                warnings.append(
                    SheetParseWarning(
                        source_file=source_path.name,
                        source_sheet=sheet_name,
                        work_date=sheet_date_az.isoformat(),
                        issue="Revision Indicator AZ2 Mismatch",
                        details=detail_az + extra,
                    )
                )

        total_sheets = max(len(selected_sheets), 1)
        preview_emitted = False
        recent_week_starts = sorted({monday_week_start(sheet_date) for sheet_date, _sheet_name in selected_sheets}, reverse=True)
        preview_cutoff = recent_week_starts[1] if len(recent_week_starts) >= 2 else (recent_week_starts[0] if recent_week_starts else None)
        preview_sheet_count = sum(
            1
            for sheet_date, _sheet_name in selected_sheets
            if preview_cutoff is not None and monday_week_start(sheet_date) >= preview_cutoff
        )
        for sheet_index, (sheet_date, sheet_name) in enumerate(selected_sheets, start=1):
            raise_if_cancelled()
            if restrict_to_sheet_names is not None and sheet_name not in restrict_to_sheet_names:
                continue
            ws = workbook[sheet_name]
            seen_names: dict[str, int] = {}
            for start_row in BLOCK_START_ROWS:
                for label, name_cols, hour_cols in (
                    ("Left", LEFT_NAME_COLS, LEFT_HOUR_COLS),
                    ("Right", RIGHT_NAME_COLS, RIGHT_HOUR_COLS),
                ):
                    raise_if_cancelled()
                    employee, cell_ref = find_name_cell(ws, start_row, name_cols)
                    st = sum_hour_row(ws, start_row, hour_cols)
                    ot = sum_hour_row(ws, start_row + 1, hour_cols)
                    dt = sum_hour_row(ws, start_row + 2, hour_cols)
                    total = round(st + ot + dt, 2)
                    if not employee:
                        if total > 0:
                            warnings.append(
                                SheetParseWarning(
                                    source_file=source_path.name,
                                    source_sheet=sheet_name,
                                    work_date=sheet_date.isoformat(),
                                    issue="Hours Without Name",
                                    details=f"{label} block at row {start_row} has {fmt_hours(total)} hours but no employee name.",
                                )
                            )
                        continue

                    seen_names[employee] = seen_names.get(employee, 0) + 1
                    if total > 24:
                        warnings.append(
                            SheetParseWarning(
                                source_file=source_path.name,
                                source_sheet=sheet_name,
                                work_date=sheet_date.isoformat(),
                                issue="Suspicious Total",
                                details=(
                                    f"{employee} in {label} block at row {start_row} totals {fmt_hours(total)} hours "
                                    f"(name cell {cell_ref})."
                                ),
                            )
                        )
                    records.append(
                        DailyRecord(
                            source_path=source_path,
                            source_file=source_path.name,
                            work_date=sheet_date,
                            source_sheet=sheet_name,
                            employee=employee,
                            st=st,
                            ot=ot,
                            dt=dt,
                            source_ranges=build_source_ranges(label, start_row, tuple(name_cols), tuple(hour_cols)),
                        )
                    )

            for employee, count in sorted(seen_names.items()):
                if count > 1:
                    warnings.append(
                        SheetParseWarning(
                            source_file=source_path.name,
                            source_sheet=sheet_name,
                            work_date=sheet_date.isoformat(),
                            issue="Duplicate Employee",
                            details=f"{employee} appears {count} times on the same sheet.",
                        )
                    )
            pf_label = extract_pf_identifier(source_path.name)
            emit_progress(
                0.2 + (0.75 * sheet_index / total_sheets),
                f"Processed {pf_label} — {sheet_name}",
            )
            if (
                preview_callback is not None
                and not preview_emitted
                and preview_sheet_count > 0
                and sheet_index >= preview_sheet_count
            ):
                preview_callback(
                    list(records),
                    list(warnings),
                    list(health),
                    "Showing the most recent one to two weeks while older DSS sheets continue loading.",
                )
                preview_emitted = True
    finally:
        workbook.close()

    if not health and restrict_to_sheet_names is None:
        health.append(
            WorkbookHealthItem(
                source_file=source_path.name,
                status="OK",
                details="No workbook health issues detected from the current workbook structure.",
            )
        )
    emit_progress(1.0, f"Finished {source_path.name}")
    return records, warnings, health


def analyze_workbook_bytes(
    source_path: Path,
    workbook_bytes: bytes,
    progress_callback: Callable[[float, str], None] | None = None,
) -> tuple[list[SheetParseWarning], list[WorkbookHealthItem]]:
    _records, warnings, health = process_workbook_bytes(
        source_path,
        workbook_bytes,
        progress_callback=progress_callback,
    )
    return warnings, health


def read_source_bytes(source_path: Path) -> bytes:
    try:
        return source_path.read_bytes()
    except PermissionError as exc:
        workbook_bytes = load_open_excel_workbook_copy_bytes(source_path)
        if workbook_bytes is not None:
            return workbook_bytes
        raise PermissionError(build_permission_denied_message(source_path)) from exc
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Workbook not found:\n{source_path}") from exc

def load_source_workbook(source_path: Path):
    return load_workbook(io.BytesIO(read_source_bytes(source_path)), data_only=True, read_only=True)


def compute_bytes_hash(workbook_bytes: bytes, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.blake2b(digest_size=16)
    for offset in range(0, len(workbook_bytes), chunk_size):
        digest.update(workbook_bytes[offset:offset + chunk_size])
    return digest.hexdigest()


def normalize_workbook_member_bytes(member_name: str, member_bytes: bytes) -> bytes:
    normalized_name = member_name.replace("\\", "/")
    if normalized_name == "xl/workbook.xml":
        return WORKBOOK_CALC_ID_PATTERN.sub(b"", member_bytes)
    return member_bytes


def cell_reference_to_position(cell_ref: str) -> tuple[int, int] | None:
    match = re.fullmatch(r"([A-Z]+)(\d+)", cell_ref.upper())
    if not match:
        return None
    letters, row_text = match.groups()
    column = 0
    for character in letters:
        column = (column * 26) + (ord(character) - 64)
    return column, int(row_text)


def normalize_hash_cell_value(raw_value: str, cell_type: str) -> str:
    value = raw_value.strip()
    if not value:
        return ""
    if cell_type in {"s", "str", "inlineStr"}:
        return " ".join(value.split())
    try:
        normalized_decimal = Decimal(value)
    except (InvalidOperation, ValueError):
        return value
    normalized = normalized_decimal.normalize()
    if normalized == normalized.to_integral():
        return format(normalized.quantize(Decimal("1")), "f")
    return format(normalized, "f")


def parse_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        payload = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(payload)
    values: list[str] = []
    for si in root.findall(f"{{{EXCEL_MAIN_NS}}}si"):
        text_parts = [node.text or "" for node in si.iterfind(f".//{{{EXCEL_MAIN_NS}}}t")]
        values.append("".join(text_parts))
    return values


def parse_workbook_sheet_targets(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(normalize_workbook_member_bytes("xl/workbook.xml", archive.read("xl/workbook.xml")))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets: dict[str, str] = {}
    for rel in rels_root.findall(f"{{{EXCEL_PACKAGE_REL_NS}}}Relationship"):
        rel_id = rel.get("Id", "").strip()
        target = rel.get("Target", "").strip()
        if rel_id and target:
            rel_targets[rel_id] = target
    sheet_targets: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{EXCEL_MAIN_NS}}}sheet"):
        sheet_name = str(sheet.get("name", "")).strip()
        rel_id = str(sheet.get(f"{{{EXCEL_REL_NS}}}id", "")).strip()
        target = rel_targets.get(rel_id, "")
        if not sheet_name or not target:
            continue
        normalized_target = target.replace("\\", "/").lstrip("/")
        if not normalized_target.startswith("xl/"):
            normalized_target = f"xl/{normalized_target}"
        sheet_targets[sheet_name] = normalized_target
    return sheet_targets


def worksheet_window_value_map(worksheet_root: ET.Element, shared_strings: list[str]) -> dict[str, str]:
    values: dict[str, str] = {}
    for cell in worksheet_root.findall(f".//{{{EXCEL_MAIN_NS}}}c"):
        cell_ref = str(cell.get("r", "")).strip()
        position = cell_reference_to_position(cell_ref)
        if position is None:
            continue
        column, row = position
        in_data_window = DSS_HASH_MIN_COL <= column <= DSS_HASH_MAX_COL and DSS_HASH_MIN_ROW <= row <= DSS_HASH_MAX_ROW
        is_az2_marker = column == DSS_HASH_AZ2_COL and row == DSS_HASH_AZ2_ROW
        if not (in_data_window or is_az2_marker):
            continue
        cell_type = str(cell.get("t", "")).strip()
        raw_value = ""
        if cell_type == "inlineStr":
            raw_value = "".join(node.text or "" for node in cell.iterfind(f".//{{{EXCEL_MAIN_NS}}}t"))
        else:
            value_node = cell.find(f"{{{EXCEL_MAIN_NS}}}v")
            raw_value = value_node.text if value_node is not None and value_node.text is not None else ""
            if cell_type == "s" and raw_value.strip():
                try:
                    raw_value = shared_strings[int(raw_value)]
                except (IndexError, ValueError):
                    pass
        normalized_value = normalize_hash_cell_value(raw_value, cell_type)
        if normalized_value:
            values[cell_ref.upper()] = normalized_value
    return values


def _digest_sheet_cells(sheet_date: date, sheet_name: str, cell_values: dict[str, str]) -> str:
    digest = hashlib.blake2b(digest_size=16)
    digest.update(sheet_date.isoformat().encode("utf-8"))
    digest.update(b"\0")
    digest.update(sheet_name.encode("utf-8"))
    digest.update(b"\0")
    for cell_ref in sorted(cell_values):
        digest.update(cell_ref.encode("utf-8"))
        digest.update(b"=")
        digest.update(cell_values[cell_ref].encode("utf-8"))
        digest.update(b"\0")
    return digest.hexdigest()


def compute_all_dated_sheet_hashes(workbook_bytes: bytes) -> dict[str, str]:
    """Per–dated-sheet digest of K25:AZ36 plus AZ2 (same window as parsing / cache invalidation)."""
    hashes: dict[str, str] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            sheet_targets = parse_workbook_sheet_targets(archive)
            shared_strings = parse_shared_strings(archive)
            for sheet_name in sheet_targets:
                sheet_date = parse_sheet_date(sheet_name)
                if sheet_date is None:
                    continue
                worksheet_target = sheet_targets.get(sheet_name, "")
                if not worksheet_target:
                    continue
                worksheet_root = ET.fromstring(archive.read(worksheet_target))
                cell_values = worksheet_window_value_map(worksheet_root, shared_strings)
                hashes[sheet_name] = _digest_sheet_cells(sheet_date, sheet_name, cell_values)
    except (OSError, ValueError, zipfile.BadZipFile, KeyError, ET.ParseError):
        return {}
    return hashes


def combine_sheet_hashes(sheet_hashes: dict[str, str]) -> str:
    if not sheet_hashes:
        return ""
    joined = "\0".join(f"{name}\x1e{value}" for name, value in sorted(sheet_hashes.items(), key=lambda item: item[0].casefold()))
    return hashlib.blake2b(joined.encode("utf-8"), digest_size=16).hexdigest()


def compute_dss_semantic_hash(workbook_bytes: bytes) -> str | None:
    sheet_hashes = compute_all_dated_sheet_hashes(workbook_bytes)
    if not sheet_hashes:
        return None
    return combine_sheet_hashes(sheet_hashes)


def compute_workbook_content_hash(workbook_bytes: bytes) -> str:
    semantic_hash = compute_dss_semantic_hash(workbook_bytes)
    if semantic_hash is not None:
        return semantic_hash
    try:
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            digest = hashlib.blake2b(digest_size=16)
            member_names = sorted(
                name
                for name in archive.namelist()
                if not name.endswith("/")
                and name not in {"docProps/core.xml", "docProps/app.xml", "docProps/custom.xml", "xl/calcChain.xml"}
            )
            if not member_names:
                return compute_bytes_hash(workbook_bytes)
            for member_name in member_names:
                digest.update(member_name.encode("utf-8"))
                digest.update(b"\0")
                digest.update(normalize_workbook_member_bytes(member_name, archive.read(member_name)))
                digest.update(b"\0")
            return digest.hexdigest()
    except (OSError, ValueError, zipfile.BadZipFile, KeyError):
        return compute_bytes_hash(workbook_bytes)


def load_open_excel_workbook_copy_bytes(source_path: Path) -> bytes | None:
    if pythoncom is None or win32com is None:
        return None

    temp_path: str | None = None
    pythoncom.CoInitialize()
    try:
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return None

        workbook = find_open_excel_workbook(excel, source_path)
        if workbook is None:
            return None

        fd, temp_path = tempfile.mkstemp(prefix="dss_hours_tracker_", suffix=".xlsx")
        os.close(fd)
        workbook.SaveCopyAs(temp_path)
        return Path(temp_path).read_bytes()
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except OSError:
                pass
        pythoncom.CoUninitialize()


def normalize_windows_path(value: str) -> str:
    return value.replace("/", "\\").strip().rstrip("\\").casefold()


def find_open_excel_workbook(excel_app, source_path: Path):
    target = normalize_windows_path(str(source_path))
    target_name = source_path.name.casefold()
    name_matches = []

    for workbook in excel_app.Workbooks:
        full_name = normalize_windows_path(str(workbook.FullName))
        if full_name == target:
            return workbook
        if Path(str(workbook.Name)).name.casefold() == target_name:
            name_matches.append(workbook)

    if len(name_matches) == 1:
        return name_matches[0]
    return None


def build_permission_denied_message(source_path: Path) -> str:
    lower_path = str(source_path).lower()
    if "onedrive" in lower_path or "sharepoint" in lower_path:
        return (
            "Windows denied access to this OneDrive / SharePoint workbook.\n\n"
            f"Path:\n{source_path}\n\n"
            "Things to try:\n"
            "- Keep the workbook open in desktop Excel, then press 'Update View' again.\n"
            "- In File Explorer, right-click the file and choose 'Always keep on this device'.\n"
            "- Wait for OneDrive to finish syncing, then try again.\n"
            "- If Excel just saved the workbook, wait a moment for OneDrive to release the file lock.\n"
            "- If needed, copy the workbook to a normal local folder and open that copy."
        )
    return (
        "Windows denied access to this workbook.\n\n"
        f"Path:\n{source_path}\n\n"
        "Make sure the file is not open in another app and that you have local read access."
    )


def parse_daily_records(workbook_path: Path) -> list[DailyRecord]:
    return parse_daily_records_from_bytes(workbook_path, read_source_bytes(workbook_path))


def parse_daily_records_from_bytes(
    workbook_path: Path,
    workbook_bytes: bytes,
    progress_callback: Callable[[float, str], None] | None = None,
    should_cancel: Callable[[], bool] | None = None,
) -> list[DailyRecord]:
    records, _warnings, _health = process_workbook_bytes(
        workbook_path,
        workbook_bytes,
        progress_callback=progress_callback,
        should_cancel=should_cancel,
    )
    return records


def monday_week_start(value: date) -> date:
    return value - timedelta(days=value.weekday())


def aggregate_weekly(records: Iterable[DailyRecord], combine_sources: bool = False) -> list[WeeklyRecord]:
    grouped: dict[tuple[str, date, str], dict[str, float]] = {}
    for record in records:
        week_start = monday_week_start(record.work_date)
        source_file = "All DSSs" if combine_sources else record.source_file
        bucket = grouped.setdefault((source_file, week_start, record.employee), {"ST": 0.0, "OT": 0.0, "DT": 0.0})
        bucket["ST"] += record.st
        bucket["OT"] += record.ot
        bucket["DT"] += record.dt

    results = []
    for (source_file, week_start, employee), totals in sorted(grouped.items(), key=lambda item: (item[0][1], item[0][0], item[0][2])):
        results.append(
            WeeklyRecord(
                source_file=source_file,
                week_start=week_start,
                week_end=week_start + timedelta(days=6),
                employee=employee,
                st=round(totals["ST"], 2),
                ot=round(totals["OT"], 2),
                dt=round(totals["DT"], 2),
            )
        )
    return results


def aggregate_daily(records: Iterable[DailyRecord], combine_sources: bool = False) -> list[DailySummaryRecord]:
    grouped: dict[tuple[str, date, str], dict[str, float]] = {}
    for record in records:
        source_file = "All DSSs" if combine_sources else record.source_file
        bucket = grouped.setdefault((source_file, record.work_date, record.employee), {"ST": 0.0, "OT": 0.0, "DT": 0.0})
        bucket["ST"] += record.st
        bucket["OT"] += record.ot
        bucket["DT"] += record.dt

    results: list[DailySummaryRecord] = []
    for (source_file, work_date, employee), totals in sorted(grouped.items(), key=lambda item: (item[0][1], item[0][0], item[0][2])):
        results.append(
            DailySummaryRecord(
                source_file=source_file,
                work_date=work_date,
                employee=employee,
                st=round(totals["ST"], 2),
                ot=round(totals["OT"], 2),
                dt=round(totals["DT"], 2),
            )
        )
    return results


def build_weekly_rollup(weekly_records: list[WeeklyRecord]) -> list[WeeklyRollupRow]:
    grouped: dict[tuple[str, date], list[WeeklyRecord]] = {}
    for record in weekly_records:
        grouped.setdefault((record.source_file, record.week_start), []).append(record)

    rows: list[WeeklyRollupRow] = []
    for source_file, week_start in sorted(grouped):
        week_entries = sorted(grouped[(source_file, week_start)], key=lambda item: item.employee)
        week_end = week_start + timedelta(days=6)
        crew_st = 0.0
        crew_ot = 0.0
        crew_dt = 0.0
        for record in week_entries:
            rows.append(
                WeeklyRollupRow(
                    source_file=source_file,
                    week_start=week_start,
                    week_end=week_end,
                    employee=record.employee,
                    st=record.st,
                    ot=record.ot,
                    dt=record.dt,
                    row_type="Employee",
                )
            )
            crew_st += record.st
            crew_ot += record.ot
            crew_dt += record.dt
        rows.append(
            WeeklyRollupRow(
                source_file=source_file,
                week_start=week_start,
                week_end=week_end,
                employee="Whole Crew",
                st=round(crew_st, 2),
                ot=round(crew_ot, 2),
                dt=round(crew_dt, 2),
                row_type="Crew Total",
            )
        )
    return rows


def build_daily_rollup(daily_summary: list[DailySummaryRecord]) -> list[DailyRollupRow]:
    grouped: dict[tuple[str, date], list[DailySummaryRecord]] = {}
    for record in daily_summary:
        grouped.setdefault((record.source_file, record.work_date), []).append(record)

    rows: list[DailyRollupRow] = []
    for source_file, work_date in sorted(grouped):
        day_entries = sorted(grouped[(source_file, work_date)], key=lambda item: item.employee)
        crew_st = 0.0
        crew_ot = 0.0
        crew_dt = 0.0
        for record in day_entries:
            rows.append(
                DailyRollupRow(
                    source_file=source_file,
                    work_date=work_date,
                    employee=record.employee,
                    st=record.st,
                    ot=record.ot,
                    dt=record.dt,
                    row_type="Employee",
                )
            )
            crew_st += record.st
            crew_ot += record.ot
            crew_dt += record.dt
        rows.append(
            DailyRollupRow(
                source_file=source_file,
                work_date=work_date,
                employee="Whole Crew",
                st=round(crew_st, 2),
                ot=round(crew_ot, 2),
                dt=round(crew_dt, 2),
                row_type="Crew Total",
            )
        )
    return rows


def build_week_totals(weekly_records: list[WeeklyRecord]) -> list[WeekTotalRow]:
    grouped: dict[date, dict[str, float]] = {}
    for record in weekly_records:
        bucket = grouped.setdefault(record.week_start, {"ST": 0.0, "OT": 0.0, "DT": 0.0})
        bucket["ST"] += record.st
        bucket["OT"] += record.ot
        bucket["DT"] += record.dt

    rows: list[WeekTotalRow] = []
    for week_start in sorted(grouped):
        totals = grouped[week_start]
        rows.append(
            WeekTotalRow(
                week_start=week_start,
                week_end=week_start + timedelta(days=6),
                st=round(totals["ST"], 2),
                ot=round(totals["OT"], 2),
                dt=round(totals["DT"], 2),
            )
        )
    return rows


def build_tracker_data(source_paths: list[Path], file_hashes: dict[Path, str], daily_records: list[DailyRecord]) -> TrackerData:
    return build_tracker_data_with_status(source_paths, file_hashes, [], source_paths, daily_records)


def build_tracker_data_with_status(
    source_paths: list[Path],
    file_hashes: dict[Path, str],
    reused_paths: list[Path],
    reloaded_paths: list[Path],
    daily_records: list[DailyRecord],
    cache_status_by_path: dict[Path, str] | None = None,
    parse_warnings: list[SheetParseWarning] | None = None,
    workbook_health: list[WorkbookHealthItem] | None = None,
) -> TrackerData:
    if not daily_records:
        raise ValueError("No daily DSS records were found in sheets named with YYYY-MM-DD.")
    weekly_summary = aggregate_weekly(daily_records, combine_sources=False)
    combined_weekly_summary = aggregate_weekly(daily_records, combine_sources=True)
    daily_summary = aggregate_daily(daily_records, combine_sources=False)
    combined_daily_summary = aggregate_daily(daily_records, combine_sources=True)
    return TrackerData(
        source_paths=source_paths,
        file_hashes=file_hashes,
        reused_paths=reused_paths,
        reloaded_paths=reloaded_paths,
        cache_status_by_path=cache_status_by_path or {},
        daily_records=sorted(daily_records, key=lambda item: (item.work_date, item.source_sheet, item.employee)),
        employee_names=sorted({record.employee for record in daily_records}),
        weekly_summary=weekly_summary,
        weekly_rollup=build_weekly_rollup(weekly_summary),
        daily_summary=daily_summary,
        daily_rollup=build_daily_rollup(daily_summary),
        week_totals=build_week_totals(combined_weekly_summary),
        combined_weekly_summary=combined_weekly_summary,
        combined_daily_summary=combined_daily_summary,
        parse_warnings=parse_warnings or [],
        workbook_health=workbook_health or [],
    )


def tracker_data_invalidated_for_cache_clear(data: TrackerData) -> TrackerData:
    """After on-disk cache files are removed, drop hashes and reuse flags so the next load does not memory-hit stale rows."""
    status_keys = {*data.source_paths, *data.cache_status_by_path}
    return replace(
        data,
        file_hashes={},
        reused_paths=[],
        cache_status_by_path={path: "Miss" for path in status_keys},
    )


def load_tracker_data(
    source_paths: Path | Iterable[Path],
    previous_data: TrackerData | None = None,
    progress_callback: Callable[[float, str], None] | None = None,
    partial_callback: Callable[[TrackerData, str], None] | None = None,
    cache_dir: Path | None = None,
    should_cancel: Callable[[], bool] | None = None,
) -> TrackerData:
    if isinstance(source_paths, Path):
        normalized_paths = [source_paths]
    else:
        normalized_paths = sorted({Path(path) for path in source_paths}, key=lambda item: item.name.lower())
    try:
        normalized_paths.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    except OSError:
        pass

    if cache_dir is None:
        _app_root, cache_dir = ensure_app_directories()
    purge_stale_cache(cache_dir)

    previous_hashes = previous_data.file_hashes if previous_data else {}
    previous_records_by_path: dict[Path, list[DailyRecord]] = {}
    previous_warnings_by_source_file: dict[str, list[SheetParseWarning]] = {}
    previous_health_by_source_file: dict[str, list[WorkbookHealthItem]] = {}
    if previous_data is not None:
        for source_path in previous_data.source_paths:
            previous_records_by_path[source_path] = [
                record for record in previous_data.daily_records if record.source_path == source_path
            ]
            previous_warnings_by_source_file[source_path.name] = [
                warning for warning in previous_data.parse_warnings if warning.source_file == source_path.name
            ]
            previous_health_by_source_file[source_path.name] = [
                item for item in previous_data.workbook_health if item.source_file == source_path.name
            ]

    file_hashes: dict[Path, str] = {}
    reused_paths: list[Path] = []
    reloaded_paths: list[Path] = []
    cache_status_by_path: dict[Path, str] = {}
    daily_records: list[DailyRecord] = []
    parse_warnings: list[SheetParseWarning] = []
    workbook_health: list[WorkbookHealthItem] = []
    total_files = max(len(normalized_paths), 1)
    check_cancel = should_cancel or (lambda: False)
    state_lock = threading.Lock()
    file_progress = {source_path: 0.0 for source_path in normalized_paths}

    def raise_if_cancelled(message: str) -> None:
        if check_cancel():
            raise OperationCancelled(message)

    def emit_overall_progress(source_path: Path, file_fraction: float, message: str) -> None:
        if progress_callback is None:
            return
        clamped_fraction = min(max(file_fraction, 0.0), 1.0)
        with state_lock:
            file_progress[source_path] = clamped_fraction
            overall = sum(file_progress.values()) / total_files
        progress_callback(overall, message)

    def emit_partial_data(
        source_path: Path,
        partial_records_for_file: list[DailyRecord],
        partial_warnings_for_file: list[SheetParseWarning],
        partial_health_for_file: list[WorkbookHealthItem],
        message: str,
    ) -> None:
        if partial_callback is None:
            return
        with state_lock:
            snapshot_records = [*daily_records, *partial_records_for_file]
            if not snapshot_records:
                return
            snapshot_status_by_path = dict(cache_status_by_path)
            snapshot_status_by_path[source_path] = snapshot_status_by_path.get(source_path, "Miss")
            tracker_snapshot = build_tracker_data_with_status(
                normalized_paths,
                dict(file_hashes),
                list(reused_paths),
                list(reloaded_paths),
                snapshot_records,
                cache_status_by_path=snapshot_status_by_path,
                parse_warnings=[*parse_warnings, *partial_warnings_for_file],
                workbook_health=[*workbook_health, *partial_health_for_file],
            )
        partial_callback(tracker_snapshot, message)

    def parse_miss_file(
        source_path: Path,
        workbook_bytes: bytes,
    ) -> tuple[Path, list[DailyRecord], list[SheetParseWarning], list[WorkbookHealthItem], dict[str, str]]:
        sheet_hashes = compute_all_dated_sheet_hashes(workbook_bytes)
        parsed_records, file_warnings, file_health = process_workbook_bytes(
            source_path,
            workbook_bytes,
            progress_callback=lambda fraction, message, current_path=source_path: emit_overall_progress(
                current_path,
                0.2 + (0.8 * min(max(fraction, 0.0), 1.0)),
                message,
            ),
            should_cancel=check_cancel,
            preview_callback=lambda partial_records, partial_warnings, partial_health, message, current_path=source_path: emit_partial_data(
                current_path,
                partial_records,
                partial_warnings,
                partial_health,
                message,
            ),
        )
        return source_path, parsed_records, file_warnings, file_health, sheet_hashes

    pending_parse_inputs: list[tuple[Path, bytes]] = []

    for source_path in normalized_paths:
        raise_if_cancelled(f"Cancelled loading {source_path.name}")
        emit_overall_progress(source_path, 0.0, f"Starting {source_path.name}")
        workbook_bytes = read_source_bytes(source_path)
        emit_overall_progress(source_path, 0.05, f"Read bytes for {source_path.name}")
        raise_if_cancelled(f"Cancelled loading {source_path.name}")
        file_hash = compute_workbook_content_hash(workbook_bytes)
        file_hashes[source_path] = file_hash
        emit_overall_progress(source_path, 0.1, f"Hashed {source_path.name}")
        if previous_hashes.get(source_path) == file_hash:
            with state_lock:
                reused_paths.append(source_path)
                cache_status_by_path[source_path] = "Memory Hit"
                daily_records.extend(previous_records_by_path.get(source_path, []))
                parse_warnings.extend(previous_warnings_by_source_file.get(source_path.name, []))
                workbook_health.extend(previous_health_by_source_file.get(source_path.name, []))
            emit_overall_progress(source_path, 1.0, f"Unchanged {source_path.name}")
        else:
            raise_if_cancelled(f"Cancelled loading {source_path.name}")
            cached_records = load_cached_daily_records(cache_dir, source_path, file_hash)
            if cached_records is not None:
                with state_lock:
                    reused_paths.append(source_path)
                    cache_status_by_path[source_path] = "Disk Hit"
                    daily_records.extend(cached_records)
                cached_analysis = load_cached_source_analysis(cache_dir, source_path, file_hash)
                if cached_analysis is not None:
                    cached_warnings, cached_health = cached_analysis
                    with state_lock:
                        parse_warnings.extend(cached_warnings)
                        workbook_health.extend(cached_health)
                else:
                    emit_overall_progress(source_path, 0.2, f"Inspecting {source_path.name}")
                    file_warnings, file_health = analyze_workbook_bytes(
                        source_path,
                        workbook_bytes,
                        progress_callback=lambda fraction, message, current_path=source_path: emit_overall_progress(
                            current_path,
                            0.2 + (0.1 * min(max(fraction, 0.0), 1.0)),
                            message,
                        ),
                    )
                    with state_lock:
                        parse_warnings.extend(file_warnings)
                        workbook_health.extend(file_health)
                    save_cached_daily_records(
                        cache_dir,
                        source_path,
                        file_hash,
                        cached_records,
                        parse_warnings=file_warnings,
                        workbook_health=file_health,
                        sheet_hashes=compute_all_dated_sheet_hashes(workbook_bytes),
                    )
                emit_overall_progress(source_path, 1.0, f"Loaded cached data for {source_path.name}")
            else:
                merged = merge_workbook_from_cache_by_sheet_hashes(
                    source_path=source_path,
                    workbook_bytes=workbook_bytes,
                    cache_dir=cache_dir,
                    should_cancel=check_cancel,
                    progress_callback=lambda fraction, message, sp=source_path: emit_overall_progress(
                        sp,
                        0.15 + 0.85 * fraction,
                        message,
                    ),
                )
                if merged is not None:
                    merged_records, merged_warnings, merged_health, merged_sheet_hashes, merged_fp = merged
                    with state_lock:
                        reloaded_paths.append(source_path)
                        cache_status_by_path[source_path] = "Partial Refresh"
                        daily_records.extend(merged_records)
                        parse_warnings.extend(merged_warnings)
                        workbook_health.extend(merged_health)
                    save_cached_daily_records(
                        cache_dir,
                        source_path,
                        file_hash,
                        merged_records,
                        parse_warnings=merged_warnings,
                        workbook_health=merged_health,
                        sheet_hashes=merged_sheet_hashes,
                    )
                    emit_overall_progress(source_path, 1.0, f"Updated changed sheets in {source_path.name}")
                else:
                    with state_lock:
                        reloaded_paths.append(source_path)
                        cache_status_by_path[source_path] = "Miss"
                    emit_overall_progress(source_path, 0.15, f"Queued {source_path.name} for parsing")
                    pending_parse_inputs.append((source_path, workbook_bytes))

    if pending_parse_inputs:
        max_workers = min(MAX_PARALLEL_PARSE_WORKERS, len(pending_parse_inputs))
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(parse_miss_file, source_path, workbook_bytes): source_path
                for source_path, workbook_bytes in pending_parse_inputs
            }
            for future in concurrent.futures.as_completed(future_map):
                raise_if_cancelled("Cancelled DSS load.")
                source_path = future_map[future]
                parsed_source_path, parsed_records, file_warnings, file_health, sheet_hashes = future.result()
                save_cached_daily_records(
                    cache_dir,
                    parsed_source_path,
                    file_hashes[parsed_source_path],
                    parsed_records,
                    parse_warnings=file_warnings,
                    workbook_health=file_health,
                    sheet_hashes=sheet_hashes,
                )
                with state_lock:
                    parse_warnings.extend(file_warnings)
                    workbook_health.extend(file_health)
                    daily_records.extend(parsed_records)
                emit_overall_progress(source_path, 1.0, f"Finished {source_path.name}")
    return build_tracker_data_with_status(
        normalized_paths,
        file_hashes,
        reused_paths,
        reloaded_paths,
        daily_records,
        cache_status_by_path=cache_status_by_path,
        parse_warnings=parse_warnings,
        workbook_health=workbook_health,
    )


def fmt_hours(value: float) -> str:
    text = f"{value:.2f}"
    return text.rstrip("0").rstrip(".")


def expanded_hours(st: float, ot: float, dt: float) -> float:
    return round(st + (ot * 1.5) + (dt * 2.0), 2)


AUTO_COLUMN_WIDTH_PAD_PX = 28
AUTO_COLUMN_MIN_WIDTH_PX = 56
AUTO_SOURCE_FILE_MAX_WIDTH_PX = 300
AUTO_PATH_LIKE_MAX_WIDTH_PX = 340
AUTO_NON_PATH_MAX_WIDTH_PX = 1200


def is_path_like_table_column(column: str, source_file_column: str | None) -> bool:
    """Columns that often hold filenames or paths: cap width when auto-fitting."""
    if column == "source_file" or (source_file_column and column == source_file_column):
        return True
    if column == "sources":
        return True
    if column.endswith("_path"):
        return True
    return False


class SortableTreeview(ttk.Treeview):
    def __init__(self, master, columns: list[str], headings: list[str]):
        super().__init__(master, columns=columns, show="headings")
        self._rows: list[tuple[tuple[str, ...], tuple[str, ...]]] = []
        self._sort_state: dict[str, bool] = {}
        self._active_sort_column: str | None = None
        self._active_sort_descending = False
        self.on_sort_changed: Callable[[], None] | None = None
        self.custom_sort_key: Callable[[str, tuple[str, ...], bool], object] | None = None
        for column, heading in zip(columns, headings):
            self.heading(column, text=heading, command=lambda col=column: self.sort_by(col))
            self.column(column, anchor="w", width=120, stretch=False)

    def set_rows(self, rows: list[tuple[str, ...]], tags: list[tuple[str, ...]] | None = None) -> None:
        self._rows = []
        tags = tags or [tuple() for _ in rows]
        for row, row_tags in zip(rows, tags):
            self._rows.append((row, row_tags))
        if self._active_sort_column and self._active_sort_column in self["columns"]:
            self._sort_rows(self._active_sort_column, self._active_sort_descending)
        self._render_rows()

    def sort_by(self, column: str, descending: bool | None = None) -> None:
        columns = list(self["columns"])
        index = columns.index(column)
        if descending is None:
            if self._active_sort_column == column:
                reverse = not self._active_sort_descending
            else:
                reverse = False
        else:
            reverse = descending

        self._sort_rows(column, reverse, index=index)
        self._active_sort_column = column
        self._active_sort_descending = reverse
        self._sort_state[column] = not reverse
        self._render_rows()
        if self.on_sort_changed is not None:
            self.on_sort_changed()

    def set_sort(self, column: str, descending: bool) -> None:
        if column not in self["columns"]:
            return
        self._active_sort_column = column
        self._active_sort_descending = descending
        self._sort_state[column] = not descending
        if self._rows:
            self._sort_rows(column, descending)
            self._render_rows()

    def current_sort(self) -> tuple[str, bool] | None:
        if not self._active_sort_column:
            return None
        return self._active_sort_column, self._active_sort_descending

    def clear_sort(self) -> None:
        self._active_sort_column = None
        self._active_sort_descending = False
        self._sort_state.clear()

    def _sort_rows(self, column: str, reverse: bool, index: int | None = None) -> None:
        if self.custom_sort_key is not None:
            self._rows.sort(key=lambda item: self.custom_sort_key(column, item[0], reverse))
            return

        columns = list(self["columns"])
        column_index = columns.index(column) if index is None else index

        def sort_key(item: tuple[tuple[str, ...], tuple[str, ...]]):
            value = item[0][column_index]
            try:
                return (0, float(value))
            except ValueError:
                return (1, value.lower())

        self._rows.sort(key=sort_key, reverse=reverse)

    def _render_rows(self) -> None:
        self.delete(*self.get_children())
        for row, row_tags in self._rows:
            self.insert("", "end", values=row, tags=row_tags)


class DataTable(ttk.Frame):
    def __init__(
        self,
        master,
        columns: list[str],
        headings: list[str],
        table_id: str | None = None,
        config_path: Path | None = None,
        default_sort_column: str | None = None,
        default_sort_descending: bool = False,
        custom_sort_key: Callable[[str, tuple[str, ...], bool], object] | None = None,
        open_source_file_callback: Callable[[str], None] | None = None,
        source_file_column: str | None = None,
        ui_theme: UiThemeColors | None = None,
    ):
        super().__init__(master)
        self._table_id = table_id
        self._config_path = config_path
        self._all_columns = list(columns)
        self._headings_by_column = dict(zip(columns, headings))
        self._column_visibility = {column: True for column in columns}
        self._default_sort_column = default_sort_column if default_sort_column in self._all_columns else None
        self._default_sort_descending = default_sort_descending
        self._all_rows: list[tuple[str, ...]] = []
        self._all_tags: list[tuple[str, ...]] = []
        self._column_filters: dict[str, set[str]] = {}
        self._filter_menu: tk.Toplevel | None = None
        self._header_drag_column: str | None = None
        self._open_source_file_callback = open_source_file_callback
        self._source_file_column = source_file_column if source_file_column in columns else None
        self._column_drag_line: tk.Frame | None = None
        self._ui_theme = ui_theme or DEFAULT_UI_THEME

        toolbar = ttk.Frame(self)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self._word_wrap_enabled = False
        ttk.Button(toolbar, text="Columns", command=self.open_column_picker).pack(side="right")
        self._wrap_btn = ttk.Button(toolbar, text="Word wrap (off)", command=self._toggle_word_wrap)
        self._wrap_btn.pack(side="right", padx=(0, 8))

        self.tree = SortableTreeview(self, columns, headings)
        self._tree_style_base = DSS_TABLE_TREEVIEW_STYLE
        self.tree.configure(style=self._tree_style_base)
        self._wrap_style_name = f"DssWrap{id(self)}.Treeview"
        wrap_style = ttk.Style(self)
        try:
            wrap_style.layout(self._wrap_style_name, wrap_style.layout(self._tree_style_base))
        except tk.TclError:
            pass
        self.tree.custom_sort_key = custom_sort_key
        y_scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.tree.grid(row=1, column=0, sticky="nsew")
        y_scroll.grid(row=1, column=1, sticky="ns")
        x_scroll.grid(row=2, column=0, sticky="ew")
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        self._apply_ui_theme_tags()
        self.tree.bind("<ButtonPress-1>", self._on_tree_button_press, add="+")
        self.tree.bind("<ButtonRelease-1>", self._on_tree_button_release, add="+")
        self.tree.bind("<B1-Motion>", self._on_tree_b1_motion, add="+")
        self.tree.bind("<ButtonRelease-3>", self._on_tree_right_click, add="+")
        self.tree.on_sort_changed = self._save_layout
        if self._open_source_file_callback and self._source_file_column:
            self.tree.bind("<Double-Button-1>", self._on_tree_double_click_open_source, add="+")
        bind_vertical_mousewheel(self.tree, self.tree, units_per_notch=4)
        bind_horizontal_mousewheel(self.tree, self.tree, units_per_notch=4)

        self._load_saved_layout()

    def _apply_ui_theme_tags(self) -> None:
        t = self._ui_theme
        self.tree.tag_configure("alert", background=t.alert_row_background, foreground=t.alert_row_foreground)
        self.tree.tag_configure("crew_total", background=t.crew_total_background, foreground=t.crew_total_foreground)

    def _sync_wrap_style_colours(self) -> None:
        t = self._ui_theme
        style = ttk.Style(self)
        bg = t.table_background
        heading_bg = lighten_hex_color(bg, -18)
        heading_fg = "#18181b"
        cell_fg = "#18181b"
        try:
            style.configure(
                self._wrap_style_name,
                background=bg,
                fieldbackground=bg,
                foreground=cell_fg,
            )
            style.configure(f"{self._wrap_style_name}.Heading", background=heading_bg, foreground=heading_fg)
            style.map(
                self._wrap_style_name,
                background=[("selected", "#bfdbfe")],
                foreground=[("selected", "#1e3a8a")],
            )
        except tk.TclError:
            pass

    def apply_ui_theme(self, theme: UiThemeColors) -> None:
        self._ui_theme = theme
        self._apply_ui_theme_tags()
        self._sync_wrap_style_colours()

    def set_rows(self, rows: list[tuple[str, ...]], tags: list[tuple[str, ...]] | None = None) -> None:
        self._all_rows = list(rows)
        self._all_tags = list(tags or [tuple() for _ in rows])
        self._apply_filters_and_render()

    def open_column_picker(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Show / Hide Columns")
        dialog.transient(self.winfo_toplevel())
        dialog.resizable(False, False)

        vars_by_column: dict[str, tk.BooleanVar] = {}
        content = ttk.Frame(dialog, padding=12)
        content.pack(fill="both", expand=True)

        for row_index, column in enumerate(self._all_columns):
            var = tk.BooleanVar(value=self._column_visibility[column])
            vars_by_column[column] = var
            ttk.Checkbutton(
                content,
                text=self._headings_by_column[column],
                variable=var,
            ).grid(row=row_index, column=0, sticky="w")

        buttons = ttk.Frame(content)
        buttons.grid(row=len(self._all_columns), column=0, sticky="ew", pady=(12, 0))
        ttk.Button(buttons, text="Show All", command=lambda: self._set_all_column_vars(vars_by_column, True)).pack(side="left")
        ttk.Button(buttons, text="Hide All", command=lambda: self._set_all_column_vars(vars_by_column, False)).pack(side="left", padx=(8, 0))
        ttk.Button(
            buttons,
            text="Apply",
            command=lambda: self._apply_column_visibility(vars_by_column, dialog),
        ).pack(side="right")

    @staticmethod
    def _set_all_column_vars(vars_by_column: dict[str, tk.BooleanVar], value: bool) -> None:
        for var in vars_by_column.values():
            var.set(value)

    def _apply_column_visibility(self, vars_by_column: dict[str, tk.BooleanVar], dialog: tk.Toplevel) -> None:
        selected_columns = [column for column in self._all_columns if vars_by_column[column].get()]
        if not selected_columns:
            messagebox.showerror("Columns", "At least one column must remain visible.")
            return
        for column in self._all_columns:
            self._column_visibility[column] = vars_by_column[column].get()
        self.tree.configure(displaycolumns=selected_columns)
        self._save_layout()
        dialog.destroy()

    def _load_saved_layout(self) -> None:
        if not self._config_path or not self._table_id:
            self._apply_default_sort()
            return
        layouts = load_table_layouts(self._config_path)
        layout = layouts.get(self._table_id)
        if not isinstance(layout, dict):
            self._column_filters.clear()
            self._apply_default_sort()
            return

        requested_columns = [
            column for column in layout.get("visible_columns", [])
            if column in self._all_columns
        ]
        if requested_columns:
            self.tree.configure(displaycolumns=requested_columns)
            for column in self._all_columns:
                self._column_visibility[column] = column in requested_columns

        sort_column = str(layout.get("sort_column", "")).strip()
        if sort_column in self._all_columns:
            self.tree.set_sort(sort_column, bool(layout.get("sort_descending", False)))
        else:
            self._apply_default_sort()
        self._column_filters.clear()
        saved_filters = layout.get("column_filters", {})
        if isinstance(saved_filters, dict):
            for column, values in saved_filters.items():
                if column not in self._all_columns:
                    continue
                if isinstance(values, set) and values:
                    self._column_filters[column] = set(values)
                elif isinstance(values, list) and values:
                    self._column_filters[column] = {str(value) for value in values}
        self._apply_filters_and_render()

    def _apply_default_sort(self) -> None:
        if self._default_sort_column:
            self.tree.set_sort(self._default_sort_column, self._default_sort_descending)

    def _save_layout(self) -> None:
        if not self._config_path or not self._table_id:
            return
        display_columns_value = self.tree.cget("displaycolumns")
        if display_columns_value == "#all":
            visible_columns = list(self._all_columns)
        else:
            display_columns = list(display_columns_value)
            visible_columns = [column for column in display_columns if column in self._all_columns]
        column_widths = {
            column: int(self.tree.column(column, "width"))
            for column in self._all_columns
        }
        current_sort = self.tree.current_sort()
        sort_column = current_sort[0] if current_sort is not None else ""
        sort_descending = current_sort[1] if current_sort is not None else False
        column_filters_payload = {
            column: values
            for column, values in self._column_filters.items()
            if column in self._all_columns and values
        }
        save_table_layout(
            self._config_path,
            self._table_id,
            visible_columns,
            column_widths,
            sort_column=sort_column,
            sort_descending=sort_descending,
            column_filters=column_filters_payload if column_filters_payload else None,
        )

    def _on_tree_button_release(self, _event=None) -> None:
        self._handle_header_drop()
        self._hide_column_drag_line()
        self.after_idle(self._save_layout)

    def _on_tree_button_press(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region != "heading":
            self._header_drag_column = None
            self._hide_column_drag_line()
            return
        self._header_drag_column = self._column_from_identified_id(self.tree.identify_column(event.x))

    def _hide_column_drag_line(self) -> None:
        if self._column_drag_line is not None:
            self._column_drag_line.place_forget()

    def _on_tree_b1_motion(self, event) -> None:
        if not self._header_drag_column:
            self._hide_column_drag_line()
            return
        if self.tree.identify_region(event.x, event.y) != "heading":
            self._hide_column_drag_line()
            return
        target_id = self.tree.identify_column(event.x)
        target_column = self._column_from_identified_id(target_id)
        if not target_column:
            self._hide_column_drag_line()
            return
        x_edge = self._heading_left_edge_x_for_column_id(target_id)
        if x_edge is None:
            self._hide_column_drag_line()
            return
        if self._column_drag_line is None:
            self._column_drag_line = tk.Frame(self.tree, width=3, bg="#0078d7")
        self._column_drag_line.place(in_=self.tree, x=max(x_edge - 1, 0), y=0, relheight=1.0)

    def _heading_left_edge_x_for_column_id(self, column_id: str) -> int | None:
        if not column_id.startswith("#"):
            return None
        try:
            index = int(column_id[1:]) - 1
        except ValueError:
            return None
        if index < 0:
            return None
        x_acc = 0
        display_columns_value = self.tree.cget("displaycolumns")
        if display_columns_value == "#all":
            ordered = list(self._all_columns)
        else:
            ordered = [c for c in list(display_columns_value) if c in self._all_columns]
        for col in ordered:
            if self._all_columns.index(col) == index:
                return x_acc
            x_acc += int(self.tree.column(col, "width"))
        return None

    def _on_tree_double_click_open_source(self, event) -> None:
        if not self._open_source_file_callback or not self._source_file_column:
            return
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        column = self._column_from_identified_id(self.tree.identify_column(event.x))
        if column != self._source_file_column:
            return
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        values = self.tree.item(row_id, "values")
        col_index = self._all_columns.index(self._source_file_column)
        if col_index >= len(values):
            return
        name = str(values[col_index]).replace("\n", "").replace("\r", "").strip()
        if name:
            self._open_source_file_callback(name)

    def _on_tree_right_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region != "heading":
            self._close_filter_menu()
            return
        column = self._column_from_identified_id(self.tree.identify_column(event.x))
        if column:
            self._open_header_filter_menu(column, event)

    def _handle_header_drop(self) -> None:
        if not self._header_drag_column:
            return
        pointer_x = self.winfo_pointerx() - self.tree.winfo_rootx()
        pointer_y = self.winfo_pointery() - self.tree.winfo_rooty()
        if self.tree.identify_region(pointer_x, pointer_y) != "heading":
            self._header_drag_column = None
            self._hide_column_drag_line()
            return
        target_column = self._column_from_identified_id(self.tree.identify_column(pointer_x))
        dragged_column = self._header_drag_column
        self._header_drag_column = None
        if not target_column or target_column == dragged_column:
            self._hide_column_drag_line()
            return
        display_columns_value = self.tree.cget("displaycolumns")
        if display_columns_value == "#all":
            display_columns = list(self._all_columns)
        else:
            display_columns = [column for column in list(display_columns_value) if column in self._all_columns]
        if dragged_column not in display_columns or target_column not in display_columns:
            self._hide_column_drag_line()
            return
        display_columns.remove(dragged_column)
        target_index = display_columns.index(target_column)
        display_columns.insert(target_index, dragged_column)
        self.tree.configure(displaycolumns=display_columns)

    def _column_from_identified_id(self, column_id: str) -> str | None:
        if not column_id.startswith("#"):
            return None
        try:
            index = int(column_id[1:]) - 1
        except ValueError:
            return None
        return self._all_columns[index] if 0 <= index < len(self._all_columns) else None

    def _apply_filters_and_render(self) -> None:
        filtered_rows: list[tuple[str, ...]] = []
        filtered_tags: list[tuple[str, ...]] = []
        for row, row_tags in zip(self._all_rows, self._all_tags):
            include = True
            for column, allowed_values in self._column_filters.items():
                if not allowed_values:
                    continue
                column_index = self._all_columns.index(column)
                if str(row[column_index]) not in allowed_values:
                    include = False
                    break
            if include:
                filtered_rows.append(row)
                filtered_tags.append(row_tags)
        self.tree.set_rows(filtered_rows, filtered_tags)
        self._autofit_column_widths_to_content()
        if self._word_wrap_enabled:
            self._apply_word_wrap_to_visible_rows()
        else:
            self.tree.configure(style=self._tree_style_base)

    def _measure_font_for_tree(self) -> tkfont.Font:
        try:
            spec = self.tree.cget("font")
        except tk.TclError:
            spec = None
        if isinstance(spec, tuple) and spec:
            family, size = spec[0], spec[1]
            opts: dict[str, object] = {"root": self, "family": family, "size": int(float(size))}
            if len(spec) > 2:
                opts["weight"] = spec[2]
            return tkfont.Font(**opts)
        if isinstance(spec, str) and spec.strip():
            try:
                return tkfont.nametofont(spec.strip())
            except tk.TclError:
                pass
        return tkfont.nametofont("TkDefaultFont")

    def _autofit_column_widths_to_content(self) -> None:
        font_obj = self._measure_font_for_tree()
        display_columns_value = self.tree.cget("displaycolumns")
        if display_columns_value == "#all":
            visible = list(self._all_columns)
        else:
            visible = [column for column in list(display_columns_value) if column in self._all_columns]
        if not visible:
            return
        for column in visible:
            heading = self._headings_by_column.get(column, column)
            max_px = font_obj.measure(str(heading)) + AUTO_COLUMN_WIDTH_PAD_PX
            col_index = self._all_columns.index(column)
            for item_id in self.tree.get_children():
                values = self.tree.item(item_id, "values")
                if col_index >= len(values):
                    continue
                cell = str(values[col_index]).replace("\n", " ")
                span = font_obj.measure(cell) + AUTO_COLUMN_WIDTH_PAD_PX
                if span > max_px:
                    max_px = span
            width = max(AUTO_COLUMN_MIN_WIDTH_PX, int(max_px))
            if is_path_like_table_column(column, self._source_file_column):
                cap = (
                    AUTO_SOURCE_FILE_MAX_WIDTH_PX
                    if column == "source_file" or column == self._source_file_column
                    else AUTO_PATH_LIKE_MAX_WIDTH_PX
                )
                width = min(width, cap)
            else:
                width = min(width, AUTO_NON_PATH_MAX_WIDTH_PX)
            self.tree.column(column, width=width, stretch=False)

    def _toggle_word_wrap(self) -> None:
        self._word_wrap_enabled = not self._word_wrap_enabled
        self._wrap_btn.configure(text=f"Word wrap {'(on)' if self._word_wrap_enabled else '(off)'}")
        self._apply_filters_and_render()

    def _wrap_plain_text_to_pixels(self, text: str, font_obj: tkfont.Font, max_px: int) -> str:
        single = str(text).replace("\r\n", "\n").replace("\r", "\n").replace("\n", " ")
        if max_px < 24 or not single.strip():
            return str(text)
        if font_obj.measure(single) <= max_px:
            return str(text)
        words = single.split(" ")
        lines: list[str] = []
        current = ""
        for word in words:
            trial = f"{current} {word}".strip() if current else word
            if font_obj.measure(trial) <= max_px:
                current = trial
            else:
                if current:
                    lines.append(current)
                if font_obj.measure(word) <= max_px:
                    current = word
                else:
                    frag = ""
                    for ch in word:
                        cand = frag + ch
                        if font_obj.measure(cand) <= max_px:
                            frag = cand
                        else:
                            if frag:
                                lines.append(frag)
                            frag = ch
                    current = frag
        if current:
            lines.append(current)
        return "\n".join(lines)

    def _apply_word_wrap_to_visible_rows(self) -> None:
        font_obj = self._measure_font_for_tree()
        pad = AUTO_COLUMN_WIDTH_PAD_PX
        max_lines = 1
        col_widths = {col: max(24, int(self.tree.column(col, "width")) - pad) for col in self._all_columns}
        for item_id in self.tree.get_children():
            raw = self.tree.item(item_id, "values")
            new_vals: list[str] = []
            for idx, column in enumerate(self._all_columns):
                if idx >= len(raw):
                    new_vals.append("")
                    continue
                wpx = col_widths.get(column, 120)
                wrapped = self._wrap_plain_text_to_pixels(str(raw[idx]), font_obj, wpx)
                max_lines = max(max_lines, wrapped.count("\n") + 1)
                new_vals.append(wrapped)
            self.tree.item(item_id, values=tuple(new_vals))
        line_px = max(1, int(font_obj.metrics("linespace")))
        height = min(max(line_px + 6, line_px * max_lines + 6), line_px * 14)
        style = ttk.Style(self)
        self._sync_wrap_style_colours()
        try:
            style.configure(self._wrap_style_name, rowheight=height)
        except tk.TclError:
            return
        self.tree.configure(style=self._wrap_style_name)

    def _open_header_filter_menu(self, column: str, event) -> None:
        self._close_filter_menu()
        dialog = tk.Toplevel(self)
        dialog.wm_overrideredirect(True)
        dialog.wm_geometry(f"+{event.x_root}+{event.y_root}")
        dialog.transient(self.winfo_toplevel())
        dialog.bind("<FocusOut>", lambda _event: self._close_filter_menu())
        self._filter_menu = dialog

        frame = ttk.Frame(dialog, padding=8, relief="solid", borderwidth=1)
        frame.pack(fill="both", expand=True)

        ttk.Button(frame, text="Sort Ascending", command=lambda: self._sort_and_close(column, False)).pack(fill="x")
        ttk.Button(frame, text="Sort Descending", command=lambda: self._sort_and_close(column, True)).pack(fill="x", pady=(4, 8))

        value_frame = ttk.Frame(frame)
        value_frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(value_frame, width=280, height=220, highlightthickness=0)
        scrollbar = ttk.Scrollbar(value_frame, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        bind_vertical_mousewheel(canvas, canvas, units_per_notch=4)

        values = sorted({str(row[self._all_columns.index(column)]) for row in self._all_rows}, key=lambda item: item.casefold())
        current_filter = set(self._column_filters.get(column, set()))
        if not current_filter:
            current_filter = set(values)
        vars_by_value: dict[str, tk.BooleanVar] = {}

        control_frame = ttk.Frame(inner)
        control_frame.pack(fill="x")
        ttk.Button(control_frame, text="Check All", command=lambda: self._set_filter_vars(vars_by_value, True)).pack(side="left")
        ttk.Button(control_frame, text="Uncheck All", command=lambda: self._set_filter_vars(vars_by_value, False)).pack(side="left", padx=(8, 0))

        for value in values:
            var = tk.BooleanVar(value=value in current_filter)
            vars_by_value[value] = var
            ttk.Checkbutton(inner, text=value or "(blank)", variable=var).pack(anchor="w")

        action_frame = ttk.Frame(frame)
        action_frame.pack(fill="x", pady=(8, 0))
        ttk.Button(action_frame, text="Apply", command=lambda: self._apply_column_filter(column, vars_by_value)).pack(side="right")
        ttk.Button(action_frame, text="Clear Filter", command=lambda: self._clear_column_filter(column)).pack(side="right", padx=(0, 8))

        dialog.focus_force()

    @staticmethod
    def _set_filter_vars(vars_by_value: dict[str, tk.BooleanVar], value: bool) -> None:
        for var in vars_by_value.values():
            var.set(value)

    def _sort_and_close(self, column: str, descending: bool) -> None:
        self.tree.sort_by(column, descending=descending)
        self._close_filter_menu()

    def _apply_column_filter(self, column: str, vars_by_value: dict[str, tk.BooleanVar]) -> None:
        selected_values = {value for value, var in vars_by_value.items() if var.get()}
        if not selected_values:
            messagebox.showerror("Column Filter", "At least one value must remain selected.")
            return
        all_values = {str(row[self._all_columns.index(column)]) for row in self._all_rows}
        if selected_values == all_values:
            self._column_filters.pop(column, None)
        else:
            self._column_filters[column] = selected_values
        self._apply_filters_and_render()
        self._close_filter_menu()
        self.after_idle(self._save_layout)

    def _clear_column_filter(self, column: str) -> None:
        self._column_filters.pop(column, None)
        self._apply_filters_and_render()
        self._close_filter_menu()
        self.after_idle(self._save_layout)

    def _close_filter_menu(self) -> None:
        if self._filter_menu is not None:
            self._filter_menu.destroy()
            self._filter_menu = None

    def reset_layout(self) -> None:
        self._word_wrap_enabled = False
        self._wrap_btn.configure(text="Word wrap (off)")
        for column in self._all_columns:
            self._column_visibility[column] = True
            self.tree.column(column, width=120, stretch=False)
        self.tree.configure(displaycolumns="#all", style=self._tree_style_base)
        self._column_filters.clear()
        self._header_drag_column = None
        self._hide_column_drag_line()
        self.tree.clear_sort()
        self._apply_default_sort()
        self._apply_filters_and_render()

    def displayed_columns_and_rows(self) -> tuple[list[str], list[tuple[str, ...]]]:
        display_columns_value = self.tree.cget("displaycolumns")
        if display_columns_value == "#all":
            columns = list(self._all_columns)
        else:
            columns = [column for column in list(display_columns_value) if column in self._all_columns]
        headings = [self._headings_by_column[column] for column in columns]
        rows: list[tuple[str, ...]] = []
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            rows.append(tuple(str(value) for value in values))
        return headings, rows


class ToolTip:
    def __init__(
        self,
        widget: tk.Widget,
        text: str,
        color_getter: Callable[[], tuple[str, str]] | None = None,
    ):
        self.widget = widget
        self.text = text
        self._color_getter = color_getter
        self.tip_window: tk.Toplevel | None = None
        self._after_id: str | None = None
        existing = getattr(widget, "_tooltip_refs", [])
        existing.append(self)
        setattr(widget, "_tooltip_refs", existing)
        widget.bind("<Enter>", self._schedule_show, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")
        widget.bind("<Destroy>", self._hide, add="+")

    def _schedule_show(self, _event=None) -> None:
        self._cancel_scheduled()
        self._after_id = self.widget.after(500, self._show)

    def _cancel_scheduled(self) -> None:
        if self._after_id is not None:
            self.widget.after_cancel(self._after_id)
            self._after_id = None

    def _show(self) -> None:
        self._after_id = None
        if self.tip_window is not None or not self.text.strip():
            return
        x = self.widget.winfo_rootx() + 18
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        if self._color_getter is not None:
            bg, fg = self._color_getter()
        else:
            bg, fg = DEFAULT_UI_THEME.tooltip_background, DEFAULT_UI_THEME.tooltip_foreground
        label = tk.Label(
            self.tip_window,
            text=self.text,
            justify="left",
            background=bg,
            foreground=fg,
            relief="solid",
            borderwidth=1,
            wraplength=320,
            padx=8,
            pady=6,
        )
        label.pack()

    def _hide(self, _event=None) -> None:
        self._cancel_scheduled()
        if self.tip_window is not None:
            self.tip_window.destroy()
            self.tip_window = None


def bind_vertical_mousewheel(widget: tk.Widget, scroll_target, units_per_notch: int = 3) -> None:
    def on_mousewheel(event) -> str:
        delta = int(getattr(event, "delta", 0))
        if delta == 0:
            return "break"
        steps = max(1, abs(delta) // 120) * units_per_notch
        direction = -1 if delta > 0 else 1
        scroll_target.yview_scroll(direction * steps, "units")
        return "break"

    def on_button4(_event) -> str:
        scroll_target.yview_scroll(-units_per_notch, "units")
        return "break"

    def on_button5(_event) -> str:
        scroll_target.yview_scroll(units_per_notch, "units")
        return "break"

    widget.bind("<MouseWheel>", on_mousewheel, add="+")
    widget.bind("<Button-4>", on_button4, add="+")
    widget.bind("<Button-5>", on_button5, add="+")


def bind_horizontal_mousewheel(widget: tk.Widget, scroll_target, units_per_notch: int = 3) -> None:
    def on_shift_mousewheel(event) -> str:
        delta = int(getattr(event, "delta", 0))
        if delta == 0:
            return "break"
        steps = max(1, abs(delta) // 120) * units_per_notch
        direction = -1 if delta > 0 else 1
        scroll_target.xview_scroll(direction * steps, "units")
        return "break"

    widget.bind("<Shift-MouseWheel>", on_shift_mousewheel, add="+")


class VerticalScrollablePage(ttk.Frame):
    """Notebook (or similar) page with a vertical scrollbar when content is taller than the viewport."""

    def __init__(self, master: tk.Misc):
        super().__init__(master)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        style = ttk.Style(self)
        canvas_bg = style.lookup("TFrame", "background") or "#f0f0f0"
        self._canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0, background=canvas_bg)
        self._vscroll = ttk.Scrollbar(self, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._vscroll.set)
        self._canvas.grid(row=0, column=0, sticky="nsew")
        self._vscroll.grid(row=0, column=1, sticky="ns")
        self.inner = ttk.Frame(self._canvas)
        self._inner_window = self._canvas.create_window((0, 0), window=self.inner, anchor="nw")

        def _on_inner_configure(_event: tk.Event | None = None) -> None:
            self._canvas.configure(scrollregion=self._canvas.bbox("all"))

        def _on_canvas_configure(event: tk.Event) -> None:
            self._canvas.itemconfigure(self._inner_window, width=max(1, int(event.width)))

        self.inner.bind("<Configure>", _on_inner_configure)
        self._canvas.bind("<Configure>", _on_canvas_configure)

    def wire_mousewheel_to_canvas(self) -> None:
        """Route wheel events from descendants to the page canvas (except self-scrolling Text/Listbox/Treeview)."""
        canvas = self._canvas

        bind_vertical_mousewheel(canvas, canvas)

        def visit(parent: tk.Misc) -> None:
            for child in parent.winfo_children():
                if isinstance(child, (tk.Text, tk.Listbox)):
                    bind_vertical_mousewheel(child, child, units_per_notch=4)
                elif isinstance(child, ttk.Treeview):
                    pass
                else:
                    bind_vertical_mousewheel(child, canvas, units_per_notch=3)
                visit(child)

        visit(self.inner)


class EmployeeListEditor(ttk.Frame):
    def __init__(
        self,
        master,
        on_email_changed: Callable[[str, str], None] | None = None,
        on_request_edit_email: Callable[[str], None] | None = None,
    ):
        super().__init__(master)
        self.names: list[str] = []
        self.email_map: dict[str, str] = {}
        self.on_email_changed = on_email_changed
        self.on_request_edit_email = on_request_edit_email

        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 8))

        self.entry = ttk.Entry(top)
        self.entry.pack(side="left", fill="x", expand=True)
        ttk.Button(top, text="Add Name", command=self.add_name).pack(side="left", padx=(8, 0))
        ttk.Button(top, text="Remove Selected", command=self.remove_selected).pack(side="left", padx=(8, 0))

        self.listbox = tk.Listbox(self, activestyle="none")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self.listbox.bind("<<ListboxSelect>>", self._on_selection_changed)
        self.listbox.bind("<Double-Button-1>", self._on_double_click)
        bind_vertical_mousewheel(self.listbox, self.listbox, units_per_notch=4)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        editor = ttk.Frame(self)
        editor.pack(fill="x", pady=(8, 0))
        ttk.Label(editor, text="Selected Employee").grid(row=0, column=0, sticky="w")
        self.selected_name_var = tk.StringVar(value="")
        ttk.Label(editor, textvariable=self.selected_name_var).grid(row=0, column=1, sticky="w", padx=(8, 0))
        ttk.Label(editor, text="Email").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.email_var = tk.StringVar(value="")
        ttk.Entry(editor, textvariable=self.email_var).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))
        ttk.Button(editor, text="Save Email", command=self.save_email).grid(row=1, column=2, padx=(8, 0), pady=(8, 0))
        editor.columnconfigure(1, weight=1)

        note = ttk.Label(
            self,
            text="Parsed employee names are grouped by exact spelling. You can also store employee email addresses here for draft emails.",
            wraplength=500,
            justify="left",
        )
        note.pack(fill="x", pady=(8, 0))

    def set_names(self, names: list[str], email_map: dict[str, str]) -> None:
        self.names = list(names)
        self.email_map = dict(email_map)
        self.refresh()

    def refresh(self) -> None:
        self.listbox.delete(0, tk.END)
        for name in self.names:
            self.listbox.insert(tk.END, name)
        self._on_selection_changed()

    def add_name(self) -> None:
        value = self.entry.get().strip()
        if not value or value in self.names:
            return
        self.names.append(value)
        self.names.sort()
        self.entry.delete(0, tk.END)
        self.refresh()

    def remove_selected(self) -> None:
        selected = self.listbox.curselection()
        if not selected:
            return
        for index in reversed(selected):
            employee = self.names[index]
            self.email_map.pop(employee, None)
            del self.names[index]
        self.refresh()

    def _on_selection_changed(self, _event=None) -> None:
        selected = self.listbox.curselection()
        if not selected:
            self.selected_name_var.set("")
            self.email_var.set("")
            return
        employee = self.names[selected[0]]
        self.selected_name_var.set(employee)
        self.email_var.set(self.email_map.get(employee, ""))

    def save_email(self) -> None:
        employee = self.selected_name_var.get().strip()
        if not employee:
            return
        email = self.email_var.get().strip()
        self.email_map[employee] = email
        if self.on_email_changed is not None:
            self.on_email_changed(employee, email)

    def _on_double_click(self, _event=None) -> None:
        employee = self.selected_name_var.get().strip()
        if employee and self.on_request_edit_email is not None:
            self.on_request_edit_email(employee)


class EmailDraftsFrame(ttk.Frame):
    def __init__(
        self,
        master,
        create_drafts_callback: Callable[[], None],
        sync_emails_callback: Callable[[], None],
        on_request_edit_email: Callable[[str], None],
        save_templates_callback: Callable[[], None],
        ui_theme: UiThemeColors | None = None,
    ):
        super().__init__(master, padding=12)
        self.create_drafts_callback = create_drafts_callback
        self.sync_emails_callback = sync_emails_callback
        self.on_request_edit_email = on_request_edit_email
        self.save_templates_callback = save_templates_callback
        self._ui_theme = ui_theme or DEFAULT_UI_THEME

        self.columnconfigure(1, weight=1)
        self.rowconfigure(4, weight=1)

        self.week_options: list[tuple[date, date]] = []
        self.week_var = tk.StringVar()

        ttk.Label(self, text="Week").grid(row=0, column=0, sticky="w")
        self.week_combo = ttk.Combobox(self, textvariable=self.week_var, state="readonly")
        self.week_combo.grid(row=0, column=1, sticky="ew", pady=(0, 8))

        ttk.Label(self, text="Subject Template").grid(row=1, column=0, sticky="nw")
        subject_wrap = ttk.Frame(self)
        subject_wrap.grid(row=1, column=1, sticky="ew", pady=(0, 8))
        subject_wrap.columnconfigure(0, weight=1)
        self.subject_template_text = tk.Text(subject_wrap, wrap="word", height=2)
        subject_scroll = ttk.Scrollbar(subject_wrap, orient="vertical", command=self.subject_template_text.yview)
        self.subject_template_text.configure(yscrollcommand=subject_scroll.set)
        self.subject_template_text.grid(row=0, column=0, sticky="nsew")
        subject_scroll.grid(row=0, column=1, sticky="ns")
        bind_vertical_mousewheel(self.subject_template_text, self.subject_template_text, units_per_notch=4)

        ttk.Button(self, text="Create Outlook Drafts", command=self.create_drafts_callback).grid(
            row=1, column=2, padx=(8, 0), pady=(0, 8)
        )
        ttk.Button(self, text="Sync Outlook Emails", command=self.sync_emails_callback).grid(
            row=1, column=3, padx=(8, 0), pady=(0, 8)
        )
        ttk.Button(self, text="Save Templates", command=self.save_templates_callback).grid(
            row=1, column=4, padx=(8, 0), pady=(0, 8)
        )

        self.summary_label = ttk.Label(
            self,
            text="Load DSS data to preview employee email drafts.",
            wraplength=700,
            justify="left",
        )
        self.summary_label.grid(row=2, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(self, text="Body Template (HTML)").grid(row=3, column=0, sticky="nw")
        body_wrap = ttk.Frame(self)
        body_wrap.grid(row=3, column=1, columnspan=4, sticky="nsew", pady=(0, 8))
        body_wrap.columnconfigure(0, weight=1)
        body_wrap.rowconfigure(0, weight=1)
        self.body_template_text = tk.Text(body_wrap, wrap="word", height=8)
        body_scroll = ttk.Scrollbar(body_wrap, orient="vertical", command=self.body_template_text.yview)
        self.body_template_text.configure(yscrollcommand=body_scroll.set)
        self.body_template_text.grid(row=0, column=0, sticky="nsew")
        body_scroll.grid(row=0, column=1, sticky="ns")
        bind_vertical_mousewheel(self.body_template_text, self.body_template_text, units_per_notch=4)

        self.preview_table = DataTable(
            self,
            columns=["employee", "email", "days", "st", "ot", "dt", "total", "expanded"],
            headings=["Employee", "Email", "Rows", "ST", "OT", "DT", "Total", "Expanded Hours"],
            ui_theme=self._ui_theme,
        )
        self.preview_table.grid(row=4, column=0, columnspan=5, sticky="nsew")
        self.preview_table.tree.bind("<Double-Button-1>", self._on_preview_double_click)

        note = (
            "Use {employee}, {first_name}, {week_start}, {week_end}, {pf_numbers}, and {hours_table} in the templates. "
            "Drafts are saved in Outlook and are not sent automatically."
        )
        ttk.Label(self, text=note, wraplength=700, justify="left").grid(row=5, column=0, columnspan=5, sticky="w", pady=(8, 0))

    def apply_ui_theme(self, theme: UiThemeColors) -> None:
        self._ui_theme = theme
        self.preview_table.apply_ui_theme(theme)

    def set_week_options(self, week_options: list[tuple[date, date]]) -> None:
        self.week_options = week_options
        labels = [format_week_label(week_start, week_end) for week_start, week_end in week_options]
        self.week_combo.configure(values=labels)
        if labels:
            self.week_var.set(labels[-1])
        else:
            self.week_var.set("")

    def selected_week_start(self) -> date | None:
        selected_label = self.week_var.get().strip()
        for week_start, week_end in self.week_options:
            if format_week_label(week_start, week_end) == selected_label:
                return week_start
        return self.week_options[-1][0] if self.week_options else None

    def _on_preview_double_click(self, _event=None) -> None:
        selected = self.preview_table.tree.selection()
        if not selected:
            return
        employee = self.preview_table.tree.item(selected[0], "values")[0]
        if employee:
            self.on_request_edit_email(employee)

    def set_templates(self, subject_template: str, body_template: str) -> None:
        self.subject_template_text.delete("1.0", tk.END)
        self.subject_template_text.insert("1.0", subject_template)
        self.body_template_text.delete("1.0", tk.END)
        self.body_template_text.insert("1.0", body_template)

    def get_subject_template(self) -> str:
        return self.subject_template_text.get("1.0", tk.END).strip()

    def get_body_template(self) -> str:
        return self.body_template_text.get("1.0", tk.END).strip()

    def selected_employees(self) -> set[str]:
        employees: set[str] = set()
        for item_id in self.preview_table.tree.selection():
            values = self.preview_table.tree.item(item_id, "values")
            if values:
                employees.add(str(values[0]))
        return employees


class EmployeeGroupsFrame(ttk.Frame):
    def __init__(
        self,
        master,
        on_groups_changed: Callable[[dict[str, list[str]]], None],
        on_request_edit_email: Callable[[str], None],
        sync_emails_callback: Callable[[], None],
    ):
        super().__init__(master, padding=12)
        self.on_groups_changed = on_groups_changed
        self.on_request_edit_email = on_request_edit_email
        self.sync_emails_callback = sync_emails_callback
        self.employee_names: list[str] = []
        self.employee_groups: dict[str, list[str]] = {}
        self.email_map: dict[str, str] = {}

        self.columnconfigure(1, weight=1)
        self.rowconfigure(1, weight=1)

        left = ttk.Frame(self)
        left.grid(row=0, column=0, rowspan=3, sticky="nsw", padx=(0, 12))
        ttk.Label(left, text="Groups").pack(anchor="w")
        list_row = ttk.Frame(left)
        list_row.pack(fill="both", expand=True, pady=(4, 8))
        self.group_listbox = tk.Listbox(list_row, exportselection=False, height=12)
        group_list_scroll = ttk.Scrollbar(list_row, orient="vertical", command=self.group_listbox.yview)
        self.group_listbox.configure(yscrollcommand=group_list_scroll.set)
        self.group_listbox.pack(side="left", fill="both", expand=True)
        group_list_scroll.pack(side="right", fill="y")
        self.group_listbox.bind("<<ListboxSelect>>", self._on_group_selected)
        bind_vertical_mousewheel(self.group_listbox, self.group_listbox, units_per_notch=4)
        ttk.Button(left, text="New Group", command=self.create_group).pack(fill="x")
        ttk.Button(left, text="Delete Group", command=self.delete_group).pack(fill="x", pady=(8, 0))
        ttk.Button(left, text="Sync Outlook Emails", command=self.sync_emails_callback).pack(fill="x", pady=(8, 0))

        right = ttk.Frame(self)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        ttk.Label(right, text="Employees in selected group").grid(row=0, column=0, sticky="w")
        self.employee_listbox = tk.Listbox(right, selectmode="extended", exportselection=False)
        self.employee_listbox.grid(row=1, column=0, sticky="nsew", pady=(4, 8))
        self.employee_listbox.bind("<Double-Button-1>", self._on_employee_double_click)
        bind_vertical_mousewheel(self.employee_listbox, self.employee_listbox, units_per_notch=4)
        employee_scroll = ttk.Scrollbar(right, orient="vertical", command=self.employee_listbox.yview)
        self.employee_listbox.configure(yscrollcommand=employee_scroll.set)
        employee_scroll.grid(row=1, column=1, sticky="ns", pady=(4, 8))

        ttk.Button(right, text="Save Group Members", command=self.save_group_members).grid(row=2, column=0, sticky="w")
        self.summary_label = ttk.Label(
            self,
            text="Create groups to filter the tables by crews, trades, or project splits.",
            wraplength=700,
            justify="left",
        )
        self.summary_label.grid(row=3, column=0, columnspan=2, sticky="w", pady=(12, 0))

    def set_data(
        self,
        employee_names: list[str],
        employee_groups: dict[str, list[str]],
        email_map: dict[str, str],
    ) -> None:
        current_group = self.selected_group_name()
        self.employee_names = list(employee_names)
        self.employee_groups = {name: list(members) for name, members in employee_groups.items()}
        self.email_map = dict(email_map)

        self.group_listbox.delete(0, tk.END)
        for group_name in sorted(self.employee_groups):
            self.group_listbox.insert(tk.END, group_name)

        self.employee_listbox.delete(0, tk.END)
        for employee in self.employee_names:
            email = self.email_map.get(employee, "").strip()
            label = f"{employee} | {email}" if email else f"{employee} | (missing)"
            self.employee_listbox.insert(tk.END, label)

        if current_group and current_group in self.employee_groups:
            group_names = sorted(self.employee_groups)
            self.group_listbox.selection_set(group_names.index(current_group))
        elif self.group_listbox.size() > 0:
            self.group_listbox.selection_set(0)
        self._on_group_selected()

    def selected_group_name(self) -> str | None:
        selection = self.group_listbox.curselection()
        if not selection:
            return None
        return self.group_listbox.get(selection[0])

    def create_group(self) -> None:
        name = simpledialog.askstring("New Employee Group", "Group name:", parent=self)
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in self.employee_groups:
            messagebox.showerror("Employee Groups", "A group with that name already exists.")
            return
        self.employee_groups[name] = []
        self._notify_change()
        self.set_data(self.employee_names, self.employee_groups)

    def delete_group(self) -> None:
        group_name = self.selected_group_name()
        if not group_name:
            return
        if not messagebox.askyesno("Delete Group", f"Delete group '{group_name}'?"):
            return
        del self.employee_groups[group_name]
        self._notify_change()
        self.set_data(self.employee_names, self.employee_groups)

    def save_group_members(self) -> None:
        group_name = self.selected_group_name()
        if not group_name:
            return
        selected_indices = self.employee_listbox.curselection()
        self.employee_groups[group_name] = [self.employee_names[index] for index in selected_indices]
        self._notify_change()
        self.summary_label.configure(
            text=f"Saved {len(self.employee_groups[group_name])} employee(s) in group '{group_name}'."
        )

    def _on_group_selected(self, _event=None) -> None:
        for index in range(self.employee_listbox.size()):
            self.employee_listbox.selection_clear(index)
        group_name = self.selected_group_name()
        if not group_name:
            return
        members = set(self.employee_groups.get(group_name, []))
        for index, employee in enumerate(self.employee_names):
            if employee in members:
                self.employee_listbox.selection_set(index)
        self.summary_label.configure(
            text=f"Group '{group_name}' contains {len(members)} employee(s)."
        )

    def _notify_change(self) -> None:
        self.on_groups_changed({name: list(members) for name, members in self.employee_groups.items()})

    def _on_employee_double_click(self, _event=None) -> None:
        selected_indices = self.employee_listbox.curselection()
        if not selected_indices:
            return
        employee = self.employee_names[selected_indices[0]]
        self.on_request_edit_email(employee)


class EmployeeNotesEditor(ttk.Frame):
    def __init__(self, master, on_notes_changed: Callable[[str, str], None]):
        super().__init__(master, padding=12)
        self.on_notes_changed = on_notes_changed
        self.names: list[str] = []
        self.notes_map: dict[str, str] = {}

        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        list_col = ttk.Frame(self)
        list_col.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        list_col.rowconfigure(0, weight=1)
        list_col.columnconfigure(0, weight=1)
        self.listbox = tk.Listbox(list_col, exportselection=False)
        notes_lb_scroll = ttk.Scrollbar(list_col, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=notes_lb_scroll.set)
        self.listbox.grid(row=0, column=0, sticky="nsew")
        notes_lb_scroll.grid(row=0, column=1, sticky="ns")
        self.listbox.bind("<<ListboxSelect>>", self._on_selection_changed)
        bind_vertical_mousewheel(self.listbox, self.listbox, units_per_notch=4)

        editor = ttk.Frame(self)
        editor.grid(row=0, column=1, sticky="nsew")
        editor.columnconfigure(0, weight=1)
        editor.rowconfigure(1, weight=1)

        self.selected_name_var = tk.StringVar(value="")
        ttk.Label(editor, textvariable=self.selected_name_var).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))
        self.note_text = tk.Text(editor, wrap="word", height=10)
        note_text_scroll = ttk.Scrollbar(editor, orient="vertical", command=self.note_text.yview)
        self.note_text.configure(yscrollcommand=note_text_scroll.set)
        self.note_text.grid(row=1, column=0, sticky="nsew")
        note_text_scroll.grid(row=1, column=1, sticky="ns")
        bind_vertical_mousewheel(self.note_text, self.note_text, units_per_notch=4)
        ttk.Button(editor, text="Save Note", command=self.save_note).grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 0))

    def set_data(self, names: list[str], notes_map: dict[str, str]) -> None:
        current = self.selected_employee()
        self.names = list(names)
        self.notes_map = dict(notes_map)
        self.listbox.delete(0, tk.END)
        for name in self.names:
            note = self.notes_map.get(name, "").strip()
            label = f"{name} *" if note else name
            self.listbox.insert(tk.END, label)
        if current and current in self.names:
            self.listbox.selection_set(self.names.index(current))
        elif self.names:
            self.listbox.selection_set(0)
        self._on_selection_changed()

    def selected_employee(self) -> str | None:
        selected = self.listbox.curselection()
        if not selected:
            return None
        return self.names[selected[0]]

    def _on_selection_changed(self, _event=None) -> None:
        employee = self.selected_employee()
        self.selected_name_var.set(employee or "")
        self.note_text.delete("1.0", tk.END)
        if employee:
            self.note_text.insert("1.0", self.notes_map.get(employee, ""))

    def save_note(self) -> None:
        employee = self.selected_employee()
        if not employee:
            return
        note = self.note_text.get("1.0", tk.END).strip()
        self.notes_map[employee] = note
        self.on_notes_changed(employee, note)
        self.set_data(self.names, self.notes_map)


class DssToolsApp(tk.Tk):
    def __init__(self, initial_source: Path | Iterable[Path] | None = None):
        super().__init__()
        self.title(DISPLAY_APP_NAME)
        apply_tk_window_icon(self)
        self.geometry("1200x760")
        self.minsize(760, 520)
        self.app_root, self.cache_dir = ensure_app_directories()
        self.updates_dir = self.app_root / UPDATE_DIRNAME
        self.updates_dir.mkdir(parents=True, exist_ok=True)
        self.config_path = self.app_root / CONFIG_FILENAME
        self.formatting_profiles, self.current_profile_name = load_formatting_profiles(self.config_path)
        self.employee_emails = load_employee_emails(self.config_path)
        self.employee_notes = load_employee_notes(self.config_path)
        self.email_subject_template, self.email_body_template = load_email_templates(self.config_path)
        self.employee_groups = load_employee_groups(self.config_path)
        self.ignored_name_typos = load_ignored_name_typos(self.config_path)
        self.job_presets = load_job_presets(self.config_path)
        self.app_settings = load_app_settings(self.config_path)
        self.filter_button_var = tk.StringVar(value="All Employees")
        self.pf_filter_button_var = tk.StringVar(value="All PFs")
        self._employee_filter_state: dict[str, bool] = {}
        self._employee_filter_vars: dict[str, tk.BooleanVar] = {}
        self._pf_filter_state: dict[str, bool] = {}
        self._pf_filter_vars: dict[str, tk.BooleanVar] = {}
        self.filter_popup: tk.Toplevel | None = None
        self.filter_popup_content: ttk.Frame | None = None
        self.pf_filter_popup: tk.Toplevel | None = None
        self.pf_filter_popup_content: ttk.Frame | None = None
        self.current_data: TrackerData | None = None
        self._has_partial_preview = False
        self._load_request_id = 0
        self._is_loading = False
        self.progress_var = tk.DoubleVar(value=0.0)
        self._hash_monitor_token = 0
        self._hash_alerted_paths: set[Path] = set()
        self._outlook_sync_in_progress = False
        self._outlook_auto_sync_done = False
        self._update_check_in_progress = False
        self._update_download_in_progress = False
        self._auto_update_check_done = False
        self._downloaded_update_path: Path | None = None
        self.update_status_var = tk.StringVar(value=f"Installed version: {APP_VERSION}")
        self.hash_poll_interval_ms = self.app_settings.hash_poll_minutes * 60 * 1000
        self._cancel_event: threading.Event | None = None
        self._active_operation_name = ""

        self._quickload_session = False
        self._quickload_cancel_sequence: str | None = None
        self._last_reports_alert: tuple[bool, bool] = (False, False)
        self._build_layout()
        self.after(AUTO_OUTLOOK_SYNC_DELAY_MS, self._auto_sync_outlook_emails)
        self.after(AUTO_UPDATE_CHECK_DELAY_MS, self._auto_check_for_updates)
        if initial_source:
            self.load_source(initial_source)
        else:
            self.after(800, self._maybe_quickload_last_sources)
        self._register_quickload_cancel_hotkey()

    def _tooltip_colours(self) -> tuple[str, str]:
        theme = self.app_settings.ui_theme
        return theme.tooltip_background, theme.tooltip_foreground

    def _build_layout(self) -> None:
        theme = self.app_settings.ui_theme
        configure_dss_table_treeview_style(self, theme)

        content_bg = theme.content_chrome_background
        self._shell_frame = tk.Frame(self, bg=content_bg, highlightthickness=0)
        self._shell_frame.pack(fill="both", expand=True)

        self._main_chrome_frame = tk.Frame(self._shell_frame, bg=content_bg, highlightthickness=0)
        self._main_chrome_frame.pack(fill="both", expand=True, padx=12, pady=12)

        top = ttk.Frame(self._main_chrome_frame)
        top.pack(fill="x", pady=(0, 12))

        self.add_dss_button = ttk.Button(top, text="Add DSS Workbook(s)", command=self.add_sources)
        self.add_dss_button.pack(side="left")
        self.remove_button = ttk.Button(top, text="Remove DSS(s)", command=self.remove_sources, state="disabled")
        self.remove_button.pack(side="left", padx=(8, 0))
        self.reload_button = ttk.Button(top, text="Update View", command=self.reload_source, state="disabled")
        self.reload_button.pack(side="left", padx=(8, 0))
        self.export_button = ttk.Button(top, text="Export Current View", command=self.export_current_view)
        self.export_button.pack(side="left", padx=(8, 0))
        ttk.Label(top, text="Filter").pack(side="left", padx=(12, 4))
        self.filter_button = ttk.Button(top, textvariable=self.filter_button_var, width=28, command=self._toggle_filter_popup)
        self.filter_button.pack(side="left")
        ttk.Label(top, text="PF").pack(side="left", padx=(8, 4))
        self.pf_filter_button = ttk.Button(
            top,
            textvariable=self.pf_filter_button_var,
            width=18,
            command=self._toggle_pf_filter_popup,
            state="disabled",
        )
        self.pf_filter_button.pack(side="left")
        self.source_label = ttk.Label(top, text="No workbook loaded", anchor="w")
        self.source_label.pack(side="left", fill="x", expand=True, padx=(12, 0))
        self.loading_label = ttk.Label(top, text="", anchor="e")
        self.loading_label.pack(side="right")

        stats = ttk.Frame(self._main_chrome_frame)
        stats.pack(fill="x", pady=(0, 12))
        self.stats_label = ttk.Label(
            stats,
            text="Load a DSS workbook to view daily and weekly labour summaries.",
            wraplength=520,
            justify="left",
        )
        self.stats_label.pack(side="left", fill="x", expand=True)

        progress_cluster = ttk.Frame(stats)
        progress_cluster.pack(side="right")
        self.quickload_hint_label = ttk.Label(progress_cluster, text="", wraplength=220, justify="right")
        self.quickload_hint_label.pack(side="left", padx=(0, 8))
        self.progress_bar = ttk.Progressbar(
            progress_cluster, variable=self.progress_var, maximum=100, mode="determinate", length=240
        )
        self.progress_bar.pack(side="left")
        self.cancel_button = ttk.Button(progress_cluster, text="Cancel", command=self._cancel_current_action, state="disabled")
        self.cancel_button.pack(side="left", padx=(8, 0))

        self.group_notebook = ttk.Notebook(self._main_chrome_frame)
        self.group_notebook.pack(fill="both", expand=True)

        self.data_group = ttk.Frame(self.group_notebook, padding=6)
        self.data_group.columnconfigure(0, weight=1)
        self.data_group.rowconfigure(0, weight=1)
        self.data_notebook = ttk.Notebook(self.data_group)
        self.data_notebook.grid(row=0, column=0, sticky="nsew")

        self.summaries_group = ttk.Frame(self.group_notebook, padding=6)
        self.summaries_group.columnconfigure(0, weight=1)
        self.summaries_group.rowconfigure(0, weight=1)
        self.summaries_notebook = ttk.Notebook(self.summaries_group)
        self.summaries_notebook.grid(row=0, column=0, sticky="nsew")

        self.reports_group = ttk.Frame(self.group_notebook, padding=6)
        self.reports_group.columnconfigure(0, weight=1)
        self.reports_group.rowconfigure(0, weight=1)
        self.reports_outline = tk.Frame(self.reports_group, highlightthickness=0, borderwidth=0)
        self.reports_outline.grid(row=0, column=0, sticky="nsew")
        self.reports_outline.columnconfigure(0, weight=1)
        self.reports_outline.rowconfigure(0, weight=1)
        self.reports_notebook = ttk.Notebook(self.reports_outline)
        self.reports_notebook.grid(row=0, column=0, sticky="nsew")

        self.settings_group = ttk.Frame(self.group_notebook, padding=6)
        self.settings_group.columnconfigure(0, weight=1)
        self.settings_group.rowconfigure(0, weight=1)
        self.settings_notebook = ttk.Notebook(self.settings_group)
        self.settings_notebook.grid(row=0, column=0, sticky="nsew")

        self.daily_table = DataTable(
            self.data_notebook,
            columns=["source_file", "date", "sheet", "employee", "st", "ot", "dt", "total", "expanded", "ranges"],
            headings=["Source File", "Date", "Source Sheet", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours", "Source Ranges Used"],
            table_id="daily_raw",
            config_path=self.config_path,
            default_sort_column="sheet",
            default_sort_descending=True,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self._employee_scroll_page = VerticalScrollablePage(self.settings_notebook)
        self.employee_editor = EmployeeListEditor(
            self._employee_scroll_page.inner,
            on_email_changed=self._update_employee_email,
            on_request_edit_email=self._prompt_edit_employee_email,
        )
        self.employee_editor.pack(fill="both", expand=True)
        self._employee_scroll_page.wire_mousewheel_to_canvas()
        self.weekly_rollup_table = DataTable(
            self.summaries_notebook,
            columns=["source_file", "week_start", "week_end", "employee", "st", "ot", "dt", "total", "expanded", "row_type"],
            headings=["Source File", "Week Start", "Week End", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours", "Row Type"],
            table_id="weekly_rollup",
            config_path=self.config_path,
            default_sort_column="week_start",
            default_sort_descending=True,
            custom_sort_key=weekly_rollup_sort_key,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self.daily_by_pf_table = DataTable(
            self.summaries_notebook,
            columns=["source_file", "work_date", "employee", "st", "ot", "dt", "total", "expanded", "row_type"],
            headings=["Source File", "Date", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours", "Row Type"],
            table_id="daily_by_pf",
            config_path=self.config_path,
            default_sort_column="work_date",
            default_sort_descending=True,
            custom_sort_key=daily_rollup_sort_key,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self.combined_weekly_summary_table = DataTable(
            self.summaries_notebook,
            columns=["week_start", "week_end", "employee", "st", "ot", "dt", "total", "expanded"],
            headings=["Week Start", "Week End", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours"],
            table_id="combined_summary",
            config_path=self.config_path,
            default_sort_column="week_start",
            default_sort_descending=True,
            ui_theme=self.app_settings.ui_theme,
        )
        self.combined_daily_summary_table = DataTable(
            self.summaries_notebook,
            columns=["work_date", "employee", "st", "ot", "dt", "total", "expanded"],
            headings=["Date", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours"],
            table_id="combined_summary_by_day",
            config_path=self.config_path,
            default_sort_column="work_date",
            default_sort_descending=True,
            ui_theme=self.app_settings.ui_theme,
        )
        self.week_totals_table = DataTable(
            self.data_notebook,
            columns=["week_start", "week_end", "st", "ot", "dt", "total", "expanded"],
            headings=["Week Start", "Week End", "Whole Crew ST", "Whole Crew OT", "Whole Crew DT", "Whole Crew Total", "Expanded Hours"],
            table_id="week_totals",
            config_path=self.config_path,
            default_sort_column="week_start",
            default_sort_descending=True,
            ui_theme=self.app_settings.ui_theme,
        )
        self.error_report_page = ttk.Frame(self.reports_notebook)
        self.error_report_page.columnconfigure(0, weight=1)
        self.error_report_page.rowconfigure(1, weight=1)
        error_toolbar = ttk.Frame(self.error_report_page, padding=(0, 0, 0, 6))
        error_toolbar.grid(row=0, column=0, sticky="ew")
        ttk.Button(error_toolbar, text="Check Name Typos", command=self._check_name_typos_manually).pack(side="left")
        self.error_report_table = DataTable(
            self.error_report_page,
            columns=[
                "employee",
                "week_start",
                "week_end",
                "rule",
                "trigger_date",
                "actual",
                "limit",
                "delta",
                "day_st",
                "day_ot",
                "day_dt",
                "sources",
                "reason",
                "breakdown",
            ],
            headings=[
                "Employee",
                "Week Start",
                "Week End",
                "Rule",
                "Trigger Date",
                "Actual",
                "Limit",
                "Delta",
                "Trigger Day ST",
                "Trigger Day OT",
                "Trigger Day DT",
                "Source Files",
                "Reason",
                "Daily Breakdown",
            ],
            table_id="error_report",
            config_path=self.config_path,
            default_sort_column="week_start",
            default_sort_descending=True,
            ui_theme=self.app_settings.ui_theme,
        )
        self.parse_warnings_table = DataTable(
            self.reports_notebook,
            columns=["source_file", "sheet", "date", "issue", "details"],
            headings=["Source File", "Sheet", "Date", "Issue", "Details"],
            table_id="parse_warnings",
            config_path=self.config_path,
            default_sort_column="sheet",
            default_sort_descending=True,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self.workbook_health_table = DataTable(
            self.reports_notebook,
            columns=["source_file", "status", "details"],
            headings=["Source File", "Status", "Details"],
            table_id="workbook_health",
            config_path=self.config_path,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self.audit_data_trail_table = DataTable(
            self.reports_notebook,
            columns=["source_file", "date", "sheet", "employee", "st", "ot", "dt", "total", "expanded", "source_ranges", "audit"],
            headings=["Source File", "Date", "Sheet", "Employee", "ST", "OT", "DT", "Total", "Expanded Hours", "Source Ranges", "Audit"],
            table_id="audit_data_trail",
            config_path=self.config_path,
            default_sort_column="sheet",
            default_sort_descending=True,
            open_source_file_callback=self._open_displayed_source_file,
            source_file_column="source_file",
            ui_theme=self.app_settings.ui_theme,
        )
        self._email_scroll_page = VerticalScrollablePage(self.reports_notebook)
        self.email_drafts_frame = EmailDraftsFrame(
            self._email_scroll_page.inner,
            create_drafts_callback=self.create_email_drafts,
            sync_emails_callback=self.sync_outlook_emails,
            on_request_edit_email=self._prompt_edit_employee_email,
            save_templates_callback=self._save_email_templates,
            ui_theme=self.app_settings.ui_theme,
        )
        self.email_drafts_frame.pack(fill="both", expand=True)
        self.email_drafts_frame.week_combo.bind("<<ComboboxSelected>>", self._on_email_week_changed)
        self.email_drafts_frame.set_templates(self.email_subject_template, self.email_body_template)
        self._email_scroll_page.wire_mousewheel_to_canvas()

        self._groups_scroll_page = VerticalScrollablePage(self.settings_notebook)
        self.groups_frame = EmployeeGroupsFrame(
            self._groups_scroll_page.inner,
            on_groups_changed=self._update_employee_groups,
            on_request_edit_email=self._prompt_edit_employee_email,
            sync_emails_callback=self.sync_outlook_emails,
        )
        self.groups_frame.pack(fill="both", expand=True)
        self._groups_scroll_page.wire_mousewheel_to_canvas()

        self._notes_scroll_page = VerticalScrollablePage(self.settings_notebook)
        self.notes_frame = EmployeeNotesEditor(self._notes_scroll_page.inner, on_notes_changed=self._update_employee_note)
        self.notes_frame.pack(fill="both", expand=True)
        self._notes_scroll_page.wire_mousewheel_to_canvas()

        self._rules_scroll_page = VerticalScrollablePage(self.settings_notebook)
        self.rules_frame = ttk.Frame(self._rules_scroll_page.inner, padding=12)
        self.rules_frame.pack(fill="both", expand=True)
        self._build_rules_tab()
        self._rules_scroll_page.wire_mousewheel_to_canvas()

        self._config_scroll_page = VerticalScrollablePage(self.settings_notebook)
        self.config_frame = ttk.Frame(self._config_scroll_page.inner, padding=12)
        self.config_frame.pack(fill="both", expand=True)
        self._build_config_tab()
        self._config_scroll_page.wire_mousewheel_to_canvas()

        self.group_notebook.add(self.data_group, text="Data")
        self.group_notebook.add(self.summaries_group, text="Summaries")
        self.group_notebook.add(self.reports_group, text="Reports")
        self.group_notebook.add(self.settings_group, text="Settings")
        self._reports_group_tab_index = self.group_notebook.index(self.reports_group)

        self._refresh_data_tabs()
        self.summaries_notebook.add(self.daily_by_pf_table, text="Daily by PF#")
        self.summaries_notebook.add(self.weekly_rollup_table, text="Weekly by PF#")
        self.summaries_notebook.add(self.combined_daily_summary_table, text="Combined Summary by Day")
        self.summaries_notebook.add(self.combined_weekly_summary_table, text="Combined Summary by Week")
        self.error_report_table.grid(row=1, column=0, sticky="nsew")
        self.reports_notebook.add(self.error_report_page, text="Error Report")
        self.reports_notebook.add(self.parse_warnings_table, text="Sheet Parse Warnings")
        self.reports_notebook.add(self.workbook_health_table, text="Workbook Health")
        self.reports_notebook.add(self.audit_data_trail_table, text="Audit Data Trail")
        self.reports_notebook.add(self._email_scroll_page, text="Email Drafts")
        self.settings_notebook.add(self._config_scroll_page, text="Configuration")
        self.settings_notebook.add(self._employee_scroll_page, text="Employee List")
        self.settings_notebook.add(self._notes_scroll_page, text="Employee Notes")
        self.settings_notebook.add(self._groups_scroll_page, text="Employee Groups")
        self.settings_notebook.add(self._rules_scroll_page, text="Formatting Rules")
        self._refresh_filter_options()
        self._apply_ui_theme_to_all_tables()
        self._sync_reports_alert_chrome(
            has_errors=self._last_reports_alert[0],
            has_parse_warnings=self._last_reports_alert[1],
        )

    def _build_rules_tab(self) -> None:
        self.rules_frame.columnconfigure(1, weight=1)

        self.profile_var = tk.StringVar(value=self.current_profile_name)
        self.job_preset_var = tk.StringVar()
        self.st_threshold_var = tk.StringVar()
        self.ot_threshold_var = tk.StringVar()
        self.daily_st_threshold_var = tk.StringVar()
        self.max_hours_per_day_var = tk.StringVar()

        ttk.Label(self.rules_frame, text="Job Preset").grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.job_preset_combo = ttk.Combobox(
            self.rules_frame,
            textvariable=self.job_preset_var,
            values=self._job_preset_names(),
        )
        self.job_preset_combo.grid(row=0, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(self.rules_frame, text="Apply Job Preset", command=self._apply_job_preset).grid(row=0, column=2, padx=(8, 0), pady=(0, 8))
        ttk.Button(self.rules_frame, text="Save Job Preset", command=self._save_job_preset).grid(row=0, column=3, padx=(8, 0), pady=(0, 8))
        ttk.Button(self.rules_frame, text="Delete Job Preset", command=self._delete_job_preset).grid(row=0, column=4, padx=(8, 0), pady=(0, 8))

        ttk.Label(self.rules_frame, text="Profile").grid(row=1, column=0, sticky="w", pady=(0, 8))
        self.profile_combo = ttk.Combobox(
            self.rules_frame,
            textvariable=self.profile_var,
            state="readonly",
            values=self._profile_names(),
        )
        self.profile_combo.grid(row=1, column=1, sticky="ew", pady=(0, 8))
        self.profile_combo.bind("<<ComboboxSelected>>", self._on_profile_selected)

        ttk.Button(self.rules_frame, text="New Profile", command=self._create_profile).grid(row=1, column=2, padx=(8, 0), pady=(0, 8))
        ttk.Button(self.rules_frame, text="Delete Profile", command=self._delete_profile).grid(row=1, column=3, padx=(8, 0), pady=(0, 8))

        daily_st_label = ttk.Label(self.rules_frame, text="Daily ST Alert")
        daily_st_label.grid(row=2, column=0, sticky="w", pady=(0, 8))
        daily_st_entry = ttk.Entry(self.rules_frame, textvariable=self.daily_st_threshold_var)
        daily_st_entry.grid(row=2, column=1, sticky="ew", pady=(0, 8))
        daily_st_help = "How many regular time hours per week an employee can work (Usually 8 or 10)."
        ToolTip(daily_st_label, daily_st_help, color_getter=self._tooltip_colours)
        ToolTip(daily_st_entry, daily_st_help, color_getter=self._tooltip_colours)

        weekly_st_label = ttk.Label(self.rules_frame, text="Weekly ST Alert")
        weekly_st_label.grid(row=3, column=0, sticky="w", pady=(0, 8))
        weekly_st_entry = ttk.Entry(self.rules_frame, textvariable=self.st_threshold_var)
        weekly_st_entry.grid(row=3, column=1, sticky="ew", pady=(0, 8))
        weekly_st_help = "How many regular time hours per week an employee can work."
        ToolTip(weekly_st_label, weekly_st_help, color_getter=self._tooltip_colours)
        ToolTip(weekly_st_entry, weekly_st_help, color_getter=self._tooltip_colours)

        weekly_ot_label = ttk.Label(self.rules_frame, text="Weekly OT Alert")
        weekly_ot_label.grid(row=4, column=0, sticky="w", pady=(0, 8))
        weekly_ot_entry = ttk.Entry(self.rules_frame, textvariable=self.ot_threshold_var)
        weekly_ot_entry.grid(row=4, column=1, sticky="ew", pady=(0, 8))
        weekly_ot_help = "Some sites have limits to how much OT you can work before you automatically start making DT."
        ToolTip(weekly_ot_label, weekly_ot_help, color_getter=self._tooltip_colours)
        ToolTip(weekly_ot_entry, weekly_ot_help, color_getter=self._tooltip_colours)

        max_hours_label = ttk.Label(self.rules_frame, text="Max Hours Per Day")
        max_hours_label.grid(row=5, column=0, sticky="w", pady=(0, 8))
        max_hours_entry = ttk.Entry(self.rules_frame, textvariable=self.max_hours_per_day_var)
        max_hours_entry.grid(row=5, column=1, sticky="ew", pady=(0, 8))
        max_hours_help = "Max hours before fatigue management per day"
        ToolTip(max_hours_label, max_hours_help, color_getter=self._tooltip_colours)
        ToolTip(max_hours_entry, max_hours_help, color_getter=self._tooltip_colours)

        ttk.Button(self.rules_frame, text="Apply Rules", command=self._apply_rule_changes).grid(row=6, column=1, sticky="w", pady=(8, 0))

        note = (
            "Enter weekly alert thresholds for this profile. Leave a field blank to disable that alert. "
            "These rules apply to employee rows in the Summaries tables (weekly and daily, per PF and combined)."
        )
        ttk.Label(self.rules_frame, text=note, wraplength=700, justify="left").grid(
            row=7, column=0, columnspan=5, sticky="w", pady=(12, 0)
        )

        self._populate_rule_editor()

    def _build_config_tab(self) -> None:
        self.config_frame.columnconfigure(1, weight=1)

        self.disable_name_typos_var = tk.BooleanVar(value=self.app_settings.disable_name_typo_notifications)
        self.show_daily_raw_var = tk.BooleanVar(value=self.app_settings.show_daily_raw_tab)
        self.quickload_last_sources_var = tk.BooleanVar(value=self.app_settings.quickload_last_sources_enabled)
        self.quickload_cancel_hotkey_var = tk.StringVar(value=self.app_settings.quickload_cancel_hotkey)
        self.hash_poll_minutes_var = tk.StringVar(value=str(self.app_settings.hash_poll_minutes))
        self.auto_update_check_var = tk.BooleanVar(value=self.app_settings.auto_update_check_enabled)
        self.auto_update_download_var = tk.BooleanVar(value=self.app_settings.auto_download_updates_on_unmetered_wifi)

        ttk.Checkbutton(
            self.config_frame,
            text="Disable name typo notifications",
            variable=self.disable_name_typos_var,
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Label(self.config_frame, text="Check source DSS(s) frequency (minutes)").grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(self.config_frame, textvariable=self.hash_poll_minutes_var, width=12).grid(
            row=1, column=1, sticky="w", pady=(0, 8)
        )

        ttk.Checkbutton(
            self.config_frame,
            text="Show Daily Raw tab",
            variable=self.show_daily_raw_var,
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Checkbutton(
            self.config_frame,
            text="Quick load last opened DSS workbook(s) on startup",
            variable=self.quickload_last_sources_var,
        ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Label(self.config_frame, text="Cancel quick-load hotkey (Tk sequence)").grid(
            row=4, column=0, sticky="nw", pady=(0, 8)
        )
        hotkey_row = ttk.Frame(self.config_frame)
        hotkey_row.grid(row=4, column=1, sticky="ew", pady=(0, 8))
        hotkey_row.columnconfigure(0, weight=1)
        self.quickload_hotkey_combo = ttk.Combobox(
            hotkey_row,
            textvariable=self.quickload_cancel_hotkey_var,
            values=list(QUICKLOAD_CANCEL_HOTKEY_PRESETS),
            width=22,
        )
        self.quickload_hotkey_combo.grid(row=0, column=0, sticky="ew")
        ttk.Button(hotkey_row, text="Press keys…", command=self._open_quickload_hotkey_capture).grid(
            row=0, column=1, sticky="e", padx=(8, 0)
        )

        ttk.Checkbutton(
            self.config_frame,
            text="Automatically check GitHub for updates on startup",
            variable=self.auto_update_check_var,
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Checkbutton(
            self.config_frame,
            text="Automatically download updates on unmetered Wi-Fi",
            variable=self.auto_update_download_var,
        ).grid(row=6, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Label(self.config_frame, text=f"Application version: {APP_VERSION}").grid(
            row=7, column=0, columnspan=2, sticky="w", pady=(0, 8)
        )

        appearance = ttk.LabelFrame(self.config_frame, text="Appearance (colours)", padding=8)
        appearance.grid(row=8, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        appearance.columnconfigure(1, weight=1)
        self._ui_theme_colour_vars: dict[str, tk.StringVar] = {}
        ut = self.app_settings.ui_theme
        for row_idx, (label, attr) in enumerate(UI_THEME_CONFIG_FIELDS):
            var = tk.StringVar(value=getattr(ut, attr))
            self._ui_theme_colour_vars[attr] = var
            ttk.Label(appearance, text=label).grid(row=row_idx, column=0, sticky="w", pady=2)
            entry_pick = ttk.Frame(appearance)
            entry_pick.grid(row=row_idx, column=1, sticky="w", padx=(8, 0), pady=2)
            ttk.Entry(entry_pick, textvariable=var, width=14).pack(side="left")
            ttk.Button(entry_pick, text="Pick…", command=lambda v=var: self._pick_ui_colour(v), width=8).pack(
                side="left", padx=(6, 0)
            )
        ttk.Button(appearance, text="Reset colours to sample defaults", command=self._reset_ui_colour_vars_to_defaults).grid(
            row=len(UI_THEME_CONFIG_FIELDS), column=0, columnspan=2, sticky="w", pady=(10, 0)
        )
        ttk.Label(
            appearance,
            text="Use #RRGGBB hex (optional short form #RGB). Applies to table row highlights, crew totals, tooltips, "
            "main window chrome, table cell backgrounds, and alert tint on report tabs.",
            wraplength=680,
            justify="left",
        ).grid(row=len(UI_THEME_CONFIG_FIELDS) + 1, column=0, columnspan=2, sticky="w", pady=(8, 0))

        ttk.Button(self.config_frame, text="Apply Settings", command=self._apply_app_settings).grid(
            row=9, column=0, sticky="w", pady=(12, 0)
        )

        maintenance = ttk.LabelFrame(self.config_frame, text="Maintenance", padding=8)
        maintenance.grid(row=10, column=0, columnspan=2, sticky="ew", pady=(16, 0))
        ttk.Button(maintenance, text="Reset All Settings to Default", command=self._reset_all_settings).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(maintenance, text="Clear Cached DSSs", command=self._clear_cached_dsss).grid(
            row=1, column=0, sticky="w", pady=(8, 0)
        )
        ttk.Button(maintenance, text="Clear Stored Emails", command=self._clear_stored_emails).grid(
            row=1, column=1, sticky="w", padx=(12, 0), pady=(8, 0)
        )
        ttk.Button(maintenance, text="Clear All Stored Data", command=self._clear_all_stored_data).grid(
            row=2, column=0, sticky="w", pady=(8, 0)
        )

        diagnostics = ttk.LabelFrame(self.config_frame, text="Diagnostics", padding=8)
        diagnostics.grid(row=11, column=0, columnspan=2, sticky="ew", pady=(16, 0))
        ttk.Button(diagnostics, text="Show App Data Folder", command=self._show_app_data_folder).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(diagnostics, text="Export Diagnostic Snapshot", command=self._export_diagnostic_snapshot).grid(
            row=0, column=1, sticky="w", padx=(12, 0)
        )
        ttk.Button(diagnostics, text="Test Outlook Connection", command=self._test_outlook_connection).grid(
            row=1, column=0, sticky="w", pady=(8, 0)
        )
        ttk.Button(diagnostics, text="Show Loaded DSS Status", command=self._show_loaded_dss_status).grid(
            row=1, column=1, sticky="w", padx=(12, 0), pady=(8, 0)
        )
        ttk.Button(diagnostics, text="Submit Bug Report", command=self._submit_bug_report).grid(
            row=2, column=0, sticky="w", pady=(8, 0)
        )
        ttk.Button(diagnostics, text="Check for Updates", command=self._check_for_updates).grid(
            row=2, column=1, sticky="w", padx=(12, 0), pady=(8, 0)
        )
        ttk.Button(diagnostics, text="Sync Outlook Emails", command=lambda: self.sync_outlook_emails(manual=True)).grid(
            row=3, column=0, sticky="w", pady=(8, 0)
        )
        ttk.Button(diagnostics, text="Check Name Typos", command=self._check_name_typos_manually).grid(
            row=3, column=1, sticky="w", padx=(12, 0), pady=(8, 0)
        )
        ttk.Label(diagnostics, textvariable=self.update_status_var, wraplength=700, justify="left").grid(
            row=4, column=0, columnspan=2, sticky="w", pady=(8, 0)
        )

        note = (
            "These settings control background notifications, how often the app checks loaded DSS files for changes, "
            "whether Daily Raw is visible, quick re-open of the last DSS set, the cancel hotkey for that load, "
            "how the app checks GitHub for downloadable updates, and optional UI colours (tables, chrome, alerts)."
        )
        ttk.Label(self.config_frame, text=note, wraplength=700, justify="left").grid(
            row=12, column=0, columnspan=2, sticky="w", pady=(12, 0)
        )

    def _pick_ui_colour(self, var: tk.StringVar) -> None:
        initial = var.get().strip()
        if not initial.startswith("#"):
            initial = "#ffffff"
        picked = colorchooser.askcolor(color=initial, title="Choose colour", parent=self)
        if picked and picked[1]:
            var.set(str(picked[1]).lower())

    def _reset_ui_colour_vars_to_defaults(self) -> None:
        defaults = DEFAULT_UI_THEME
        for _label, attr in UI_THEME_CONFIG_FIELDS:
            self._ui_theme_colour_vars[attr].set(getattr(defaults, attr))

    def _coerce_ui_theme_from_config_vars(self) -> tuple[UiThemeColors | None, str]:
        kwargs: dict[str, str] = {}
        for label, attr in UI_THEME_CONFIG_FIELDS:
            var = self._ui_theme_colour_vars.get(attr)
            if var is None:
                return None, "Appearance fields are not initialised."
            normalized = normalize_ui_hex_color(var.get().strip())
            if normalized is None:
                return None, f'Invalid colour for "{label}". Use #RRGGBB (digits 0-9 and letters a-f), for example #e0f2f1.'
            kwargs[attr] = normalized
        return UiThemeColors(**kwargs), ""

    def _profile_names(self) -> list[str]:
        return sorted(self.formatting_profiles)

    def _job_preset_names(self) -> list[str]:
        return sorted(self.job_presets)

    def _active_profile(self) -> FormattingProfile:
        return self.formatting_profiles[self.current_profile_name]

    def _populate_rule_editor(self) -> None:
        profile = self._active_profile()
        self.profile_var.set(profile.name)
        self.profile_combo.configure(values=self._profile_names())
        self.job_preset_combo.configure(values=self._job_preset_names())
        self.st_threshold_var.set("" if profile.st_threshold is None else fmt_hours(profile.st_threshold))
        self.ot_threshold_var.set("" if profile.ot_threshold is None else fmt_hours(profile.ot_threshold))
        self.daily_st_threshold_var.set("" if profile.daily_st_threshold is None else fmt_hours(profile.daily_st_threshold))
        self.max_hours_per_day_var.set("" if profile.max_hours_per_day is None else fmt_hours(profile.max_hours_per_day))

    def _on_profile_selected(self, _event=None) -> None:
        selected = self.profile_var.get()
        if selected in self.formatting_profiles:
            self.current_profile_name = selected
            self._populate_rule_editor()
            self._persist_profiles()
            self._refresh_alert_rendering()

    def _apply_rule_changes(self) -> None:
        try:
            profile = FormattingProfile(
                name=self.current_profile_name,
                st_threshold=parse_threshold_value(self.st_threshold_var.get()),
                ot_threshold=parse_threshold_value(self.ot_threshold_var.get()),
                daily_st_threshold=parse_threshold_value(self.daily_st_threshold_var.get()),
                max_hours_per_day=parse_threshold_value(self.max_hours_per_day_var.get()),
            )
        except ValueError:
            messagebox.showerror("Formatting Rules", "Thresholds must be numbers or blank.")
            return

        self.formatting_profiles[profile.name] = profile
        self._persist_profiles()
        self._refresh_alert_rendering()
        messagebox.showinfo("Formatting Rules", f"Saved rules for profile '{profile.name}'.")

    def _create_profile(self) -> None:
        name = simpledialog.askstring("New Profile", "Profile name:", parent=self)
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in self.formatting_profiles:
            messagebox.showerror("Formatting Rules", "A profile with that name already exists.")
            return

        base = self._active_profile()
        self.formatting_profiles[name] = FormattingProfile(
            name=name,
            st_threshold=base.st_threshold,
            ot_threshold=base.ot_threshold,
            daily_st_threshold=base.daily_st_threshold,
            max_hours_per_day=base.max_hours_per_day,
        )
        self.current_profile_name = name
        self._persist_profiles()
        self._populate_rule_editor()

    def _delete_profile(self) -> None:
        if len(self.formatting_profiles) == 1:
            messagebox.showerror("Formatting Rules", "At least one profile must remain.")
            return
        profile_name = self.current_profile_name
        if not messagebox.askyesno("Delete Profile", f"Delete profile '{profile_name}'?"):
            return
        del self.formatting_profiles[profile_name]
        self.current_profile_name = self._profile_names()[0]
        self._persist_profiles()
        self._populate_rule_editor()
        self._refresh_alert_rendering()

    def _persist_profiles(self) -> None:
        save_formatting_profiles(self.config_path, self.formatting_profiles, self.current_profile_name)

    def _persist_employee_emails(self) -> None:
        save_employee_emails(self.config_path, self.employee_emails)

    def _persist_employee_groups(self) -> None:
        save_employee_groups(self.config_path, self.employee_groups)

    def _persist_employee_notes(self) -> None:
        save_employee_notes(self.config_path, self.employee_notes)

    def _persist_app_settings(self) -> None:
        save_app_settings(self.config_path, self.app_settings)

    def _persist_job_presets(self) -> None:
        save_job_presets(self.config_path, self.job_presets)

    def _all_layout_tables(self) -> list[DataTable]:
        return [
            self.daily_table,
            self.weekly_rollup_table,
            self.daily_by_pf_table,
            self.combined_weekly_summary_table,
            self.combined_daily_summary_table,
            self.week_totals_table,
            self.error_report_table,
            self.parse_warnings_table,
            self.workbook_health_table,
            self.audit_data_trail_table,
        ]

    def _apply_ui_theme_to_all_tables(self) -> None:
        theme = self.app_settings.ui_theme
        configure_dss_table_treeview_style(self, theme)
        for table in self._all_layout_tables():
            table.apply_ui_theme(theme)
        self.email_drafts_frame.apply_ui_theme(theme)
        self._apply_main_chrome_theme()

    def _apply_main_chrome_theme(self) -> None:
        theme = self.app_settings.ui_theme
        content_bg = theme.content_chrome_background
        try:
            self.configure(bg=content_bg)
        except tk.TclError:
            pass
        shell = getattr(self, "_shell_frame", None)
        if isinstance(shell, tk.Frame):
            shell.configure(bg=content_bg)
        main_chrome = getattr(self, "_main_chrome_frame", None)
        if isinstance(main_chrome, tk.Frame):
            main_chrome.configure(bg=content_bg)
        self._refresh_quickload_hint_label()

    def _reload_defaults_into_ui(self) -> None:
        self.disable_name_typos_var.set(self.app_settings.disable_name_typo_notifications)
        self.show_daily_raw_var.set(self.app_settings.show_daily_raw_tab)
        self.quickload_last_sources_var.set(self.app_settings.quickload_last_sources_enabled)
        self.quickload_cancel_hotkey_var.set(self.app_settings.quickload_cancel_hotkey)
        self.hash_poll_minutes_var.set(str(self.app_settings.hash_poll_minutes))
        self.auto_update_check_var.set(self.app_settings.auto_update_check_enabled)
        self.auto_update_download_var.set(self.app_settings.auto_download_updates_on_unmetered_wifi)
        ut = self.app_settings.ui_theme
        if getattr(self, "_ui_theme_colour_vars", None):
            for _label, attr in UI_THEME_CONFIG_FIELDS:
                if attr in self._ui_theme_colour_vars:
                    self._ui_theme_colour_vars[attr].set(getattr(ut, attr))
        self.email_drafts_frame.set_templates(self.email_subject_template, self.email_body_template)
        self._populate_rule_editor()
        self._refresh_data_tabs()
        self._refresh_filter_options()
        self.groups_frame.set_data(
            self.current_data.employee_names if self.current_data is not None else [],
            self.employee_groups,
            self.employee_emails,
        )
        employee_names = self.current_data.employee_names if self.current_data is not None else []
        self.employee_editor.set_names(employee_names, self.employee_emails)
        self.notes_frame.set_data(employee_names, self.employee_notes)
        self.job_preset_combo.configure(values=self._job_preset_names())
        for table in self._all_layout_tables():
            table.reset_layout()
        if self.current_data is not None:
            self._render_data(self.current_data)
        else:
            self._last_reports_alert = (False, False)
        self._apply_ui_theme_to_all_tables()
        self._sync_reports_alert_chrome(
            has_errors=self._last_reports_alert[0],
            has_parse_warnings=self._last_reports_alert[1],
        )

    def _save_email_templates(self) -> None:
        self.email_subject_template = self.email_drafts_frame.get_subject_template()
        self.email_body_template = self.email_drafts_frame.get_body_template()
        save_email_templates(self.config_path, self.email_subject_template, self.email_body_template)
        messagebox.showinfo("Email Templates", "Saved email templates.")

    def _apply_app_settings(self) -> None:
        try:
            hash_poll_minutes = max(1, int(self.hash_poll_minutes_var.get().strip()))
        except ValueError:
            messagebox.showerror("Configuration", "Hash poll frequency must be a whole number of minutes.")
            return

        hotkey_raw = self.quickload_cancel_hotkey_var.get().strip()
        hotkey_norm = normalize_quickload_cancel_hotkey(hotkey_raw)
        if not is_allowed_quickload_cancel_hotkey(hotkey_norm):
            messagebox.showerror(
                "Configuration",
                "That cancel hotkey is not allowed. Use Escape, F-keys, Shift+F-keys, or Control+letter "
                "(see presets), then click Apply again.",
            )
            return

        ui_theme, ui_err = self._coerce_ui_theme_from_config_vars()
        if ui_theme is None:
            messagebox.showerror("Configuration", ui_err or "Invalid appearance colours.")
            return

        self.app_settings = AppSettings(
            disable_name_typo_notifications=bool(self.disable_name_typos_var.get()),
            hash_poll_minutes=hash_poll_minutes,
            show_daily_raw_tab=bool(self.show_daily_raw_var.get()),
            quickload_last_sources_enabled=bool(self.quickload_last_sources_var.get()),
            quickload_cancel_hotkey=hotkey_norm,
            auto_update_check_enabled=bool(self.auto_update_check_var.get()),
            auto_download_updates_on_unmetered_wifi=bool(self.auto_update_download_var.get()),
            ui_theme=ui_theme,
        )
        self.hash_poll_interval_ms = self.app_settings.hash_poll_minutes * 60 * 1000
        self._persist_app_settings()
        self._register_quickload_cancel_hotkey()
        self._refresh_data_tabs()
        self._schedule_hash_monitor()
        self._apply_ui_theme_to_all_tables()
        self._sync_reports_alert_chrome(
            has_errors=self._last_reports_alert[0],
            has_parse_warnings=self._last_reports_alert[1],
        )
        messagebox.showinfo("Configuration", "Saved application settings.")

    def _reset_all_settings(self) -> None:
        if not messagebox.askyesno(
            "Reset Settings",
            "Reset saved settings, formatting rules, email templates, and table layouts to defaults?\n\n"
            "This does not delete stored employee emails, employee groups, or cached DSS data.",
        ):
            return

        remove_config_keys(
            self.config_path,
            [
                "app_settings",
                "profiles",
                "current_profile",
                "email_subject_template",
                "email_body_template",
                "table_layouts",
                "job_presets",
            ],
        )
        self.app_settings = AppSettings()
        self.hash_poll_interval_ms = self.app_settings.hash_poll_minutes * 60 * 1000
        self.formatting_profiles, self.current_profile_name = load_formatting_profiles(self.config_path)
        self.email_subject_template, self.email_body_template = load_email_templates(self.config_path)
        self.job_presets = {}
        self._reload_defaults_into_ui()
        self._register_quickload_cancel_hotkey()
        self._schedule_hash_monitor()
        messagebox.showinfo("Configuration", "Settings were reset to defaults.")

    def _clear_cached_dsss(self) -> None:
        if not messagebox.askyesno(
            "Clear Cached DSSs",
            "Delete all cached parsed DSS files on disk and clear in-memory reuse flags for the "
            "currently loaded workbooks?\n\n"
            "The on-screen data stays as-is until you run Update View, which will re-read and re-parse as needed.",
        ):
            return
        deleted = clear_cache_files(self.cache_dir)
        if self.current_data is not None:
            self.current_data = tracker_data_invalidated_for_cache_clear(self.current_data)
            self._refresh_stats_summary()
        messagebox.showinfo("Configuration", f"Deleted {deleted} cached DSS file(s).")

    def _clear_stored_emails(self) -> None:
        if not messagebox.askyesno("Clear Stored Emails", "Delete all saved employee email addresses?"):
            return
        self.employee_emails = {}
        remove_config_keys(self.config_path, ["employee_emails"])
        if self.current_data is not None:
            self._render_data(self.current_data)
        else:
            self.employee_editor.set_names([], self.employee_emails)
            self.groups_frame.set_data([], self.employee_groups, self.employee_emails)
        messagebox.showinfo("Configuration", "Stored employee emails were cleared.")

    def _clear_all_stored_data(self) -> None:
        if not messagebox.askyesno(
            "Clear All Stored Data",
            "Delete all stored tracker data?\n\n"
            "This clears cached DSS data, employee emails, employee groups, formatting rules, "
            "templates, saved layouts, and app settings.",
        ):
            return

        clear_cache_files(self.cache_dir)
        if self.config_path.exists():
            self.config_path.unlink()

        self.employee_emails = {}
        self.employee_groups = {}
        self.employee_notes = {}
        self.job_presets = {}
        self.app_settings = AppSettings()
        self.hash_poll_interval_ms = self.app_settings.hash_poll_minutes * 60 * 1000
        self.formatting_profiles, self.current_profile_name = load_formatting_profiles(self.config_path)
        self.email_subject_template, self.email_body_template = load_email_templates(self.config_path)
        self._reload_defaults_into_ui()
        self._register_quickload_cancel_hotkey()
        self._schedule_hash_monitor()
        messagebox.showinfo("Configuration", "All stored tracker data was cleared.")

    def _show_app_data_folder(self) -> None:
        try:
            if os.name == "nt":
                os.startfile(self.app_root)  # type: ignore[attr-defined]
            else:
                messagebox.showinfo("App Data Folder", str(self.app_root))
        except Exception as exc:
            messagebox.showerror("App Data Folder", f"Could not open the app data folder.\n\n{exc}")

    def _export_diagnostic_snapshot(self) -> None:
        try:
            export_path = self._write_diagnostic_snapshot()
        except OSError as exc:
            messagebox.showerror("Diagnostic Snapshot", f"Could not write the diagnostic snapshot.\n\n{exc}")
            return
        messagebox.showinfo("Diagnostic Snapshot", f"Saved diagnostic snapshot:\n{export_path}")

    def _write_diagnostic_snapshot(self) -> Path:
        snapshot = {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "app_name": DISPLAY_APP_NAME,
            "app_root": str(self.app_root),
            "cache_dir": str(self.cache_dir),
            "config_path": str(self.config_path),
            "python_version": sys.version,
            "app_version": APP_VERSION,
            "os_name": os.name,
            "app_settings": {
                "disable_name_typo_notifications": self.app_settings.disable_name_typo_notifications,
                "hash_poll_minutes": self.app_settings.hash_poll_minutes,
                "show_daily_raw_tab": self.app_settings.show_daily_raw_tab,
                "quickload_last_sources_enabled": self.app_settings.quickload_last_sources_enabled,
                "quickload_cancel_hotkey": self.app_settings.quickload_cancel_hotkey,
                "auto_update_check_enabled": self.app_settings.auto_update_check_enabled,
                "auto_download_updates_on_unmetered_wifi": self.app_settings.auto_download_updates_on_unmetered_wifi,
                "ui_theme": asdict(self.app_settings.ui_theme),
            },
            "formatting_profiles": sorted(self.formatting_profiles),
            "current_profile_name": self.current_profile_name,
            "employee_email_count": len([email for email in self.employee_emails.values() if email.strip()]),
            "employee_group_count": len(self.employee_groups),
            "cache_file_count": len(list(self.cache_dir.glob("*.json"))),
            "update_download_dir": str(self.updates_dir),
            "downloaded_update_path": str(self._downloaded_update_path) if self._downloaded_update_path else "",
            "loaded_dss": [],
        }

        if self.current_data is not None:
            for source_path in self.current_data.source_paths:
                snapshot["loaded_dss"].append(
                    {
                        "path": str(source_path),
                        "hash": self.current_data.file_hashes.get(source_path, ""),
                        "reused": source_path in self.current_data.reused_paths,
                        "reloaded": source_path in self.current_data.reloaded_paths,
                        "hash_alerted": source_path in self._hash_alerted_paths,
                    }
                )
            snapshot["current_data"] = {
                "daily_record_count": len(self.current_data.daily_records),
                "employee_count": len(self.current_data.employee_names),
                "week_count": len(self.current_data.week_totals),
                "source_count": len(self.current_data.source_paths),
            }
        else:
            snapshot["current_data"] = None

        export_path = self.app_root / f"diagnostic_snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        export_path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
        return export_path

    def _test_outlook_connection(self) -> None:
        if pythoncom is None or win32com is None:
            messagebox.showerror("Outlook Connection", "Outlook integration requires pywin32 and desktop Outlook.")
            return

        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            current_user = getattr(namespace, "CurrentUser", None)
            user_name = str(getattr(current_user, "Name", "")).strip() or "Unknown user"
        except Exception as exc:
            messagebox.showerror("Outlook Connection", f"Could not connect to Outlook.\n\n{exc}")
            return
        finally:
            pythoncom.CoUninitialize()

        messagebox.showinfo("Outlook Connection", f"Connected to desktop Outlook successfully.\nUser: {user_name}")

    def _show_loaded_dss_status(self) -> None:
        if self.current_data is None or not self.current_data.source_paths:
            messagebox.showinfo("Loaded DSS Status", "No DSS workbooks are currently loaded.")
            return

        lines = [
            f"Loaded DSS files: {len(self.current_data.source_paths)}",
            f"Daily records: {len(self.current_data.daily_records)}",
            f"Employees: {len(self.current_data.employee_names)}",
            f"Weeks: {len(self.current_data.week_totals)}",
            "",
        ]
        for source_path in self.current_data.source_paths:
            status_bits = []
            if source_path in self.current_data.reloaded_paths:
                status_bits.append("reloaded")
            if source_path in self.current_data.reused_paths:
                status_bits.append("reused")
            if source_path in self._hash_alerted_paths:
                status_bits.append("changed since load")
            if not status_bits:
                status_bits.append("loaded")
            lines.append(str(source_path))
            lines.append(f"Hash: {self.current_data.file_hashes.get(source_path, '')}")
            lines.append(f"Cache: {self.current_data.cache_status_by_path.get(source_path, 'Unknown')}")
            lines.append(f"Status: {', '.join(status_bits)}")
            lines.append("")

        messagebox.showinfo("Loaded DSS Status", "\n".join(lines).rstrip())

    def _auto_check_for_updates(self) -> None:
        if self._auto_update_check_done or not self.app_settings.auto_update_check_enabled:
            return
        self._auto_update_check_done = True
        self._check_for_updates(manual=False)

    def _check_for_updates(self, manual: bool = True) -> None:
        if self._update_check_in_progress or self._update_download_in_progress:
            return
        self._update_check_in_progress = True
        if manual:
            self.update_status_var.set(f"Checking GitHub releases from version {APP_VERSION}...")
        else:
            self.update_status_var.set(f"Background update check from version {APP_VERSION}...")
        threading.Thread(target=self._check_for_updates_worker, args=(manual,), daemon=True).start()

    def _check_for_updates_worker(self, manual: bool) -> None:
        try:
            release_info = fetch_latest_release_info()
            latest_version = str(release_info.get("version", "")).strip()
            network_profile = get_windows_network_profile() if latest_version and is_newer_version(latest_version, APP_VERSION) else None
        except Exception as exc:
            self.after(0, lambda exc=exc, manual=manual: self._handle_update_check_error(exc, manual))
            return
        self.after(0, lambda release_info=release_info, network_profile=network_profile, manual=manual: self._handle_update_check_result(release_info, network_profile, manual))

    def _handle_update_check_error(self, exc: Exception, manual: bool) -> None:
        self._update_check_in_progress = False
        self.update_status_var.set(f"Installed version: {APP_VERSION}")
        if manual:
            messagebox.showerror("Check for Updates", f"Could not check for updates.\n\n{exc}")

    def _handle_update_check_result(self, release_info: dict[str, object], network_profile: dict[str, object] | None, manual: bool) -> None:
        self._update_check_in_progress = False
        latest_version = str(release_info.get("version", "")).strip()
        latest_tag = str(release_info.get("tag_name", "")).strip()
        html_url = str(release_info.get("html_url", "")).strip()
        published_at = str(release_info.get("published_at", "")).strip()
        asset_names = release_info.get("asset_names", [])
        assets_text = ", ".join(asset_names) if isinstance(asset_names, list) and asset_names else "No assets listed"
        if latest_version and is_newer_version(latest_version, APP_VERSION):
            self.update_status_var.set(f"Update available: {latest_tag or latest_version} (installed: {APP_VERSION})")
            installer_asset = choose_release_installer_asset(release_info)
            can_auto_download = installer_asset is not None and self.app_settings.auto_download_updates_on_unmetered_wifi and is_unmetered_wifi_profile(network_profile or {})
            if can_auto_download:
                self._start_update_download(release_info, manual=manual)
                if manual:
                    messagebox.showinfo(
                        "Check for Updates",
                        "A newer version is available\n\n"
                        f"Installed: {APP_VERSION}\n"
                        f"Latest: {latest_tag or latest_version}\n"
                        f"Published: {published_at or 'Unknown'}\n"
                        f"Assets: {assets_text}\n\n"
                        "This machine is on unmetered Wi-Fi, so the installer is being downloaded automatically.",
                    )
                return
            network_text = describe_network_profile(network_profile or {}) if network_profile is not None else "network state unavailable"
            self.update_status_var.set(f"Update available: {latest_tag or latest_version} (installed: {APP_VERSION}; {network_text})")
            if manual:
                if installer_asset is not None:
                    if messagebox.askyesno(
                        "Update Available",
                        "A newer version is available\n\n"
                        f"Installed: {APP_VERSION}\n"
                        f"Latest: {latest_tag or latest_version}\n"
                        f"Published: {published_at or 'Unknown'}\n"
                        f"Assets: {assets_text}\n"
                        f"Network: {network_text}\n\n"
                        "Automatic download is only offered on unmetered Wi‑Fi.\n\n"
                        "Download the installer to this PC now?",
                    ):
                        self._start_update_download(release_info, manual=manual)
                    else:
                        messagebox.showinfo(
                            "Check for Updates",
                            "You can install later from the release page:\n\n" + html_url,
                        )
                else:
                    messagebox.showinfo(
                        "Check for Updates",
                        "A newer version is available\n\n"
                        f"Installed: {APP_VERSION}\n"
                        f"Latest: {latest_tag or latest_version}\n"
                        f"Published: {published_at or 'Unknown'}\n"
                        f"Assets: {assets_text}\n"
                        f"Network: {network_text}\n\n"
                        "No downloadable installer asset was found on the release.\n\n"
                        f"Release page:\n{html_url}",
                    )
            return
        self.update_status_var.set(f"Installed version: {APP_VERSION} (up to date)")
        if manual:
            messagebox.showinfo(
                "Check for Updates",
                "You are up to date.\n\n"
                f"Installed: {APP_VERSION}\n"
                f"Latest release: {latest_tag or latest_version or 'Unknown'}",
            )

    def _start_update_download(self, release_info: dict[str, object], manual: bool) -> None:
        if self._update_download_in_progress:
            return
        self._update_download_in_progress = True
        latest_tag = str(release_info.get("tag_name", "")).strip() or str(release_info.get("version", "")).strip() or "update"
        self.update_status_var.set(f"Downloading update {latest_tag}...")
        threading.Thread(target=self._download_update_worker, args=(release_info, manual), daemon=True).start()

    def _download_update_worker(self, release_info: dict[str, object], manual: bool) -> None:
        try:
            installer_asset = choose_release_installer_asset(release_info)
            if installer_asset is None:
                raise RuntimeError("No installer asset was found in the latest GitHub release.")
            asset_name = str(installer_asset.get("name", "")).strip()
            download_url = str(installer_asset.get("download_url", "")).strip()
            if not asset_name or not download_url:
                raise RuntimeError("The release installer asset is missing its download URL.")
            destination = self.updates_dir / asset_name
            download_release_asset(download_url, destination)
            checksum_verified = False
            checksum_asset = choose_release_checksum_asset(release_info)
            if checksum_asset is not None:
                checksum_url = str(checksum_asset.get("download_url", "")).strip()
                if checksum_url:
                    checksum_text = download_url_bytes(checksum_url, timeout=60).decode("utf-8", errors="replace")
                    expected_checksum = checksum_for_asset_name(checksum_text, asset_name)
                    if expected_checksum and sha256_file(destination) != expected_checksum:
                        raise RuntimeError("Downloaded update failed SHA-256 verification.")
                    checksum_verified = bool(expected_checksum)
            self.after(0, lambda release_info=release_info, destination=destination, checksum_verified=checksum_verified, manual=manual: self._handle_update_download_success(release_info, destination, checksum_verified, manual))
        except Exception as exc:
            self.after(0, lambda exc=exc, manual=manual: self._handle_update_download_error(exc, manual))

    def _handle_update_download_error(self, exc: Exception, manual: bool) -> None:
        self._update_download_in_progress = False
        self.update_status_var.set(f"Update download failed for installed version {APP_VERSION}")
        if manual:
            messagebox.showerror("Update Download", f"Could not download the update.\n\n{exc}")

    def _handle_update_download_success(self, release_info: dict[str, object], destination: Path, checksum_verified: bool, manual: bool) -> None:
        self._update_download_in_progress = False
        self._downloaded_update_path = destination
        latest_tag = str(release_info.get("tag_name", "")).strip() or str(release_info.get("version", "")).strip() or destination.name
        verification_text = " and verified" if checksum_verified else ""
        self.update_status_var.set(f"Downloaded update {latest_tag}{verification_text}: {destination.name}")
        should_install = messagebox.askyesno(
            "Install Update",
            "The update installer has been downloaded.\n\n"
            f"Release: {latest_tag}\n"
            f"File: {destination.name}\n"
            f"Saved to: {destination}\n"
            f"Checksum verified: {'Yes' if checksum_verified else 'No checksum asset found'}\n\n"
            f"Install it now? The {DISPLAY_APP_NAME} window will close first.",
        )
        if should_install:
            self._launch_update_installer(destination)
        elif manual:
            messagebox.showinfo("Update Download", f"The installer is ready at:\n{destination}")

    def _updater_exe_for_handoff(self) -> Path | None:
        """Resolve ``DSSToolsUpdater.exe``: install dir, materialized from the frozen bundle, or None."""
        if os.name != "nt":
            return None
        sibling = Path(sys.executable).resolve().parent / UPDATER_EXE_NAME
        if sibling.is_file():
            return sibling
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            bundled = Path(sys._MEIPASS) / UPDATER_EXE_NAME
            if bundled.is_file():
                dest = self.app_root / UPDATER_EXE_NAME
                try:
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(bundled, dest)
                except OSError:
                    return None
                return dest if dest.is_file() else None
        return None

    def _resolve_updater_handoff_argv(self, installer_path: Path) -> list[str] | None:
        """Return argv to run the sidecar updater, or None if unavailable (legacy handoff)."""
        inst = installer_path.resolve()
        if os.name != "nt":
            return None
        updater_exe = self._updater_exe_for_handoff()
        if updater_exe is not None:
            return [str(updater_exe.resolve()), str(inst), str(os.getpid())]
        if not getattr(sys, "frozen", False):
            dev_script = Path(__file__).resolve().parent / "dss_tools_updater.py"
            if dev_script.is_file():
                return [sys.executable, str(dev_script), str(inst), str(os.getpid())]
        return None

    def _stage_installer_for_updater(self, installer_path: Path) -> Path:
        """Copy the setup into %TEMP% when it lives under app data so the updater can delete ``updates\\``."""
        inst = installer_path.expanduser().resolve()
        try:
            inst.relative_to(self.app_root.resolve())
        except ValueError:
            return inst
        fd, staged = tempfile.mkstemp(prefix="dss_tools_setup_", suffix=".exe", dir=str(tempfile.gettempdir()))
        os.close(fd)
        staged_path = Path(staged)
        try:
            shutil.copy2(inst, staged_path)
        except OSError:
            try:
                staged_path.unlink(missing_ok=True)
            except OSError:
                pass
            raise
        return staged_path

    def _stage_updater_exe_to_temp(self, updater_src: Path) -> Path:
        """Run the helper from %TEMP% so a silent uninstall can remove the old ``Program Files`` install."""
        fd, staged = tempfile.mkstemp(prefix="dss_tools_updater_", suffix=".exe", dir=str(tempfile.gettempdir()))
        os.close(fd)
        dest = Path(staged)
        shutil.copy2(updater_src, dest)
        return dest

    def _launch_update_installer(self, installer_path: Path) -> None:
        if not installer_path.exists():
            messagebox.showerror("Install Update", f"The downloaded installer could not be found.\n\n{installer_path}")
            return
        try:
            inst = self._stage_installer_for_updater(Path(installer_path)).resolve()
        except OSError as exc:
            messagebox.showerror("Install Update", f"Could not stage the installer for the update helper.\n\n{exc}")
            return
        argv = self._resolve_updater_handoff_argv(inst)
        if argv is not None:
            if (
                os.name == "nt"
                and len(argv) >= 1
                and Path(argv[0]).suffix.lower() == ".exe"
                and Path(argv[0]).name.casefold() == UPDATER_EXE_NAME.casefold()
            ):
                try:
                    staged_u = self._stage_updater_exe_to_temp(Path(argv[0]).resolve())
                except OSError as exc:
                    messagebox.showerror("Install Update", f"Could not stage the update helper.\n\n{exc}")
                    return
                argv = [str(staged_u), *argv[1:]]
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "DETACHED_PROCESS", 0) | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                creationflags |= getattr(subprocess, "CREATE_NO_WINDOW", 0)
            try:
                subprocess.Popen(
                    argv,
                    stdin=subprocess.DEVNULL,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    close_fds=False,
                    cwd=str(inst.parent),
                    creationflags=creationflags,
                )
            except OSError as exc:
                messagebox.showerror("Install Update", f"Could not start the update helper.\n\n{exc}")
                return
            self.destroy()
            return
        pid = os.getpid()
        escaped_path = str(inst).replace("'", "''")
        command = f"while (Get-Process -Id {pid} -ErrorAction SilentlyContinue) {{ Start-Sleep -Milliseconds 500 }}; Start-Process -FilePath '{escaped_path}'"
        log_path = self.app_root / "update_handoff.log"
        try:
            log_path.parent.mkdir(parents=True, exist_ok=True)
            with log_path.open("a", encoding="utf-8") as log_handle:
                log_handle.write(
                    f"\n---\n{datetime.now().isoformat(timespec='seconds')} legacy PowerShell handoff pid={pid}\n"
                )
        except OSError:
            pass
        creationflags = 0
        startupinfo = None
        if os.name == "nt":
            creationflags = (
                getattr(subprocess, "DETACHED_PROCESS", 0)
                | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                | getattr(subprocess, "CREATE_NO_WINDOW", 0)
            )
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = getattr(subprocess, "SW_HIDE", 0)
        try:
            subprocess.Popen(
                ["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command", command],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                stdin=subprocess.DEVNULL,
                close_fds=False,
                creationflags=creationflags,
                startupinfo=startupinfo,
            )
        except OSError as exc:
            messagebox.showerror("Install Update", f"Could not launch the installer.\n\n{exc}")
            return
        self.destroy()

    def _submit_bug_report(self) -> None:
        try:
            snapshot_path = self._write_diagnostic_snapshot()
            loaded_sources = self.current_data.source_paths if self.current_data is not None else []
            cache_status_by_path = self.current_data.cache_status_by_path if self.current_data is not None else {}
            subject = f"{DISPLAY_APP_NAME} Bug Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            html_body = build_bug_report_html(
                self.current_profile_name,
                self.app_root,
                snapshot_path,
                loaded_sources,
                cache_status_by_path,
            )
            attach_warn = create_bug_report_draft(
                BUG_REPORT_EMAIL,
                subject,
                html_body,
                attachment_path=snapshot_path,
            )
        except Exception as exc:
            messagebox.showerror("Submit Bug Report", f"Could not create the bug report draft.\n\n{exc}")
            return

        if attach_warn:
            messagebox.showwarning(
                "Submit Bug Report",
                "Created an Outlook draft bug report addressed to "
                f"{BUG_REPORT_EMAIL}.\n\n{attach_warn}\n\n"
                f"Snapshot on disk:\n{snapshot_path}",
            )
        else:
            messagebox.showinfo(
                "Submit Bug Report",
                "Created an Outlook draft bug report addressed to "
                f"{BUG_REPORT_EMAIL} and attached a diagnostic snapshot.",
            )

    def _refresh_data_tabs(self) -> None:
        current_tabs = set(self.data_notebook.tabs())
        daily_raw_tab = str(self.daily_table)
        week_totals_tab = str(self.week_totals_table)

        if self.app_settings.show_daily_raw_tab:
            if daily_raw_tab not in current_tabs:
                if current_tabs:
                    self.data_notebook.insert(0, self.daily_table, text="Daily Raw")
                else:
                    self.data_notebook.add(self.daily_table, text="Daily Raw")
        elif daily_raw_tab in current_tabs:
            self.data_notebook.forget(self.daily_table)

        current_tabs = set(self.data_notebook.tabs())
        if week_totals_tab not in current_tabs:
            self.data_notebook.add(self.week_totals_table, text="Week Totals")

    def _update_employee_email(self, employee: str, email: str) -> None:
        self.employee_emails[employee] = email.strip()
        self._persist_employee_emails()
        if self.current_data is not None:
            self._refresh_email_preview(self.current_data)

    def _prompt_edit_employee_email(self, employee: str) -> None:
        current_email = self.employee_emails.get(employee, "")
        new_email = simpledialog.askstring(
            "Edit Employee Email",
            f"Email for {employee}:",
            initialvalue=current_email,
            parent=self,
        )
        if new_email is None:
            return
        self._update_employee_email(employee, new_email.strip())
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _update_employee_groups(self, employee_groups: dict[str, list[str]]) -> None:
        self.employee_groups = {name: sorted(members) for name, members in employee_groups.items()}
        self._persist_employee_groups()
        self._refresh_filter_options()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _update_employee_note(self, employee: str, note: str) -> None:
        self.employee_notes[employee] = note.strip()
        self._persist_employee_notes()

    def _save_job_preset(self) -> None:
        job_name = self.job_preset_var.get().strip()
        if not job_name:
            messagebox.showerror("Formatting Rules", "Enter a job preset name first.")
            return
        self.job_presets[job_name] = self.current_profile_name
        self._persist_job_presets()
        self.job_preset_combo.configure(values=self._job_preset_names())
        messagebox.showinfo("Formatting Rules", f"Saved job preset '{job_name}' -> profile '{self.current_profile_name}'.")

    def _apply_job_preset(self) -> None:
        job_name = self.job_preset_var.get().strip()
        profile_name = self.job_presets.get(job_name, "")
        if not profile_name or profile_name not in self.formatting_profiles:
            messagebox.showerror("Formatting Rules", "That job preset is missing or points to a deleted profile.")
            return
        self.current_profile_name = profile_name
        self._populate_rule_editor()
        self._persist_profiles()
        self._refresh_alert_rendering()

    def _delete_job_preset(self) -> None:
        job_name = self.job_preset_var.get().strip()
        if not job_name or job_name not in self.job_presets:
            return
        if not messagebox.askyesno("Delete Job Preset", f"Delete job preset '{job_name}'?"):
            return
        del self.job_presets[job_name]
        self._persist_job_presets()
        self.job_preset_var.set("")
        self.job_preset_combo.configure(values=self._job_preset_names())

    def export_current_view(self) -> None:
        table = self._current_export_table()
        if table is None:
            messagebox.showerror("Export Current View", "The current page does not support table export.")
            return
        headings, rows = table.displayed_columns_and_rows()
        if not rows:
            messagebox.showerror("Export Current View", "There is no data to export in the current view.")
            return
        target = filedialog.asksaveasfilename(
            title="Export Current View",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Excel Workbook", "*.xlsx")],
        )
        if not target:
            return
        export_path = Path(target)
        try:
            if export_path.suffix.lower() == ".xlsx":
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Export"
                worksheet.append(headings)
                for row in rows:
                    worksheet.append(list(row))
                workbook.save(export_path)
            else:
                def csv_escape(value: str) -> str:
                    return '"' + value.replace('"', '""') + '"'

                lines = [",".join(csv_escape(value) for value in headings)]
                for row in rows:
                    lines.append(",".join(csv_escape(value) for value in row))
                export_path.write_text("\n".join(lines), encoding="utf-8")
        except Exception as exc:
            messagebox.showerror("Export Current View", f"Could not export the current view.\n\n{exc}")
            return
        messagebox.showinfo("Export Current View", f"Exported current view to:\n{export_path}")

    def _current_export_table(self) -> DataTable | None:
        current_group = self.group_notebook.select()
        if current_group == str(self.data_group):
            current_page = self.data_notebook.select()
            mapping = {
                str(self.daily_table): self.daily_table,
                str(self.week_totals_table): self.week_totals_table,
            }
            return mapping.get(current_page)
        if current_group == str(self.summaries_group):
            current_page = self.summaries_notebook.select()
            mapping = {
                str(self.weekly_rollup_table): self.weekly_rollup_table,
                str(self.daily_by_pf_table): self.daily_by_pf_table,
                str(self.combined_weekly_summary_table): self.combined_weekly_summary_table,
                str(self.combined_daily_summary_table): self.combined_daily_summary_table,
            }
            return mapping.get(current_page)
        if current_group == str(self.reports_group):
            current_page = self.reports_notebook.select()
            mapping = {
                str(self.error_report_table): self.error_report_table,
                str(self.parse_warnings_table): self.parse_warnings_table,
                str(self.workbook_health_table): self.workbook_health_table,
                str(self.audit_data_trail_table): self.audit_data_trail_table,
                str(self._email_scroll_page): self.email_drafts_frame.preview_table,
            }
            return mapping.get(current_page)
        return None

    def _refresh_alert_rendering(self) -> None:
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _refresh_outlook_sync_button(self) -> None:
        return

    def _begin_cancellable_action(self, operation_name: str, cancel_event: threading.Event) -> None:
        self._cancel_event = cancel_event
        self._active_operation_name = operation_name
        self.cancel_button.configure(state="normal")
        if not self._is_loading:
            self.loading_label.configure(text=operation_name)
            self.stats_label.configure(text=f"{operation_name} in progress...")

    def _end_cancellable_action(self, cancel_event: threading.Event | None = None) -> None:
        if cancel_event is not None and self._cancel_event is not cancel_event:
            return
        self._cancel_event = None
        self._active_operation_name = ""
        self.cancel_button.configure(state="disabled")
        if not self._is_loading:
            self.loading_label.configure(text="")
            self._refresh_stats_summary()

    def _cancel_current_action(self) -> None:
        if self._cancel_event is None:
            return
        self._cancel_event.set()
        self.loading_label.configure(text="Cancelling...")
        if self._active_operation_name:
            self.stats_label.configure(text=f"Cancelling {self._active_operation_name}...")

    def _refresh_stats_summary(self) -> None:
        if self.current_data is None:
            self.stats_label.configure(text="Load a DSS workbook to view daily and weekly labour summaries.")
            return
        memory_hits = sum(1 for status in self.current_data.cache_status_by_path.values() if status == "Memory Hit")
        disk_hits = sum(1 for status in self.current_data.cache_status_by_path.values() if status == "Disk Hit")
        misses = sum(1 for status in self.current_data.cache_status_by_path.values() if status == "Miss")
        self.stats_label.configure(
            text=(
                f"{len(self.current_data.daily_records)} daily records, "
                f"{len(self.current_data.employee_names)} employees, "
                f"{len(self.current_data.week_totals)} weeks, "
                f"{len(self.current_data.source_paths)} DSS file(s), "
                f"rules: {self.current_profile_name}, "
                f"cache: {memory_hits} memory hit / {disk_hits} disk hit / {misses} miss"
            )
        )

    def _current_filter_selection(self) -> FilterSelection:
        employee_names = self.current_data.employee_names if self.current_data is not None else []
        selected_names = tuple(
            employee for employee in employee_names
            if self._employee_filter_state.get(employee, True)
        )
        if not employee_names or len(selected_names) == len(employee_names):
            return FilterSelection(mode="all", value="All Employees")
        return FilterSelection(mode="employee_multi", value=", ".join(selected_names), values=selected_names)

    def _pf_identifier_by_source_path(self) -> dict[Path, str]:
        if self.current_data is None:
            return {}
        return {
            source_path: extract_pf_identifier(source_path.name)
            for source_path in self.current_data.source_paths
        }

    def _current_pf_selection(self) -> set[str]:
        pf_values = set(self._pf_identifier_by_source_path().values())
        selected = {
            pf_value
            for pf_value in pf_values
            if self._pf_filter_state.get(pf_value, True)
        }
        return selected if selected else pf_values

    def _refresh_filter_options(self) -> None:
        employee_names = self.current_data.employee_names if self.current_data is not None else []
        previous_state = dict(self._employee_filter_state)
        if not previous_state:
            self._employee_filter_state = {employee: True for employee in employee_names}
        else:
            self._employee_filter_state = {
                employee: previous_state.get(employee, True)
                for employee in employee_names
            }

        pf_values = sorted(set(self._pf_identifier_by_source_path().values()))
        previous_pf_state = dict(self._pf_filter_state)
        if not previous_pf_state:
            self._pf_filter_state = {pf_value: True for pf_value in pf_values}
        else:
            self._pf_filter_state = {
                pf_value: previous_pf_state.get(pf_value, True)
                for pf_value in pf_values
            }
        if len(pf_values) <= 1:
            self._pf_filter_state = {pf_value: True for pf_value in pf_values}

        self._employee_filter_vars = {}
        self._pf_filter_vars = {}
        self._rebuild_filter_popup_content()
        self._rebuild_pf_filter_popup_content()
        self._update_filter_button_label()
        self._update_pf_filter_button_label()
        self.pf_filter_button.configure(state="normal" if len(pf_values) > 1 else "disabled")

    def _toggle_filter_popup(self) -> None:
        if self.filter_popup is not None and self.filter_popup.winfo_exists():
            self._close_filter_popup()
        else:
            self._open_filter_popup()

    def _toggle_pf_filter_popup(self) -> None:
        if self.pf_filter_popup is not None and self.pf_filter_popup.winfo_exists():
            self._close_pf_filter_popup()
        else:
            self._open_pf_filter_popup()

    def _open_filter_popup(self) -> None:
        if self.filter_popup is not None and self.filter_popup.winfo_exists():
            return
        self.filter_popup = tk.Toplevel(self)
        self.filter_popup.withdraw()
        self.filter_popup.wm_overrideredirect(True)
        self.filter_popup.transient(self)
        self.filter_popup.bind("<FocusOut>", self._on_filter_popup_focus_out, add="+")
        self.filter_popup.bind("<Escape>", lambda _event: self._close_filter_popup(), add="+")
        self.filter_popup_content = ttk.Frame(self.filter_popup, padding=6, relief="solid", borderwidth=1)
        self.filter_popup_content.pack(fill="both", expand=True)
        self._rebuild_filter_popup_content()
        x = self.filter_button.winfo_rootx()
        y = self.filter_button.winfo_rooty() + self.filter_button.winfo_height() + 2
        self.filter_popup.wm_geometry(f"+{x}+{y}")
        self.filter_popup.deiconify()
        self.filter_popup.lift()
        self.filter_popup.focus_force()

    def _close_filter_popup(self) -> None:
        if self.filter_popup is not None and self.filter_popup.winfo_exists():
            self.filter_popup.destroy()
        self.filter_popup = None
        self.filter_popup_content = None

    def _open_pf_filter_popup(self) -> None:
        if self.pf_filter_popup is not None and self.pf_filter_popup.winfo_exists():
            return
        if self.pf_filter_button.instate(("disabled",)):
            return
        self.pf_filter_popup = tk.Toplevel(self)
        self.pf_filter_popup.withdraw()
        self.pf_filter_popup.wm_overrideredirect(True)
        self.pf_filter_popup.transient(self)
        self.pf_filter_popup.bind("<FocusOut>", self._on_pf_filter_popup_focus_out, add="+")
        self.pf_filter_popup.bind("<Escape>", lambda _event: self._close_pf_filter_popup(), add="+")
        self.pf_filter_popup_content = ttk.Frame(self.pf_filter_popup, padding=6, relief="solid", borderwidth=1)
        self.pf_filter_popup_content.pack(fill="both", expand=True)
        self._rebuild_pf_filter_popup_content()
        x = self.pf_filter_button.winfo_rootx()
        y = self.pf_filter_button.winfo_rooty() + self.pf_filter_button.winfo_height() + 2
        self.pf_filter_popup.wm_geometry(f"+{x}+{y}")
        self.pf_filter_popup.deiconify()
        self.pf_filter_popup.lift()
        self.pf_filter_popup.focus_force()

    def _close_pf_filter_popup(self) -> None:
        if self.pf_filter_popup is not None and self.pf_filter_popup.winfo_exists():
            self.pf_filter_popup.destroy()
        self.pf_filter_popup = None
        self.pf_filter_popup_content = None

    def _rebuild_filter_popup_content(self) -> None:
        if self.filter_popup_content is None or not self.filter_popup_content.winfo_exists():
            return
        for child in self.filter_popup_content.winfo_children():
            child.destroy()

        ttk.Button(self.filter_popup_content, text="All Employees", command=self._select_all_employees).pack(
            fill="x", anchor="w"
        )
        ttk.Button(self.filter_popup_content, text="Uncheck All", command=self._clear_employee_selection).pack(
            fill="x", anchor="w", pady=(4, 0)
        )

        employee_names = self.current_data.employee_names if self.current_data is not None else []
        if employee_names:
            ttk.Separator(self.filter_popup_content, orient="horizontal").pack(fill="x", pady=6)

        checklist = ttk.Frame(self.filter_popup_content)
        checklist.pack(fill="both", expand=True)
        for employee in employee_names:
            var = tk.BooleanVar(value=self._employee_filter_state.get(employee, True))
            self._employee_filter_vars[employee] = var
            ttk.Checkbutton(
                checklist,
                text=employee,
                variable=var,
                command=self._on_employee_filter_menu_changed,
            ).pack(anchor="w", fill="x")

    def _rebuild_pf_filter_popup_content(self) -> None:
        if self.pf_filter_popup_content is None or not self.pf_filter_popup_content.winfo_exists():
            return
        for child in self.pf_filter_popup_content.winfo_children():
            child.destroy()

        ttk.Button(self.pf_filter_popup_content, text="All PFs", command=self._select_all_pfs).pack(
            fill="x", anchor="w"
        )
        if not self.pf_filter_button.instate(("disabled",)):
            ttk.Button(self.pf_filter_popup_content, text="Uncheck All", command=self._clear_pf_selection).pack(
                fill="x", anchor="w", pady=(4, 0)
            )

        pf_values = sorted(self._pf_filter_state)
        if pf_values:
            ttk.Separator(self.pf_filter_popup_content, orient="horizontal").pack(fill="x", pady=6)

        checklist = ttk.Frame(self.pf_filter_popup_content)
        checklist.pack(fill="both", expand=True)
        disable_changes = len(pf_values) <= 1
        for pf_value in pf_values:
            var = tk.BooleanVar(value=self._pf_filter_state.get(pf_value, True))
            self._pf_filter_vars[pf_value] = var
            ttk.Checkbutton(
                checklist,
                text=pf_value,
                variable=var,
                command=self._on_pf_filter_menu_changed,
                state="disabled" if disable_changes else "normal",
            ).pack(anchor="w", fill="x")

    def _on_filter_popup_focus_out(self, _event=None) -> None:
        self.after(1, self._maybe_close_filter_popup)

    def _maybe_close_filter_popup(self) -> None:
        if self.filter_popup is None or not self.filter_popup.winfo_exists():
            return
        focused_widget = self.focus_displayof()
        if focused_widget is None:
            self._close_filter_popup()
            return
        current = focused_widget
        while current is not None:
            if current == self.filter_popup or current == self.filter_button:
                return
            current = current.master
        self._close_filter_popup()

    def _on_pf_filter_popup_focus_out(self, _event=None) -> None:
        self.after(1, self._maybe_close_pf_filter_popup)

    def _maybe_close_pf_filter_popup(self) -> None:
        if self.pf_filter_popup is None or not self.pf_filter_popup.winfo_exists():
            return
        focused_widget = self.focus_displayof()
        if focused_widget is None:
            self._close_pf_filter_popup()
            return
        current = focused_widget
        while current is not None:
            if current == self.pf_filter_popup or current == self.pf_filter_button:
                return
            current = current.master
        self._close_pf_filter_popup()

    def _select_all_employees(self) -> None:
        for employee in self._employee_filter_state:
            self._employee_filter_state[employee] = True
        for var in self._employee_filter_vars.values():
            var.set(True)
        self._update_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _clear_employee_selection(self) -> None:
        for employee in self._employee_filter_state:
            self._employee_filter_state[employee] = False
        for var in self._employee_filter_vars.values():
            var.set(False)
        self._update_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _select_all_pfs(self) -> None:
        for pf_value in self._pf_filter_state:
            self._pf_filter_state[pf_value] = True
        for var in self._pf_filter_vars.values():
            var.set(True)
        self._update_pf_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _clear_pf_selection(self) -> None:
        if len(self._pf_filter_state) <= 1:
            self._select_all_pfs()
            return
        for pf_value in self._pf_filter_state:
            self._pf_filter_state[pf_value] = False
        for var in self._pf_filter_vars.values():
            var.set(False)
        self._update_pf_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _on_employee_filter_menu_changed(self) -> None:
        for employee, var in self._employee_filter_vars.items():
            self._employee_filter_state[employee] = bool(var.get())
        self._update_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _on_pf_filter_menu_changed(self) -> None:
        if len(self._pf_filter_state) <= 1:
            self._select_all_pfs()
            return
        for pf_value, var in self._pf_filter_vars.items():
            self._pf_filter_state[pf_value] = bool(var.get())
        self._update_pf_filter_button_label()
        if self.current_data is not None:
            self._render_data(self.current_data)

    def _update_filter_button_label(self) -> None:
        employee_names = self.current_data.employee_names if self.current_data is not None else []
        if not employee_names:
            self.filter_button_var.set("All Employees")
            return
        selected = [employee for employee in employee_names if self._employee_filter_state.get(employee, True)]
        if len(selected) == len(employee_names):
            self.filter_button_var.set("All Employees")
        elif not selected:
            self.filter_button_var.set("No Employees")
        elif len(selected) == 1:
            self.filter_button_var.set(selected[0])
        else:
            self.filter_button_var.set(f"{len(selected)} Employees")

    def _update_pf_filter_button_label(self) -> None:
        pf_values = sorted(self._pf_filter_state)
        if not pf_values:
            self.pf_filter_button_var.set("All PFs")
            return
        selected = [pf_value for pf_value in pf_values if self._pf_filter_state.get(pf_value, True)]
        if len(selected) == len(pf_values):
            self.pf_filter_button_var.set("All PFs")
        elif not selected:
            self.pf_filter_button_var.set("No PFs")
        elif len(selected) == 1:
            self.pf_filter_button_var.set(selected[0])
        else:
            self.pf_filter_button_var.set(f"{len(selected)} PFs")

    def _auto_sync_outlook_emails(self) -> None:
        if self._outlook_auto_sync_done:
            return
        self._outlook_auto_sync_done = True
        if self.current_data is not None:
            self.sync_outlook_emails(manual=False)

    def sync_outlook_emails(self, manual: bool = True) -> None:
        if self.current_data is None:
            if manual:
                messagebox.showinfo("Outlook Email Sync", "Load DSS data first.")
            return
        if self._outlook_sync_in_progress or self._cancel_event is not None:
            if manual:
                messagebox.showinfo(
                    "Outlook Email Sync",
                    "Another operation is in progress (for example Outlook sync or DSS loading). Try again when it finishes.",
                )
            return

        missing_names = [
            employee
            for employee in self.current_data.employee_names
            if not self.employee_emails.get(employee, "").strip()
        ]
        if not missing_names:
            if manual:
                messagebox.showinfo("Outlook Email Sync", "All loaded employees already have email addresses saved.")
            return

        self._outlook_sync_in_progress = True
        cancel_event = threading.Event()
        self._begin_cancellable_action("Outlook email sync", cancel_event)
        self._refresh_outlook_sync_button()
        worker = threading.Thread(
            target=self._sync_outlook_emails_worker,
            args=(missing_names, manual, cancel_event),
            daemon=True,
        )
        worker.start()

    def _sync_outlook_emails_worker(self, employee_names: list[str], manual: bool, cancel_event: threading.Event) -> None:
        try:
            results = lookup_outlook_emails(employee_names, should_cancel=cancel_event.is_set)
        except OperationCancelled:
            self.after(0, lambda: self._handle_outlook_sync_cancelled(cancel_event, manual))
            return
        except Exception as exc:
            self.after(0, lambda: self._handle_outlook_sync_error(exc, manual, cancel_event))
            return
        self.after(0, lambda: self._handle_outlook_sync_success(results, employee_names, manual, cancel_event))

    def _handle_outlook_sync_cancelled(self, cancel_event: threading.Event, manual: bool) -> None:
        self._outlook_sync_in_progress = False
        self._end_cancellable_action(cancel_event)
        self._refresh_outlook_sync_button()
        if manual:
            messagebox.showinfo("Outlook Email Sync", "Outlook email sync was cancelled.")

    def _handle_outlook_sync_error(self, exc: Exception, manual: bool, cancel_event: threading.Event | None = None) -> None:
        self._outlook_sync_in_progress = False
        self._end_cancellable_action(cancel_event)
        self._refresh_outlook_sync_button()
        if manual and not isinstance(exc, OperationCancelled):
            messagebox.showerror("Outlook Email Sync", f"Could not query Outlook emails.\n\n{exc}")

    def _handle_outlook_sync_success(
        self,
        results: dict[str, str],
        employee_names: list[str],
        manual: bool,
        cancel_event: threading.Event | None = None,
    ) -> None:
        self._outlook_sync_in_progress = False
        self._end_cancellable_action(cancel_event)
        updated = 0
        for employee, email in results.items():
            if email and not self.employee_emails.get(employee, "").strip():
                self.employee_emails[employee] = email
                updated += 1
        if updated:
            self._persist_employee_emails()
            if self.current_data is not None:
                self._render_data(self.current_data)
        self._refresh_outlook_sync_button()

        typo_warnings: list[NameTypoWarning] = []
        if self.current_data is not None:
            unresolved_names = [employee for employee in employee_names if not results.get(employee, "").strip()]
            typo_warnings = [
                warning
                for warning in find_potential_name_typos(
                    unresolved_names,
                    self.current_data.employee_names,
                    self.current_data.daily_records,
                )
                if typo_warning_key(warning.employee, warning.similar_employee) not in self.ignored_name_typos
            ]

        if manual:
            missing_after = sum(1 for employee in employee_names if not self.employee_emails.get(employee, "").strip())
            messagebox.showinfo(
                "Outlook Email Sync",
                f"Matched emails: {updated}\nStill missing: {missing_after}",
            )
        if typo_warnings and not self.app_settings.disable_name_typo_notifications:
            self._show_typo_warning_dialog(typo_warnings)

    def _persist_ignored_name_typos(self) -> None:
        save_ignored_name_typos(self.config_path, self.ignored_name_typos)

    def _check_name_typos_manually(self) -> None:
        if self.current_data is None or not self.current_data.employee_names:
            messagebox.showinfo("Check Name Typos", "Load DSS data first.")
            return
        names_to_check = [
            employee for employee in self.current_data.employee_names
            if not self.employee_emails.get(employee, "").strip()
        ]
        if names_to_check:
            raw_warnings = find_potential_name_typos(
                names_to_check,
                self.current_data.employee_names,
                self.current_data.daily_records,
            )
        else:
            raw_warnings = find_similar_employee_name_pairs(
                self.current_data.employee_names,
                self.current_data.daily_records,
            )
        warnings = [
            warning
            for warning in raw_warnings
            if typo_warning_key(warning.employee, warning.similar_employee) not in self.ignored_name_typos
        ]
        if not warnings:
            messagebox.showinfo("Check Name Typos", "No likely name typos were found.")
            return
        self._show_typo_warning_dialog(warnings)

    def _show_typo_warning_dialog(self, typo_warnings: list[NameTypoWarning]) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Potential Name Typos")
        dialog.transient(self)
        dialog.geometry("780x420")

        ttk.Label(
            dialog,
            text="Possible name typo(s) were found for unresolved Outlook matches. Select any entry and choose Ignore Selected to stop warning on that inconsistency going forward. This warning can also be disabled in Settings > Configuration.",
            wraplength=740,
            justify="left",
        ).pack(fill="x", padx=12, pady=(12, 8))

        content = ttk.Frame(dialog, padding=(12, 0, 12, 12))
        content.pack(fill="both", expand=True)
        content.columnconfigure(2, weight=1)
        content.rowconfigure(0, weight=1)

        listbox = tk.Listbox(content, selectmode="extended", exportselection=False)
        listbox.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        scrollbar = ttk.Scrollbar(content, orient="vertical", command=listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        listbox.configure(yscrollcommand=scrollbar.set)
        bind_vertical_mousewheel(listbox, listbox, units_per_notch=4)

        details = tk.Text(content, wrap="word", height=12)
        details.grid(row=0, column=2, sticky="nsew")
        bind_vertical_mousewheel(details, details, units_per_notch=4)

        def refresh_details(_event=None) -> None:
            selection = listbox.curselection()
            details.configure(state="normal")
            details.delete("1.0", tk.END)
            if selection:
                warning = typo_warnings[selection[0]]
                lines = [
                    f"{warning.employee} may match {warning.similar_employee}",
                    f"Similarity: {warning.similarity * 100:.0f}%",
                    "",
                    "Locations:",
                    *(warning.locations or ["(No locations found)"]),
                ]
                details.insert("1.0", "\n".join(lines))
            details.configure(state="disabled")

        for warning in typo_warnings:
            listbox.insert(
                tk.END,
                f"{warning.employee} -> {warning.similar_employee} ({warning.similarity * 100:.0f}% similar)",
            )
        if typo_warnings:
            listbox.selection_set(0)
        listbox.bind("<<ListboxSelect>>", refresh_details)
        refresh_details()

        buttons = ttk.Frame(dialog, padding=(12, 0, 12, 12))
        buttons.pack(fill="x")

        def ignore_selected() -> None:
            selected = listbox.curselection()
            if not selected:
                return
            for index in selected:
                warning = typo_warnings[index]
                self.ignored_name_typos.add(typo_warning_key(warning.employee, warning.similar_employee))
            self._persist_ignored_name_typos()
            dialog.destroy()

        ttk.Button(buttons, text="Ignore Selected", command=ignore_selected).pack(side="left")
        ttk.Button(buttons, text="Close", command=dialog.destroy).pack(side="right")

    def add_sources(self) -> None:
        if self._cancel_event is not None:
            return
        selected = filedialog.askopenfilenames(
            title="Add DSS workbook(s)",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not selected:
            return
        existing = list(self.current_data.source_paths) if self.current_data is not None else []
        combined = existing + [Path(path) for path in selected]
        self.load_source(combined)

    def remove_sources(self) -> None:
        if self.current_data is None or self._cancel_event is not None:
            return
        source_paths = list(self.current_data.source_paths)
        if not source_paths:
            return

        dialog = tk.Toplevel(self)
        dialog.title("Remove DSS Workbook(s)")
        dialog.transient(self)
        dialog.geometry("720x320")

        ttk.Label(dialog, text="Select the loaded DSS workbook(s) to remove.").pack(anchor="w", padx=12, pady=(12, 8))

        list_frame = ttk.Frame(dialog, padding=(12, 0, 12, 0))
        list_frame.pack(fill="both", expand=True)
        listbox = tk.Listbox(list_frame, selectmode="extended", exportselection=False)
        listbox.pack(side="left", fill="both", expand=True)
        bind_vertical_mousewheel(listbox, listbox, units_per_notch=4)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="left", fill="y")

        for source_path in source_paths:
            listbox.insert(tk.END, str(source_path))

        buttons = ttk.Frame(dialog, padding=12)
        buttons.pack(fill="x")

        def remove_selected() -> None:
            selected_indices = listbox.curselection()
            if not selected_indices:
                dialog.destroy()
                return
            keep_paths = [path for index, path in enumerate(source_paths) if index not in selected_indices]
            dialog.destroy()
            if keep_paths:
                self.load_source(keep_paths)
            else:
                self.current_data = None
                self._hash_alerted_paths.clear()
                self.source_label.configure(text="No workbook loaded")
                self._employee_filter_state = {}
                self._refresh_filter_options()
                self._refresh_stats_summary()
                self._clear_all_views()
                self._set_loading_state(False)

        ttk.Button(buttons, text="Remove Selected", command=remove_selected).pack(side="right")
        ttk.Button(buttons, text="Cancel", command=dialog.destroy).pack(side="right", padx=(0, 8))

    def reload_source(self) -> None:
        if not self.current_data or self._is_loading:
            return
        self.load_source(self.current_data.source_paths, show_success=True)

    def load_source(self, source_paths: Path | Iterable[Path], show_success: bool = False) -> None:
        if self._cancel_event is not None:
            return
        normalized_paths = self._normalize_source_paths(source_paths)
        if not normalized_paths:
            return

        self._load_request_id += 1
        request_id = self._load_request_id
        cancel_event = threading.Event()
        self._has_partial_preview = False
        self._begin_cancellable_action("DSS load", cancel_event)
        self._set_loading_state(True, normalized_paths)

        worker = threading.Thread(
            target=self._load_source_worker,
            args=(request_id, normalized_paths, show_success, cancel_event),
            daemon=True,
        )
        worker.start()

    @staticmethod
    def _normalize_source_paths(source_paths: Path | Iterable[Path]) -> list[Path]:
        if isinstance(source_paths, Path):
            return [source_paths.expanduser().resolve()]
        return [Path(path).expanduser().resolve() for path in source_paths]

    def _set_loading_state(self, is_loading: bool, source_paths: list[Path] | None = None) -> None:
        self._is_loading = is_loading
        self.add_dss_button.configure(state="disabled" if is_loading else "normal")
        self.remove_button.configure(state="disabled" if is_loading or self.current_data is None else "normal")
        self.reload_button.configure(
            state="disabled" if is_loading or self.current_data is None else "normal"
        )
        if is_loading:
            if source_paths and len(source_paths) == 1:
                target = source_paths[0].name
            elif source_paths:
                target = f"{len(source_paths)} DSS files"
            else:
                target = "selected DSS files"
            self.loading_label.configure(text=f"Loading {target}...")
            self.stats_label.configure(text="Reading workbook data in the background...")
            self.progress_var.set(0.0)
            self._refresh_quickload_hint_label()
        else:
            self.loading_label.configure(text="")
            self.progress_var.set(0.0)
            self._quickload_session = False
            self._refresh_quickload_hint_label()
        self._refresh_outlook_sync_button()

    def _clear_all_views(self) -> None:
        self.daily_table.set_rows([])
        self.weekly_rollup_table.set_rows([])
        self.daily_by_pf_table.set_rows([])
        self.combined_weekly_summary_table.set_rows([])
        self.combined_daily_summary_table.set_rows([])
        self.week_totals_table.set_rows([])
        self.error_report_table.set_rows([])
        self.parse_warnings_table.set_rows([])
        self.workbook_health_table.set_rows([])
        self.audit_data_trail_table.set_rows([])
        self.email_drafts_frame.preview_table.set_rows([])
        self.email_drafts_frame.summary_label.configure(text="Load DSS data to preview employee email drafts.")
        self.email_drafts_frame.set_week_options([])
        self.employee_editor.set_names([], self.employee_emails)
        self.notes_frame.set_data([], self.employee_notes)
        self.groups_frame.set_data([], self.employee_groups, self.employee_emails)
        self._sync_reports_alert_chrome(has_errors=False, has_parse_warnings=False)

    def _load_source_worker(
        self,
        request_id: int,
        source_paths: list[Path],
        show_success: bool,
        cancel_event: threading.Event,
    ) -> None:
        def report_progress(progress_fraction: float, message: str) -> None:
            self.after(0, lambda: self._update_progress(request_id, progress_fraction, message))

        def report_partial(tracker_data: TrackerData, message: str) -> None:
            self.after(0, lambda: self._handle_partial_load_update(request_id, tracker_data, message))

        try:
            tracker_data = load_tracker_data(
                source_paths,
                previous_data=self.current_data,
                progress_callback=report_progress,
                partial_callback=report_partial,
                cache_dir=self.cache_dir,
                should_cancel=cancel_event.is_set,
            )
        except OperationCancelled:
            self.after(0, lambda: self._handle_load_cancelled(request_id, cancel_event))
            return
        except Exception as exc:
            self.after(0, lambda: self._handle_load_error(request_id, exc, cancel_event))
            return
        self.after(0, lambda: self._handle_load_success(request_id, tracker_data, show_success, cancel_event))

    def _update_progress(self, request_id: int, progress_fraction: float, message: str) -> None:
        if request_id != self._load_request_id:
            return
        percentage = round(min(max(progress_fraction, 0.0), 1.0) * 100, 1)
        self.progress_var.set(percentage)
        self.loading_label.configure(text=f"{percentage:.1f}%")
        self.stats_label.configure(text=message)

    def _handle_load_cancelled(self, request_id: int, cancel_event: threading.Event | None = None) -> None:
        if request_id != self._load_request_id:
            return
        self._set_loading_state(False)
        self._end_cancellable_action(cancel_event)
        if self._has_partial_preview and self.current_data is not None:
            self.stats_label.configure(text="Load cancelled. Showing partially loaded recent DSS data.")
        else:
            self.stats_label.configure(text="Load cancelled.")

    def _handle_load_error(self, request_id: int, exc: Exception, cancel_event: threading.Event | None = None) -> None:
        if request_id != self._load_request_id:
            return
        self._set_loading_state(False)
        self._end_cancellable_action(cancel_event)
        messagebox.showerror(DISPLAY_APP_NAME, f"Failed to open workbook.\n\n{exc}")

    def _handle_partial_load_update(self, request_id: int, tracker_data: TrackerData, message: str) -> None:
        if request_id != self._load_request_id:
            return
        self._has_partial_preview = True
        self.current_data = tracker_data
        self._refresh_filter_options()
        self._refresh_outlook_sync_button()
        if len(tracker_data.source_paths) == 1:
            source_text = str(tracker_data.source_paths[0])
        else:
            source_text = f"{len(tracker_data.source_paths)} DSS workbooks loading"
        self.source_label.configure(text=source_text)
        self._render_data(tracker_data)
        self.stats_label.configure(text=message)

    def _handle_load_success(
        self,
        request_id: int,
        tracker_data: TrackerData,
        show_success: bool,
        cancel_event: threading.Event | None = None,
    ) -> None:
        if request_id != self._load_request_id:
            return

        self._has_partial_preview = False
        self.current_data = tracker_data
        self._hash_alerted_paths.clear()
        self._set_loading_state(False)
        self._end_cancellable_action(cancel_event)
        self._refresh_filter_options()
        self._refresh_outlook_sync_button()
        if len(tracker_data.source_paths) == 1:
            source_text = str(tracker_data.source_paths[0])
        else:
            source_text = f"{len(tracker_data.source_paths)} DSS workbooks loaded"
        self.source_label.configure(text=source_text)
        self._refresh_stats_summary()
        self._render_data(tracker_data)
        self._schedule_hash_monitor()
        save_last_open_dss_paths(self.config_path, tracker_data.source_paths)
        if show_success:
            summary_lines = [
                f"Reloaded: {len(tracker_data.reloaded_paths)}",
                f"Unchanged: {len(tracker_data.reused_paths)}",
            ]
            if tracker_data.reloaded_paths:
                summary_lines.append("")
                summary_lines.append("Reprocessed files:")
                summary_lines.extend(str(path) for path in tracker_data.reloaded_paths)
            messagebox.showinfo(
                DISPLAY_APP_NAME,
                "\n".join(summary_lines),
            )

    def _render_data(self, tracker_data: TrackerData) -> None:
        filter_selection = self._current_filter_selection()
        selected_pfs = self._current_pf_selection()
        filtered_daily_records = [
            record
            for record in tracker_data.daily_records
            if extract_pf_identifier(record.source_file) in selected_pfs
        ]
        filtered_employee_names = sorted({record.employee for record in filtered_daily_records})
        allowed_employees = filter_employee_names(filtered_employee_names, filter_selection, self.employee_groups)
        filtered_daily_records = [record for record in filtered_daily_records if record.employee in allowed_employees]
        filtered_weekly_rollup = [
            record for record in tracker_data.weekly_rollup
            if (
                extract_pf_identifier(record.source_file) in selected_pfs
                and (record.row_type == "Crew Total" or record.employee in allowed_employees)
            )
        ]
        filtered_daily_rollup = [
            record for record in tracker_data.daily_rollup
            if (
                extract_pf_identifier(record.source_file) in selected_pfs
                and (record.row_type == "Crew Total" or record.employee in allowed_employees)
            )
        ]
        filtered_combined_summary = [
            record for record in tracker_data.combined_weekly_summary if record.employee in allowed_employees
        ]
        filtered_combined_daily = [
            record for record in tracker_data.combined_daily_summary if record.employee in allowed_employees
        ]
        filtered_week_totals = build_week_totals(filtered_combined_summary) if filtered_combined_summary else []
        filtered_source_files = {record.source_file for record in filtered_daily_records}
        filtered_parse_warnings = [
            warning
            for warning in tracker_data.parse_warnings
            if warning.source_file in filtered_source_files or not filtered_daily_records
        ]

        self.daily_table.set_rows(
            [
                (
                    record.source_file,
                    record.work_date.isoformat(),
                    record.source_sheet,
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                    record.source_ranges,
                )
                for record in filtered_daily_records
            ]
        )

        self.employee_editor.set_names(tracker_data.employee_names, self.employee_emails)
        self.notes_frame.set_data(tracker_data.employee_names, self.employee_notes)
        self.groups_frame.set_data(tracker_data.employee_names, self.employee_groups, self.employee_emails)

        weekly_rollup_rows: list[tuple[str, ...]] = []
        weekly_rollup_tags: list[tuple[str, ...]] = []
        for record in filtered_weekly_rollup:
            weekly_rollup_rows.append(
                (
                    record.source_file,
                    record.week_start.isoformat(),
                    record.week_end.isoformat(),
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                    record.row_type,
                )
            )
            if record.row_type == "Crew Total":
                weekly_rollup_tags.append(("crew_total",))
            else:
                weekly_rollup_tags.append(self._threshold_tags(record.st, record.ot, record.dt))
        self.weekly_rollup_table.set_rows(weekly_rollup_rows, weekly_rollup_tags)

        daily_rollup_rows: list[tuple[str, ...]] = []
        daily_rollup_tags: list[tuple[str, ...]] = []
        for record in filtered_daily_rollup:
            daily_rollup_rows.append(
                (
                    record.source_file,
                    record.work_date.isoformat(),
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                    record.row_type,
                )
            )
            if record.row_type == "Crew Total":
                daily_rollup_tags.append(("crew_total",))
            else:
                daily_rollup_tags.append(self._threshold_tags(record.st, record.ot, record.dt))
        self.daily_by_pf_table.set_rows(daily_rollup_rows, daily_rollup_tags)

        combined_summary_rows: list[tuple[str, ...]] = []
        combined_summary_tags: list[tuple[str, ...]] = []
        for record in filtered_combined_summary:
            combined_summary_rows.append(
                (
                    record.week_start.isoformat(),
                    record.week_end.isoformat(),
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                )
            )
            combined_summary_tags.append(self._threshold_tags(record.st, record.ot, record.dt))
        self.combined_weekly_summary_table.set_rows(combined_summary_rows, combined_summary_tags)

        combined_daily_rows: list[tuple[str, ...]] = []
        combined_daily_tags: list[tuple[str, ...]] = []
        for record in filtered_combined_daily:
            combined_daily_rows.append(
                (
                    record.work_date.isoformat(),
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                )
            )
            combined_daily_tags.append(self._threshold_tags(record.st, record.ot, record.dt))
        self.combined_daily_summary_table.set_rows(combined_daily_rows, combined_daily_tags)

        self.week_totals_table.set_rows(
            [
                (
                    record.week_start.isoformat(),
                    record.week_end.isoformat(),
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                )
                for record in filtered_week_totals
            ]
        )
        error_findings = build_error_findings(filtered_daily_records, self._active_profile())
        self.error_report_table.set_rows(
            [
                (
                    finding.employee,
                    finding.week_start.isoformat(),
                    finding.week_end.isoformat(),
                    f"{finding.hour_type} > {fmt_hours(finding.threshold)}",
                    finding.trigger_date.isoformat(),
                    fmt_hours(finding.actual_total),
                    fmt_hours(finding.threshold),
                    fmt_hours(finding.delta),
                    fmt_hours(finding.trigger_day_st),
                    fmt_hours(finding.trigger_day_ot),
                    fmt_hours(finding.trigger_day_dt),
                    finding.source_files,
                    finding.reason,
                    finding.breakdown,
                )
                for finding in error_findings
            ],
            tags=[("alert",) for _ in error_findings],
        )
        self.parse_warnings_table.set_rows(
            [
                (
                    warning.source_file,
                    warning.source_sheet,
                    warning.work_date,
                    warning.issue,
                    warning.details,
                )
                for warning in filtered_parse_warnings
            ],
            tags=[("alert",) for _ in filtered_parse_warnings],
        )
        self.workbook_health_table.set_rows(
            [
                (
                    item.source_file,
                    item.status,
                    item.details,
                )
                for item in tracker_data.workbook_health
            ],
            tags=[("alert",) if item.status.lower() == "warning" else tuple() for item in tracker_data.workbook_health],
        )
        self.audit_data_trail_table.set_rows(
            [
                (
                    record.source_file,
                    record.work_date.isoformat(),
                    record.source_sheet,
                    record.employee,
                    fmt_hours(record.st),
                    fmt_hours(record.ot),
                    fmt_hours(record.dt),
                    fmt_hours(record.total),
                    fmt_hours(expanded_hours(record.st, record.ot, record.dt)),
                    record.source_ranges,
                    f"Derived from parsed block totals for {record.employee} on {record.source_sheet}.",
                )
                for record in filtered_daily_records
            ]
        )
        self._sync_reports_alert_chrome(
            has_errors=bool(error_findings),
            has_parse_warnings=bool(filtered_parse_warnings),
        )
        self._refresh_email_preview(tracker_data, allowed_employees)

    def _sync_reports_alert_chrome(self, *, has_errors: bool, has_parse_warnings: bool) -> None:
        er_idx = self.reports_notebook.index(self.error_report_page)
        pw_idx = self.reports_notebook.index(self.parse_warnings_table)
        base_er = "Error Report"
        base_pw = "Sheet Parse Warnings"
        parent_alert = has_errors or has_parse_warnings
        self._last_reports_alert = (has_errors, has_parse_warnings)
        theme = self.app_settings.ui_theme
        alert_hex = normalize_ui_hex_color(theme.alert_row_background) or theme.alert_row_background
        try:
            alert_er = solid_tab_swatch_photo(self, alert_hex)
            alert_pw = solid_tab_swatch_photo(self, alert_hex)
            alert_reports = solid_tab_swatch_photo(self, alert_hex)
            self._reports_tab_photo_refs = [alert_er, alert_pw, alert_reports]
            self.reports_notebook.tab(
                er_idx,
                text=f"{base_er} (!)" if has_errors else base_er,
                image=alert_er if has_errors else "",
                compound="left" if has_errors else "none",
            )
            self.reports_notebook.tab(
                pw_idx,
                text=f"{base_pw} (!)" if has_parse_warnings else base_pw,
                image=alert_pw if has_parse_warnings else "",
                compound="left" if has_parse_warnings else "none",
            )
            rg_idx = self._reports_group_tab_index
            self.group_notebook.tab(
                rg_idx,
                text="Reports (!)" if parent_alert else "Reports",
                image=alert_reports if parent_alert else "",
                compound="left" if parent_alert else "none",
            )
        except tk.TclError:
            self.reports_notebook.tab(er_idx, text=f"{base_er} (!)" if has_errors else base_er)
            self.reports_notebook.tab(pw_idx, text=f"{base_pw} (!)" if has_parse_warnings else base_pw)
            self.group_notebook.tab(self._reports_group_tab_index, text="Reports (!)" if parent_alert else "Reports")

    def _open_displayed_source_file(self, source_display_name: str) -> None:
        if not source_display_name or self.current_data is None:
            return
        target: Path | None = None
        for path in self.current_data.source_paths:
            if path.name == source_display_name:
                target = path
                break
        if target is None or not target.exists():
            messagebox.showerror("Open Source File", f"Could not find workbook:\n{source_display_name}")
            return
        try:
            if sys.platform == "win32":
                os.startfile(target)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.run(["open", str(target)], check=False)
            else:
                subprocess.run(["xdg-open", str(target)], check=False)
        except OSError as exc:
            messagebox.showerror("Open Source File", f"Could not open file.\n\n{exc}")

    def _register_quickload_cancel_hotkey(self) -> None:
        new_seq = normalize_quickload_cancel_hotkey(self.app_settings.quickload_cancel_hotkey)
        if not is_allowed_quickload_cancel_hotkey(new_seq):
            new_seq = "<Escape>"
        if self._quickload_cancel_sequence == new_seq:
            return
        if self._quickload_cancel_sequence:
            try:
                self.unbind_all(self._quickload_cancel_sequence)
            except tk.TclError:
                pass
        try:
            self.bind_all(new_seq, self._on_quickload_cancel_hotkey_event)
        except tk.TclError:
            return
        self._quickload_cancel_sequence = new_seq

    def _on_quickload_cancel_hotkey_event(self, _event: tk.Event) -> str | None:
        if self._is_loading and self._quickload_session and self._cancel_event is not None:
            self._cancel_current_action()
            return "break"
        return None

    def _refresh_quickload_hint_label(self) -> None:
        if self._is_loading and self._quickload_session:
            self.quickload_hint_label.configure(
                text="Quick load — you can turn this off under Settings → Configuration."
            )
        else:
            self.quickload_hint_label.configure(text="")

    def _open_quickload_hotkey_capture(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Set cancel hotkey")
        dialog.transient(self)
        dialog.resizable(False, False)
        ttk.Label(
            dialog,
            text=(
                "Click here, then press the key or combination you want.\n"
                "Escape, F-keys, Shift+F-keys, and Control+letter are supported.\n"
                "Use Cancel to close without changing the hotkey field."
            ),
            padding=12,
            wraplength=420,
            justify="left",
        ).pack(fill="x")
        status = ttk.Label(dialog, text="Waiting for keys…", padding=(12, 4))
        status.pack(fill="x")

        def finish(seq: str) -> None:
            self.quickload_cancel_hotkey_var.set(seq)
            status.configure(text=f"Captured: {seq}")
            dialog.destroy()

        def on_key_press(event: tk.Event) -> str | None:
            seq = binding_sequence_from_keypress_event(event)
            if seq is None:
                return None
            if not is_allowed_quickload_cancel_hotkey(seq):
                status.configure(text=f"Not allowed: {seq}")
                return "break"
            finish(seq)
            return "break"

        dialog.bind("<KeyPress>", on_key_press)
        ttk.Button(dialog, text="Cancel", command=dialog.destroy).pack(pady=(0, 12))
        dialog.grab_set()
        dialog.focus_set()
        dialog.wait_window(dialog)

    def _maybe_quickload_last_sources(self) -> None:
        if not self.app_settings.quickload_last_sources_enabled:
            return
        if self.current_data is not None:
            return
        paths = [path for path in load_last_open_dss_paths(self.config_path) if path.expanduser().resolve().exists()]
        if not paths:
            return
        self._quickload_session = True
        self.load_source(paths, show_success=False)

    def _refresh_email_preview(self, tracker_data: TrackerData, allowed_employees: set[str] | None = None) -> None:
        filtered_daily_records = [
            record for record in tracker_data.daily_records
            if allowed_employees is None or record.employee in allowed_employees
        ]
        week_options = collect_week_ranges(filtered_daily_records)
        previous_selection = self.email_drafts_frame.week_var.get()
        self.email_drafts_frame.set_week_options(week_options)
        if previous_selection in self.email_drafts_frame.week_combo.cget("values"):
            self.email_drafts_frame.week_var.set(previous_selection)

        week_start = self.email_drafts_frame.selected_week_start()
        if week_start is None:
            self.email_drafts_frame.summary_label.configure(text="Load DSS data to preview employee email drafts.")
            self.email_drafts_frame.preview_table.set_rows([])
            return

        requests = build_email_draft_requests(filtered_daily_records, self.employee_emails, week_start)
        preview_rows: list[tuple[str, ...]] = []
        for request in requests:
            st_total = round(sum(record.st for record in request.records), 2)
            ot_total = round(sum(record.ot for record in request.records), 2)
            dt_total = round(sum(record.dt for record in request.records), 2)
            preview_rows.append(
                (
                    request.employee,
                    request.email or "(missing)",
                    str(len(request.records)),
                    fmt_hours(st_total),
                    fmt_hours(ot_total),
                    fmt_hours(dt_total),
                    fmt_hours(st_total + ot_total + dt_total),
                    fmt_hours(expanded_hours(st_total, ot_total, dt_total)),
                )
            )
        self.email_drafts_frame.preview_table.set_rows(preview_rows)
        missing = sum(1 for request in requests if not request.email)
        week_end = week_start + timedelta(days=6)
        self.email_drafts_frame.summary_label.configure(
            text=(
                f"Previewing {len(requests)} employee draft(s) for {format_week_label(week_start, week_end)}. "
                f"Missing email addresses: {missing}."
            )
        )

    def create_email_drafts(self) -> None:
        if self.current_data is None:
            messagebox.showerror("Email Drafts", "Load DSS data before creating draft emails.")
            return

        week_start = self.email_drafts_frame.selected_week_start()
        if week_start is None:
            messagebox.showerror("Email Drafts", "No week is available for draft creation.")
            return

        allowed_employees = filter_employee_names(
            self.current_data.employee_names,
            self._current_filter_selection(),
            self.employee_groups,
        )
        filtered_daily_records = [
            record for record in self.current_data.daily_records if record.employee in allowed_employees
        ]
        requests = build_email_draft_requests(filtered_daily_records, self.employee_emails, week_start)
        selected_employees = self.email_drafts_frame.selected_employees()
        if selected_employees:
            requests = [request for request in requests if request.employee in selected_employees]
        if not requests:
            messagebox.showerror("Email Drafts", "No employee records were found for the selected week.")
            return

        try:
            created, skipped = create_outlook_drafts(
                requests,
                self.email_drafts_frame.get_subject_template(),
                self.email_drafts_frame.get_body_template(),
            )
        except Exception as exc:
            messagebox.showerror("Email Drafts", f"Could not create Outlook drafts.\n\n{exc}")
            return

        summary_lines = [f"Created drafts: {created}", f"Skipped (missing email): {len(skipped)}"]
        if skipped:
            summary_lines.append("")
            summary_lines.append("Missing email addresses:")
            summary_lines.extend(skipped)
        messagebox.showinfo("Email Drafts", "\n".join(summary_lines))

    def _schedule_hash_monitor(self) -> None:
        self._hash_monitor_token += 1
        token = self._hash_monitor_token
        self.after(self.hash_poll_interval_ms, lambda: self._run_hash_monitor(token))

    def _run_hash_monitor(self, token: int) -> None:
        if token != self._hash_monitor_token:
            return
        if self.current_data is None or self._is_loading or not self.current_data.source_paths:
            self._schedule_hash_monitor()
            return

        source_paths = list(self.current_data.source_paths)
        expected_hashes = dict(self.current_data.file_hashes)
        worker = threading.Thread(
            target=self._hash_monitor_worker,
            args=(token, source_paths, expected_hashes),
            daemon=True,
        )
        worker.start()

    def _hash_monitor_worker(self, token: int, source_paths: list[Path], expected_hashes: dict[Path, str]) -> None:
        changed_paths: list[Path] = []
        for source_path in source_paths:
            try:
                current_hash = compute_workbook_content_hash(read_source_bytes(source_path))
            except Exception:
                continue
            if expected_hashes.get(source_path) != current_hash:
                changed_paths.append(source_path)
        self.after(0, lambda: self._handle_hash_monitor_result(token, changed_paths))

    def _handle_hash_monitor_result(self, token: int, changed_paths: list[Path]) -> None:
        if token != self._hash_monitor_token:
            return
        unseen_changes = [path for path in changed_paths if path not in self._hash_alerted_paths]
        if unseen_changes:
            self._hash_alerted_paths.update(unseen_changes)
            changed_names = ", ".join(path.name for path in unseen_changes)
            self.loading_label.configure(text="Changed")
            self.stats_label.configure(
                text=f"DSS changed since last view: {changed_names}. Press Update View to refresh."
            )
            messagebox.showinfo(
                "DSS Changed",
                "One or more loaded DSS workbooks changed since the last inspected state.\n\n"
                "Press 'Update View' to refresh the summaries.",
            )
        self._schedule_hash_monitor()

    def _on_email_week_changed(self, _event=None) -> None:
        if self.current_data is not None:
            self._refresh_email_preview(self.current_data)

    def _threshold_tags(self, st: float, ot: float, dt: float) -> tuple[str, ...]:
        if is_alert_triggered(st, ot, dt, self._active_profile()):
            return ("alert",)
        return tuple()


def main() -> int:
    parser = argparse.ArgumentParser(description="DSS Tools — desktop DSS labour-hour summaries and workflows.")
    parser.add_argument("source", nargs="*", help="Optional DSS workbook(s) to open on launch.")
    args = parser.parse_args()

    initial_source = [Path(path).expanduser().resolve() for path in args.source] if args.source else None
    app = DssToolsApp(initial_source=initial_source)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

